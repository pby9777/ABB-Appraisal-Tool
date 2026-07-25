#!/usr/bin/env python3
"""
ABB Energy Appraisal – CEE Report Generator
Template: EA_Report_Template_CEE.docx
Regions:  France, CEE countries (Czech, Poland-CEE, Hungary, etc.)

Differences from Standard:
  * Extra "Investment – CEE" column (net investment after CEE government subsidy)
  * Header row auto-detected (default 35 vs Standard's 36)
  * Payback sensitivity from col 19 (price delta) / col 21 (payback)
  * Appendix tables borrowed from EA_Report_Template_Standard.docx when NA > 10

Usage:
  python generate_report_cee.py  <saving_calc.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]

Requirements:
  pip install openpyxl
"""

import sys, os, re, math, io, zipfile
from datetime import datetime
import openpyxl

# ── CLI ───────────────────────────────────────────────────────────────────────
args = sys.argv[1:]
if len(args) < 3:
    print('Usage: python generate_report_cee.py  <excel.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]')
    sys.exit(1)

XLSX_PATH   = args[0]
CUSTOMER    = args[1].upper()
PLANT       = args[2]
RPT_DATE    = args[3] if len(args) > 3 else datetime.now().strftime("%m.%d.%Y")
DATA_SOURCE = args[4] if len(args) > 4 else "Customer Input"

_script_dir   = os.path.dirname(os.path.abspath(__file__))
_template_dir = os.path.join(_script_dir, "report_templates")
TEMPLATE_PATH     = os.path.join(_template_dir, "ea_report_template_cee.docx")
# Standard template needed to borrow Appendix table XML when NA > 10
TEMPLATE_STD_PATH = os.path.join(_template_dir, "ea_report_template_standard.docx")

for path, label in [(XLSX_PATH, "Excel file"), (TEMPLATE_PATH, "Template docx")]:
    if not os.path.exists(path):
        print(f"ERROR: {label} not found: {path}")
        sys.exit(1)

def date_to_iso(d):
    for fmt in ('%m.%d.%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(d, fmt).strftime('%Y-%m-%d')
        except: pass
    return d

RPT_DATE_ISO = date_to_iso(RPT_DATE)

# ── Read Excel ────────────────────────────────────────────────────────────────
print(f"Reading Excel: {os.path.basename(XLSX_PATH)}")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

def _is_sc_sheet(name):
    n = name.strip().lower().replace(' ', '_')
    prefixes = ('saving_calculations', 'savings_calculations',
                'saving_calculatios',  'savings_calculatios')
    return any(n.startswith(p) for p in prefixes)

_sc_candidates = [s for s in wb.sheetnames if _is_sc_sheet(s)]
_sc_sheet = _sc_candidates[0] if _sc_candidates else None
if _sc_sheet is None:
    print(f"ERROR: No Saving_Calculations sheet found. Sheets: {wb.sheetnames}")
    sys.exit(1)
if len(_sc_candidates) > 1:
    _vn = [s for s in _sc_candidates if re.search(r'_v\d+$', s.strip().lower())]
    if _vn:
        _sc_sheet = _vn[0]
if _sc_sheet not in ("Saving_Calculations", "Savings_Calculations", "Saving Calculations"):
    print(f"  Note: using sheet '{_sc_sheet}'")
ws = wb[_sc_sheet]

_XL_ERRORS = frozenset({
    '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!', '#ERROR!',
    '-', '\u2013', '\u2014', 'N/A', 'n/a', 'NA', 'na', 'n.a.', 'N.A.',
})

def cv(row, col):
    """Read a cell value; return None for Excel error strings."""
    v = ws.cell(row=row, column=col).value
    return None if isinstance(v, str) and v.strip() in _XL_ERRORS else v

# ── Auto-detect header row (CEE default 35; Standard uses 36) ─────────────────
HEADER_ROW = 35
for _hr in range(30, 42):
    _h5 = str(ws.cell(_hr, 5).value or '').strip().lower()
    if 'annual energy' in _h5 or 'energy. cons' in _h5:
        HEADER_ROW = _hr
        break
ASSET_START_ROW = HEADER_ROW + 1

col_map = {}
for c in range(1, 70):
    h = cv(HEADER_ROW, c)
    if h:
        col_map[str(h).strip().lower()] = c

def find_col(names, default):
    for n in names:
        if n.lower() in col_map:
            return col_map[n.lower()]
    return default

# ── Column map (CEE Excel: extra Investment–CEE column after Investment gross) ─
COL_NUM        = find_col(['#', '# no.', 'no.'], 2)
COL_E_CONS     = find_col(['annual energy. cons (kwh)', 'annual energy cons (kwh)'], 5)
COL_E_COST     = find_col(['annual energy cost'], 6)
COL_CO2_CONS   = find_col(['annual co2 cons.'], 7)
COL_SAV_KWH    = find_col(['annual energy savings, kwh'], 8)
COL_INVEST     = find_col(['investment'], 9)           # gross investment
COL_CEE        = find_col(['cee'], 10)                 # CEE government subsidy
COL_INVEST_CEE = find_col(['investment - cee', 'investment-cee',
                            'investment \u2013 cee'], 11)  # net investment
COL_SAV_COST   = find_col(['annual energy cost savings'], 12)
COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 13)
COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 14)
COL_SAV_PCT    = find_col(['annual energy savings (%)'], 15)
COL_NPV        = find_col(['npv'], 16)

COL_IE       = find_col(['ie eff class', 'ie'], 40)
COL_LOAD     = find_col(['driven load', 'application'], 41)
COL_CONN     = find_col(['dol vsd', 'dol/vsd', 'connection'], 42)
COL_FLOW     = find_col(['flow control', 'flow control method'], 43)
COL_OUTPUT   = find_col(['output (kw)', 'rated power, kw'], 44)
COL_SHAFT    = find_col(['shaft height (frame)', 'shaft height'], 45)
COL_RUNHRS   = find_col(['annual running hours', 'running hours', 'running time (hours)'], 46)
COL_AVG_LOAD = find_col(['average loading'], 47)
COL_AVG_FREQ = find_col(['average freqency', 'average frequency'], 48)
COL_ESS      = find_col(['recommended ess motor'], 49)
COL_ESSC     = find_col(['ess connection'], 50)

# ── KPI values ────────────────────────────────────────────────────────────────
CURRENCY       = str(cv(3, 3) or "EUR").strip()
ELEC_PRICE     = cv(5, 4) or 0.14
CO2_INTENSITY  = cv(6, 4) or 0.055
TAX_RATE       = cv(8, 4) or 0.25
DISCOUNT_RATE  = float(cv(7, 4) or 0.065)
_NPV_YEARS     = 20
TOTAL_ASSETS   = int(cv(13, 3) or 0)
NPV_POS_CNT    = int(cv(14, 3) or 0)
ANNUAL_SAVINGS = cv(15, 3)
CONSUMP_BEFORE = cv(16, 3)
SAVINGS_KWH    = cv(17, 3)
NPV_VALUE      = cv(19, 3)
INVEST_COST    = cv(20, 3)    # Investment–CEE (net) for NPV+ summary
PAYBACK_TIME   = cv(21, 3)
CO2_SAVINGS    = cv(22, 3)
IRR_VALUE      = cv(23, 3)
BEV_COUNT      = cv(25, 4)
TOP10_PAYBACK  = cv(21, 4)
TOP10_NPV      = cv(19, 4)
TOP10_INVEST   = cv(20, 4)
TOP10_IRR      = cv(23, 4)

# Payback sensitivity (CEE Excel: col 19 = price delta, col 21 = payback)
SENSITIVITY = []
for r in range(19, 24):
    delta   = cv(r, 19)
    payback = cv(r, 21)
    try:
        SENSITIVITY.append((float(delta), float(payback)))
    except (TypeError, ValueError):
        pass
_SENSITIVITY_PENDING = len(SENSITIVITY) < 5

# ── Load assets ───────────────────────────────────────────────────────────────
assets = []
_scan_limit = max(int(TOTAL_ASSETS or 0), 200)
for r in range(ASSET_START_ROW, ASSET_START_ROW + _scan_limit):
    num = cv(r, COL_NUM)
    if num is None: break
    try: num = int(float(num))
    except: break
    if not (1 <= num <= 9999): break

    ess_motor = str(cv(r, COL_ESS) or '')

    shaft_raw = cv(r, COL_SHAFT)
    try:
        shaft_val = str(int(float(shaft_raw))) if shaft_raw is not None else ''
    except:
        shaft_val = str(shaft_raw) if shaft_raw else ''

    # Average frequency preferred; fall back to average loading
    avg_freq = cv(r, COL_AVG_FREQ)
    avg_load = cv(r, COL_AVG_LOAD)
    if avg_freq is not None and str(avg_freq).strip() not in ('', '0'):
        try:
            _af = float(avg_freq)
            avg_val = f"{round(_af * 100)}" if 0 < _af <= 1 else str(avg_freq).strip()
        except:
            avg_val = str(avg_freq).strip()
    elif avg_load is not None and str(avg_load).strip() not in ('', '0'):
        try:
            _al = float(avg_load)
            avg_val = f"{round(_al * 100)}" if 0 < _al <= 1 else str(avg_load).strip()
        except:
            avg_val = str(avg_load).strip()
    else:
        fm = re.search(r'([\d.]+)\s*Hz', ess_motor)
        avg_val = fm.group(1) if fm else ''

    conn_raw = cv(r, COL_CONN)
    conn_val = str(conn_raw) if conn_raw else 'DOL'

    invest_raw = cv(r, COL_INVEST)
    cee_raw    = cv(r, COL_CEE)
    invest_cee = cv(r, COL_INVEST_CEE)
    if invest_cee is None:
        invest_cee = float(invest_raw or 0) - float(cee_raw or 0)

    payback_raw = cv(r, COL_PAYBACK)
    assets.append({
        'num':        num,
        'tag':        str(cv(r, 3) or ''),
        'load':       str(cv(r, COL_LOAD) or cv(r, 4) or ''),
        'e_cons':     cv(r, COL_E_CONS)    or 0,
        'e_cost':     cv(r, COL_E_COST)    or 0,
        'co2_cons':   cv(r, COL_CO2_CONS)  or 0,
        'e_sav_kwh':  cv(r, COL_SAV_KWH)   or 0,
        'e_sav_cost': cv(r, COL_SAV_COST)  or 0,
        'e_sav_pct':  cv(r, COL_SAV_PCT)   or 0,
        'invest':     float(invest_raw or 0),
        'cee':        float(cee_raw or 0),
        'invest_cee': float(invest_cee or 0),
        'payback':    payback_raw if not isinstance(payback_raw, str) else None,
        'co2_sav':    cv(r, COL_CO2_SAV)   or 0,
        'npv':        cv(r, COL_NPV),
        'ie':         str(cv(r, COL_IE)    or ''),
        'flow_ctrl':  str(cv(r, COL_FLOW)  or 'Information not available'),
        'connection': conn_val,
        'output_kw':  str(cv(r, COL_OUTPUT) or ''),
        'shaft_h':    shaft_val,
        'run_hrs':    str(cv(r, COL_RUNHRS) or ''),
        'avg_val':    avg_val,
        'ess_motor':  ess_motor,
        'ess_conn':   str(cv(r, COL_ESSC)  or ''),
    })

if not assets:
    print("ERROR: No assets found. Check asset data in Excel.")
    sys.exit(1)

# ── Per-asset fallback (when Excel formulas are uncached) ─────────────────────
_needs_fallback = any(a['e_sav_cost'] == 0 and a['e_sav_kwh'] and a['npv'] is None
                      for a in assets)
if _needs_fallback:
    print("  Note: Excel formula cells uncached — computing e_cost / NPV / payback from raw inputs")

for a in assets:
    ec   = float(a['e_cons']    or 0)
    esav = float(a['e_sav_kwh'] or 0)
    inv  = float(a['invest_cee'] or 0)   # CEE uses net investment for NPV/payback
    ep   = float(ELEC_PRICE or 0.14)
    co2i = float(CO2_INTENSITY or 0.055)

    if not a['e_cost']     and ec   > 0: a['e_cost']     = ec   * ep
    if not a['e_sav_cost'] and esav > 0: a['e_sav_cost'] = esav * ep
    if not a['e_sav_pct']  and ec   > 0: a['e_sav_pct']  = esav / ec
    if not a['co2_cons']   and ec   > 0: a['co2_cons']   = round(ec   * co2i, 2)
    if not a['co2_sav']    and esav > 0: a['co2_sav']    = round(esav * co2i, 2)

    ann_sav = float(a['e_sav_cost'] or 0)
    if a['npv'] is None and ann_sav > 0 and inv > 0:
        r         = DISCOUNT_RATE
        after_tax = ann_sav * (1 - float(TAX_RATE or 0.25))
        pv_factor = ((1 - (1 + r) ** (-_NPV_YEARS)) / r) if r > 0 else _NPV_YEARS
        a['npv']  = round(after_tax * pv_factor - inv, 2)

    if a['payback'] is None and ann_sav > 0 and inv > 0:
        npv_val = float(a['npv']) if a['npv'] is not None else -1
        if npv_val > 0:
            a['payback'] = round(inv / ann_sav, 2)

NA = len(assets)
npv_pos = [a for a in assets if a['npv'] is not None and float(a['npv']) > 0]

if TOTAL_ASSETS == 0:
    TOTAL_ASSETS = NA
if NPV_POS_CNT == 0 and npv_pos:
    NPV_POS_CNT = len(npv_pos)
    print(f"  Note: NPV_POS_CNT recomputed from assets: {NPV_POS_CNT}")

# ── Sort and build top-10 (same logic as Standard) ────────────────────────────
if npv_pos:
    npv_sorted = sorted(
        npv_pos,
        key=lambda a: (float(a['payback']) if a['payback'] else 9999,
                       -(float(a['e_sav_cost'] or 0)))
    )[:10]
    if len(npv_sorted) < 10:
        npv_neg = [a for a in assets if a['npv'] is None or float(a['npv'] or 0) <= 0]
        npv_neg_sorted = sorted(npv_neg, key=lambda a: -(float(a['e_sav_cost'] or 0)))
        slots = 10 - len(npv_sorted)
        top10 = npv_sorted + npv_neg_sorted[:slots]
        print(f"  Top table: {len(npv_sorted)} NPV+ + {min(slots, len(npv_neg_sorted))} NPV- (white-shaded)")
    else:
        top10 = npv_sorted
    NO_NPV_MODE = False
else:
    top10 = sorted(assets, key=lambda a: -(float(a['e_sav_cost'] or 0)))[:10]
    NO_NPV_MODE = True
    print("  NOTE: No NPV-positive assets — using all assets sorted by energy savings")

NT = len(top10)
assets_by_num = sorted(assets, key=lambda a: a['num'])

# ── NO_NPV_MODE: re-derive summary KPIs from all assets ───────────────────────
if NO_NPV_MODE:
    ANNUAL_SAVINGS = sum(float(a['e_sav_cost'] or 0) for a in assets)
    CO2_SAVINGS    = sum(float(a['co2_sav']    or 0) for a in assets)
    CONSUMP_BEFORE = sum(float(a['e_cons']     or 0) for a in assets)
    SAVINGS_KWH    = sum(float(a['e_sav_kwh']  or 0) for a in assets)
    BEV_COUNT      = round(float(SAVINGS_KWH) / 3500) if SAVINGS_KWH else 0
    print("  NO_NPV_MODE: summary KPIs computed from all assets")

    _deltas = []
    for r in range(19, 24):
        try: _deltas.append(float(cv(r, 19)))
        except (TypeError, ValueError): pass
    if not _deltas:
        _deltas = [-0.2, -0.1, 0.0, 0.1, 0.2]
    _total_invest = sum(float(a['invest_cee'] or 0) for a in assets)
    _base_savings = float(ANNUAL_SAVINGS or 0)
    SENSITIVITY = []
    for d in _deltas:
        adj = _base_savings * (1 + d)
        if adj > 0:
            SENSITIVITY.append((d, _total_invest / adj))
    if SENSITIVITY:
        print(f"  NO_NPV_MODE sensitivity: {[f'{s[0]:+.0%}→{s[1]:.1f}yrs' for s in SENSITIVITY]}")

# ── General KPI fallback (uncached Excel formula cells) ───────────────────────
_kpi_src = assets if NO_NPV_MODE else npv_pos
if ANNUAL_SAVINGS is None or float(ANNUAL_SAVINGS or 0) == 0:
    ANNUAL_SAVINGS = sum(float(a['e_sav_cost'] or 0) for a in _kpi_src)
    print(f"  KPI fallback: ANNUAL_SAVINGS = {ANNUAL_SAVINGS:,.0f}")
if CONSUMP_BEFORE is None or float(CONSUMP_BEFORE or 0) == 0:
    CONSUMP_BEFORE = sum(float(a['e_cons'] or 0) for a in _kpi_src)
if SAVINGS_KWH is None or float(SAVINGS_KWH or 0) == 0:
    SAVINGS_KWH = sum(float(a['e_sav_kwh'] or 0) for a in _kpi_src)
if CO2_SAVINGS is None or float(CO2_SAVINGS or 0) == 0:
    CO2_SAVINGS = sum(float(a['co2_sav'] or 0) for a in _kpi_src)
if INVEST_COST is None or float(INVEST_COST or 0) == 0:
    INVEST_COST = sum(float(a['invest_cee'] or 0) for a in _kpi_src)   # net
if BEV_COUNT is None or float(BEV_COUNT or 0) == 0:
    BEV_COUNT = round(float(SAVINGS_KWH or 0) / 3500)
if PAYBACK_TIME is None or float(PAYBACK_TIME or 0) == 0:
    _sav = float(ANNUAL_SAVINGS or 0)
    PAYBACK_TIME = round(float(INVEST_COST or 0) / _sav, 2) if _sav > 0 else None
if NPV_VALUE is None or float(NPV_VALUE or 0) == 0:
    NPV_VALUE = sum(float(a['npv'] or 0) for a in _kpi_src if a['npv'] is not None)
if TOP10_NPV is None or float(TOP10_NPV or 0) == 0:
    TOP10_NPV = sum(float(a['npv'] or 0) for a in top10
                    if a['npv'] is not None and float(a['npv']) > 0)
if TOP10_INVEST is None or float(TOP10_INVEST or 0) == 0:
    TOP10_INVEST = sum(float(a['invest_cee'] or 0) for a in top10)   # net
if TOP10_PAYBACK is None or float(TOP10_PAYBACK or 0) == 0:
    _t10s = sum(float(a['e_sav_cost'] or 0) for a in top10)
    TOP10_PAYBACK = round(float(TOP10_INVEST) / _t10s, 2) if _t10s > 0 else None

# ── IRR fallback (Newton-Raphson on flat-annuity cash flow) ───────────────────
def _compute_irr(invest, annual_cf, n_years):
    if invest <= 0 or annual_cf <= 0 or n_years <= 0:
        return None
    r = annual_cf / invest
    for _ in range(200):
        try:
            if abs(r) < 1e-10: break
            pow_neg_n  = (1 + r) ** (-n_years)
            pow_neg_n1 = (1 + r) ** (-n_years - 1)
            annuity    = (1 - pow_neg_n) / r
            npv_val    = -invest + annual_cf * annuity
            d_annuity  = (n_years * pow_neg_n1 * r + pow_neg_n - 1) / (r * r)
            d_npv      = annual_cf * d_annuity
            if abs(d_npv) < 1e-12: break
            r_new = r - npv_val / d_npv
            if abs(r_new - r) < 1e-9:
                r = r_new; break
            r = r_new
        except (ValueError, ZeroDivisionError, OverflowError):
            break
    return round(r, 6) if 0 < r < 10 else None

if TOP10_IRR is None or float(TOP10_IRR or 0) == 0:
    _t10i = float(TOP10_INVEST or 0)
    _t10s = sum(float(a['e_sav_cost'] or 0) for a in top10)
    _irr  = _compute_irr(_t10i, _t10s, _NPV_YEARS)
    if _irr is not None:
        TOP10_IRR = _irr
        print(f"  IRR fallback (Top-{NT}): {_irr:.1%}")

if IRR_VALUE is None or float(IRR_VALUE or 0) == 0:
    _src_irr = assets if NO_NPV_MODE else npv_pos
    _all_s   = sum(float(a['e_sav_cost'] or 0) for a in _src_irr)
    _all_i   = float(INVEST_COST or 0)
    _irr_all = _compute_irr(_all_i, _all_s, _NPV_YEARS)
    if _irr_all is not None:
        IRR_VALUE = _irr_all

# ── Sensitivity fallback ──────────────────────────────────────────────────────
if _SENSITIVITY_PENDING and npv_pos:
    _ns = sum(float(a['e_sav_cost'] or 0) for a in npv_pos)
    _ni = sum(float(a['invest_cee'] or 0) for a in npv_pos)
    if _ns > 0:
        SENSITIVITY = [(_d, round(_ni / (_ns * (1 + _d)), 4))
                       for _d in [-0.2, -0.1, 0.0, 0.1, 0.2] if _ns * (1 + _d) > 0]
        print(f"  Sensitivity fallback: {[f'{s[0]:+.0%}→{s[1]:.2f}yrs' for s in SENSITIVITY]}")

print(f"  {NA} assets, {len(npv_pos)} NPV+, {NT} in Top-{NT} "
      f"{'(no NPV+ mode)' if NO_NPV_MODE else '(sorted by payback)'}")

# ── Top-N totals (all NT assets — mirrors Standard behaviour) ─────────────────
t10_e_cons   = sum(float(a['e_cons']     or 0) for a in top10)
t10_e_cost   = sum(float(a['e_cost']     or 0) for a in top10)
t10_co2_c    = sum(float(a['co2_cons']   or 0) for a in top10)
t10_sav_kwh  = sum(float(a['e_sav_kwh']  or 0) for a in top10)
t10_invest_g = sum(float(a['invest']     or 0) for a in top10)
t10_invest_c = sum(float(a['invest_cee'] or 0) for a in top10)   # net (CEE extra col)
t10_sav_c    = sum(float(a['e_sav_cost'] or 0) for a in top10)
t10_co2_sav  = sum(float(a['co2_sav']   or 0) for a in top10)
t10_pct      = t10_sav_kwh / t10_e_cons if t10_e_cons else 0
t10_npv      = sum(float(a['npv'] or 0) for a in top10
                   if a['npv'] is not None and float(a['npv']) > 0)

t10_simple_payback = t10_invest_c / t10_sav_c if t10_sav_c else 0
TOP10_PAYBACK = t10_simple_payback if t10_simple_payback > 0 else TOP10_PAYBACK
TOP10_NPV     = t10_npv
TOP10_INVEST  = t10_invest_c

# ── All NPV+ totals (for "Total NPV Positive Assets" row) ─────────────────────
_npv_src     = npv_pos if not NO_NPV_MODE else assets
all_e_cons   = sum(float(a['e_cons']     or 0) for a in _npv_src)
all_e_cost   = sum(float(a['e_cost']     or 0) for a in _npv_src)
all_co2_c    = sum(float(a['co2_cons']   or 0) for a in _npv_src)
all_sav_kwh  = sum(float(a['e_sav_kwh']  or 0) for a in _npv_src)
all_invest_g = sum(float(a['invest']     or 0) for a in _npv_src)
all_invest_c = sum(float(a['invest_cee'] or 0) for a in _npv_src)
all_sav_c    = sum(float(a['e_sav_cost'] or 0) for a in _npv_src)
all_co2      = sum(float(a['co2_sav']    or 0) for a in _npv_src)
all_pct          = all_sav_kwh / all_e_cons if all_e_cons else 0
all_payback      = round(all_invest_c / all_sav_c, 2) if all_sav_c else None
all_std_payback  = round(all_invest_g / all_sav_c, 2) if all_sav_c else None

# Top-N standard payback (gross investment basis, no CEE subsidy)
t10_std_payback  = round(t10_invest_g / t10_sav_c, 2) if t10_sav_c else None

# ── Formatting helpers ────────────────────────────────────────────────────────
SYM = CURRENCY

def fmt(n, dp=0):
    if n is None or (isinstance(n, float) and math.isnan(n)): return '\u2013'
    try:
        n = float(n); s = f"{abs(n):,.{dp}f}"
        return s if n >= 0 else f"-{s}"
    except: return str(n)

def fmt_parens(n):
    if n is None: return '\u2013'
    try:
        n = float(n)
        return f"({fmt(abs(n))})" if n < 0 else fmt(n)
    except: return str(n)

def fmtpct(n):
    if n is None: return '\u2013'
    try:
        v = round(float(n) * 100)
        return f"{v}%" if v >= 0 else f"-{abs(v)}%"
    except: return '\u2013'

def fmtyrs(n):
    if n is None: return ''
    try: return f"{float(n):.1f}"
    except: return ''

def fmtirr(n):
    if n is None: return '\u2013'
    try: return f"{round(float(n) * 100)}%"
    except: return str(n)

def pct_str(n):
    if n is None: return '\u2013'
    return f"{round(float(n) * 100, 1)}%"

# ── XML helpers ───────────────────────────────────────────────────────────────
def xe(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

def remove_nth_tc(row_xml, n):
    """Remove the nth <w:tc>...</w:tc> (0-indexed) from a table row."""
    tcs = [(m.start(), row_xml.find('</w:tc>', m.start()) + len('</w:tc>'))
           for m in re.finditer(r'<w:tc>', row_xml)]
    if n >= len(tcs): return row_xml
    s, e = tcs[n]
    return row_xml[:s] + row_xml[e:]

def _inject_wt_in_cell(row_xml, cell_idx):
    """Inject a bare <w:r><w:t></w:t></w:r> into the nth cell's paragraph.
    Used to make cells that have no text node fillable by fill_row()."""
    tcs = []
    pos = 0
    while True:
        s = row_xml.find('<w:tc>', pos)
        if s == -1: break
        e = row_xml.find('</w:tc>', s) + len('</w:tc>')
        tcs.append((s, e))
        pos = e
    if cell_idx >= len(tcs): return row_xml
    cs, ce = tcs[cell_idx]
    tc_xml = row_xml[cs:ce]
    pp_end = tc_xml.rfind('</w:p>')
    if pp_end == -1: return row_xml
    tc_new = tc_xml[:pp_end] + '<w:r><w:t></w:t></w:r>' + tc_xml[pp_end:]
    return row_xml[:cs] + tc_new + row_xml[ce:]

def get_tr_pos(xml):
    pos = []
    for m in re.finditer(r'<w:tr[ >]', xml):
        s = m.start()
        e = xml.find('</w:tr>', s)
        if e == -1: continue
        pos.append((s, e + len('</w:tr>')))
    return pos

def nth_wt(xml_str, n, new_val):
    ms = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', xml_str))
    if n >= len(ms): return xml_str
    m = ms[n]
    return xml_str[:m.start()] + m.group(1) + xe(new_val) + m.group(3) + xml_str[m.end():]

def strip_ids(s):
    for attr in ['w14:paraId', 'w14:textId', 'w:rsidR', 'w:rsidRPr',
                 'w:rsidRDefault', 'w:rsidTr', 'w:rsidP', 'w:rsidDel']:
        s = re.sub(f' {re.escape(attr)}="[^"]*"', '', s)
    return s

def fill_row(tmpl, vals):
    row  = strip_ids(tmpl)
    ms   = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', row))
    vals = list(vals)[:len(ms)]
    for i, v in reversed(list(enumerate(vals))):
        m   = ms[i]
        row = row[:m.start()] + m.group(1) + xe(v) + m.group(3) + row[m.end():]
    return row

def repl(xml, old, new):
    o, n = xe(old), xe(new)
    xml = xml.replace(f'<w:t>{o}</w:t>', f'<w:t>{n}</w:t>')
    xml = xml.replace(f'<w:t xml:space="preserve">{o}</w:t>',
                      f'<w:t xml:space="preserve">{n}</w:t>')
    return xml

def find_para_start(xml, search_text, after=0):
    idx = xml.find(search_text, after)
    if idx == -1: return -1
    return xml.rfind('<w:p ', 0, idx)

def find_para_end(xml, para_start):
    end = xml.find('</w:p>', para_start)
    return end + len('</w:p>') if end != -1 else -1

def _find_tbl(x, tr_start):
    s = x.rfind('<w:tbl>', 0, tr_start)
    depth = 0; i = s
    while i < len(x):
        if x[i:i+7] == '<w:tbl>': depth += 1
        elif x[i:i+8] == '</w:tbl>':
            depth -= 1
            if depth == 0: return s, i + 8
        i += 1
    return s, len(x)

def _extract_para(x, keyword):
    pos = x.rfind(f'>{keyword}<')
    if pos < 0: return ''
    ps = x.rfind('<w:p ', 0, pos)
    pe = x.find('</w:p>', pos) + len('</w:p>')
    return x[ps:pe]

# ── Load CEE template ─────────────────────────────────────────────────────────
print(f"Reading template: {os.path.basename(TEMPLATE_PATH)}")
with open(TEMPLATE_PATH, 'rb') as f:
    tmpl_bytes = f.read()
zin = zipfile.ZipFile(io.BytesIO(tmpl_bytes), 'r')
all_files = {name: zin.read(name) for name in zin.namelist()}
zin.close()

try:    xml = all_files['word/document.xml'].decode('utf-8')
except: xml = all_files['word/document.xml'].decode('latin-1')

def _finalize_xml(xml_str):
    return re.sub(r'<w:trHeight[^/]*/>', '', xml_str)

# ── Step 1: Scalar text replacements ─────────────────────────────────────────
print("Updating scalar values...")
for old, new in [
    ('Saipol Lezoux',  CUSTOMER),
    ('Saipol',         CUSTOMER),
    ('France',         PLANT),
    ('Customer Input', DATA_SOURCE),
    ('09.19.2025',     RPT_DATE),
    ('0.14 EUR/kWh',   f'{float(ELEC_PRICE):.2f} {CURRENCY}/kWh'),
    ('0.055 kg/kWh',   f'{float(CO2_INTENSITY):.3f} kg/kWh'),
    ('25%',            pct_str(TAX_RATE)),
    ('EUR 3,312,551',  f'{SYM} {fmt(ANNUAL_SAVINGS)}'),
    ('EUR 15,679,593', f'{SYM} {fmt(INVEST_COST)}'),
    ('INR 3,312,551',  f'{SYM} {fmt(ANNUAL_SAVINGS)}'),
    ('INR 15,679,593', f'{SYM} {fmt(INVEST_COST)}'),
    ('16%',            fmtirr(TOP10_IRR)),
    ('224 tCO2',       f"{round(float(CO2_SAVINGS or 0)/1000)} tCO2"),
    ('118 Vehicles\xa0', f"{round(float(BEV_COUNT or 0))} Vehicles\xa0"),
    ('11 / 18',        f'{NPV_POS_CNT} / {NA}'),
    ('11/18',          f'{NPV_POS_CNT}/{NA}'),
    ('Energy Cost (EUR)',          f'Energy Cost ({CURRENCY})'),
    ('Investment (EUR)',           f'Investment ({CURRENCY})'),
    ('Energy Cost Savings (EUR)',  f'Energy Cost Savings ({CURRENCY})'),
    ('Energy Cost (INR)',          f'Energy Cost ({CURRENCY})'),
    ('Investment (INR)',           f'Investment ({CURRENCY})'),
    ('Energy Cost Savings (INR)',  f'Energy Cost Savings ({CURRENCY})'),
]:
    xml = repl(xml, old, new)

# Replace bare currency in headers
xml = re.sub(r'(<w:t[^>]*>)(INR)(</w:t>)',
             lambda m: m.group(1) + xe(CURRENCY) + m.group(3), xml)
xml = re.sub(r'(<w:t[^>]*>)(EUR)(</w:t>)',
             lambda m: m.group(1) + xe(CURRENCY) + m.group(3), xml)

# Update heading
xml = repl(xml,
    'Energy savings with ABB premium efficiency solutions \u2013 Top 10',
    f'Energy savings with ABB premium efficiency solutions \u2013 Top {NT}')

# Remove "Top 10" from summary labels when NA <= 10
if NA <= 10:
    xml = repl(xml, 'Payback time \u2013 Top 10*', 'Payback time*')
    xml = repl(xml, 'Internal rate of return \u2013 Top 10*', 'Internal rate of return*')
    print(f"  Summary: 'Top 10' removed from labels (NA={NA} \u2264 10)")
else:
    print(f"  Summary: 'Top 10' kept in labels (NA={NA} > 10)")

# ── Step 2: Footer / header dates ────────────────────────────────────────────
print(f"Updating footer dates \u2192 {RPT_DATE_ISO}")
for fname in ['word/footer1.xml', 'word/footer2.xml', 'word/footer3.xml',
              'word/header1.xml', 'word/header2.xml', 'word/header3.xml']:
    if fname in all_files:
        try:    txt = all_files[fname].decode('utf-8')
        except: txt = all_files[fname].decode('latin-1')
        txt = re.sub(r'\d{4}-\d{2}-\d{2}', RPT_DATE_ISO, txt)
        all_files[fname] = txt.encode('utf-8')

# ── Step 3: Charts ────────────────────────────────────────────────────────────
cons_before = float(CONSUMP_BEFORE or 0)
cons_after  = cons_before - float(SAVINGS_KWH or 0)

if cons_before >= 1_000_000:
    unit = 'GWh'; dlfmt = r'#.0,,\ &quot;GWh&quot;'
    disp_b = f'{cons_before/1e6:.1f} GWh'; disp_a = f'{cons_after/1e6:.1f} GWh'
elif cons_before >= 1_000:
    unit = 'MWh'; dlfmt = r'#.0,\ &quot;MWh&quot;'
    disp_b = f'{cons_before/1e3:.1f} MWh'; disp_a = f'{cons_after/1e3:.1f} MWh'
else:
    unit = 'kWh'; dlfmt = r'#,##0\ &quot;kWh&quot;'
    disp_b = f'{cons_before:.0f} kWh'; disp_a = f'{cons_after:.0f} kWh'

def numlit_1pt(fmt_code, value):
    return (f'<c:numLit><c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="1"/>'
            f'<c:pt idx="0"><c:v>{value:.10f}</c:v></c:pt></c:numLit>')

def numlit_5pt(fmt_code, values):
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v:.10f}</c:v></c:pt>'
                  for i, v in enumerate(values))
    return (f'<c:numLit><c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="5"/>{pts}</c:numLit>')

if 'word/charts/chart1.xml' in all_files:
    ctxt = all_files['word/charts/chart1.xml'].decode('utf-8')
    nfmt = r'#,##0_);[Red]\(#,##0\)'
    ctxt = re.sub(r'<c:numRef><c:f>Saving_Calculations!\$C\$16</c:f>.*?</c:numRef>',
                  numlit_1pt(nfmt, cons_before), ctxt, flags=re.DOTALL)
    ctxt = re.sub(r'<c:numRef><c:f>Saving_Calculations!\$C\$18</c:f>.*?</c:numRef>',
                  numlit_1pt(nfmt, cons_after),  ctxt, flags=re.DOTALL)
    ctxt = ctxt.replace('#.0,,\\ &quot;GWh&quot;', dlfmt)
    # Also patch template-specific numeric literals
    for old_v, new_v in [('1013151.16', f'{cons_before:.10f}'),
                          ('976805.69000000006', f'{cons_after:.10f}'),
                          ('976805.6900000001',  f'{cons_after:.10f}'),
                          ('976805.69',          f'{cons_after:.10f}')]:
        ctxt = ctxt.replace(old_v, new_v)
    ctxt = ctxt.replace('#,##0,\\ &quot;MWh&quot;', dlfmt)
    ctxt = ctxt.replace('#.0,\\ &quot;MWh&quot;',   dlfmt)
    all_files['word/charts/chart1.xml'] = ctxt.encode('utf-8')
    print(f"  Chart 1 (Energy Consumption): {disp_b} \u2192 {disp_a}  [{unit}]")

if 'word/embeddings/Microsoft_Excel_Worksheet.xlsx' in all_files:
    try:
        import openpyxl as _xl
        emb = _xl.load_workbook(io.BytesIO(
            all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx']))
        ews = emb.active
        ews['B2'] = NPV_POS_CNT
        ews['B3'] = NA - NPV_POS_CNT
        ews['A2'] = 'NPV Positive'
        ews['A3'] = 'Remaining Assets'
        buf2 = io.BytesIO(); emb.save(buf2)
        all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx'] = buf2.getvalue()
        if 'word/charts/chart2.xml' in all_files:
            c2 = all_files['word/charts/chart2.xml'].decode('utf-8')
            c2 = c2.replace('<c:v>192</c:v>', f'<c:v>{NPV_POS_CNT}</c:v>')
            c2 = c2.replace('<c:v>6</c:v>',   f'<c:v>{NA - NPV_POS_CNT}</c:v>')
            c2 = c2.replace('<c:v>2nd Qtr</c:v>', '<c:v>Remaining Assets</c:v>')
            all_files['word/charts/chart2.xml'] = c2.encode('utf-8')
        print(f"  Chart 2 (NPV pie): {NPV_POS_CNT} NPV+ / {NA - NPV_POS_CNT} remaining")
    except Exception as e:
        print(f"  Warning: could not update embedded Excel: {e}")

if 'word/charts/chart3.xml' in all_files and SENSITIVITY:
    ctxt = all_files['word/charts/chart3.xml'].decode('utf-8')
    # Pass 1: formula ref → numLit (Standard col T/V and CEE col S/U patterns)
    for col_pair in [('T', 'V'), ('S', 'U')]:
        ctxt = re.sub(
            rf'<c:numRef><c:f>[^<]*\${col_pair[0]}\$19:\${col_pair[0]}\$23</c:f>.*?</c:numRef>',
            numlit_5pt('0%', [s[0] for s in SENSITIVITY]), ctxt, flags=re.DOTALL)
        ctxt = re.sub(
            rf'<c:numRef><c:f>[^<]*\${col_pair[1]}\$19:\${col_pair[1]}\$23</c:f>.*?</c:numRef>',
            numlit_5pt('0.00', [s[1] for s in SENSITIVITY]), ctxt, flags=re.DOTALL)
    ctxt = re.sub(r'<c:f>[^<]+</c:f>', '', ctxt)
    # Pass 2: numCache patch for any remaining numRef blocks
    _refs = list(re.finditer(r'<c:numRef>', ctxt))
    if len(_refs) >= 2:
        _rs  = _refs[1].start()
        _end = ctxt.find('</c:numRef>', _rs) + len('</c:numRef>')
        _seg = ctxt[_rs:_end]
        _pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>'
                       for i, (_, v) in enumerate(SENSITIVITY))
        _seg = re.sub(r'<c:ptCount[^/]*/>', f'<c:ptCount val="{len(SENSITIVITY)}"/>', _seg)
        _seg = re.sub(r'(<c:ptCount[^/]*/>)(.*?)(?=</c:numCache>)',
                      lambda m: m.group(1) + _pts, _seg, flags=re.DOTALL)
        ctxt = ctxt[:_rs] + _seg + ctxt[_end:]
    # Remove external data link — prevents Word from refreshing chart on open
    ctxt = re.sub(r'<c:externalData\b[^>]*/>', '', ctxt)
    ctxt = re.sub(r'<c:externalData\b.*?</c:externalData>', '', ctxt, flags=re.DOTALL)
    all_files['word/charts/chart3.xml'] = ctxt.encode('utf-8')
    _rels_name = 'word/charts/_rels/chart3.xml.rels'
    if _rels_name in all_files:
        _rels = all_files[_rels_name].decode('utf-8')
        _rels = re.sub(r'<Relationship\b[^>]*Id="rId4"[^/]*/>', '', _rels)
        all_files[_rels_name] = _rels.encode('utf-8')
    print(f"  Chart 3 (Payback sensitivity): {[round(s[1], 2) for s in SENSITIVITY]}")

# ── Step 4a: Remove IRR row when NPV ≤ 0 ─────────────────────────────────────
NO_NPV = float(TOP10_NPV or 0) <= 0
if NO_NPV:
    tr_pre = get_tr_pos(xml)
    xml = xml[:tr_pre[15][0]] + xml[tr_pre[15][1]:]
    print("  Summary: NPV \u2264 0 \u2014 IRR row removed")
irr_adj = -1 if NO_NPV else 0

# ── Step 4: Cover table TR-level updates ─────────────────────────────────────
# Uses nth_wt (position-based) for all cover rows so updates work regardless of
# what placeholder values the template currently holds (avoids repl() split-run failures).
tr = get_tr_pos(xml)

# TR 1: Customer name  (node[0]=label, node[1]=value)
r1 = strip_ids(xml[tr[1][0]:tr[1][1]])
r1 = nth_wt(r1, 1, CUSTOMER)
xml = xml[:tr[1][0]] + r1 + xml[tr[1][1]:]
tr  = get_tr_pos(xml)

# TR 2: Plant / site  (node[0]=label, node[1]=value)
r2 = strip_ids(xml[tr[2][0]:tr[2][1]])
r2 = nth_wt(r2, 1, PLANT)
xml = xml[:tr[2][0]] + r2 + xml[tr[2][1]:]
tr  = get_tr_pos(xml)

# TR 3: Data Source  (node[0]=label, node[1]=value)
r3 = strip_ids(xml[tr[3][0]:tr[3][1]])
r3 = nth_wt(r3, 1, DATA_SOURCE)
xml = xml[:tr[3][0]] + r3 + xml[tr[3][1]:]
tr  = get_tr_pos(xml)

# TR 4: Date of Report  (node[0]=label, nodes[1..6]=split date digits → write full date to [1])
r4 = strip_ids(xml[tr[4][0]:tr[4][1]])
r4 = nth_wt(r4, 1, RPT_DATE)
for _i in range(2, 7): r4 = nth_wt(r4, _i, '')
xml = xml[:tr[4][0]] + r4 + xml[tr[4][1]:]
tr  = get_tr_pos(xml)

# TR 5: NPV+ count / total assets  (node[1]=NPV+, node[3]=total)
r5 = strip_ids(xml[tr[5][0]:tr[5][1]])
r5 = nth_wt(r5, 1, str(NPV_POS_CNT))
r5 = nth_wt(r5, 3, str(NA))
xml = xml[:tr[5][0]] + r5 + xml[tr[5][1]:]
tr  = get_tr_pos(xml)

# TR 6: Electricity Cost  (node[0]=label, nodes[1..4]=split value → write full to [1])
# Template splits e.g. "0." + "14" + " EUR" + "/kWh" across 4 runs.
r6 = strip_ids(xml[tr[6][0]:tr[6][1]])
r6 = nth_wt(r6, 1, f'{float(ELEC_PRICE):.2f} {CURRENCY}/kWh')
for _i in range(2, 5): r6 = nth_wt(r6, _i, '')
xml = xml[:tr[6][0]] + r6 + xml[tr[6][1]:]
tr  = get_tr_pos(xml)

# TR 7: Carbon Intensity  (node[0]=label, nodes[1..4]=split value → write full to [1])
r7 = strip_ids(xml[tr[7][0]:tr[7][1]])
r7 = nth_wt(r7, 1, f'{float(CO2_INTENSITY):.3f} kg/kWh')
for _i in range(2, 5): r7 = nth_wt(r7, _i, '')
xml = xml[:tr[7][0]] + r7 + xml[tr[7][1]:]
tr  = get_tr_pos(xml)

# TR 8: Tax Rate  (node[0]=label, nodes[1..3]=split digits+% → write full to [1])
r8 = strip_ids(xml[tr[8][0]:tr[8][1]])
r8 = nth_wt(r8, 1, f'{round(float(TAX_RATE) * 100)}%')
for _i in range(2, 4): r8 = nth_wt(r8, _i, '')
xml = xml[:tr[8][0]] + r8 + xml[tr[8][1]:]
tr  = get_tr_pos(xml)
print(f"  Cover: Customer={CUSTOMER}  Plant={PLANT}  Date={RPT_DATE}"
      f"  Elec={float(ELEC_PRICE):.2f} {CURRENCY}/kWh  CO2={float(CO2_INTENSITY):.3f} kg/kWh"
      f"  Tax={round(float(TAX_RATE)*100)}%")

# TR 12: Annual savings + bar chart asset counts
r12 = strip_ids(xml[tr[12][0]:tr[12][1]])
r12 = nth_wt(r12,  0, f'{SYM} {fmt(ANNUAL_SAVINGS)}')
r12 = nth_wt(r12,  1, ''); r12 = nth_wt(r12, 2, ''); r12 = nth_wt(r12, 3, '')
r12 = nth_wt(r12, 10, f'0{NPV_POS_CNT}' if NPV_POS_CNT < 10 else str(NPV_POS_CNT))
r12 = nth_wt(r12, 12, str(NA))
xml = xml[:tr[12][0]] + r12 + xml[tr[12][1]:]
tr  = get_tr_pos(xml)

# TR 14: Payback (×2) + NPV value
r14    = strip_ids(xml[tr[14][0]:tr[14][1]])
pay_str = fmtyrs(PAYBACK_TIME)   # all NPV+ assets payback on cover page
r14 = nth_wt(r14, 0, pay_str); r14 = nth_wt(r14, 1, '')
r14 = nth_wt(r14, 2, '');      r14 = nth_wt(r14, 3, ' yrs')
r14 = nth_wt(r14, 4, pay_str); r14 = nth_wt(r14, 5, '')
r14 = nth_wt(r14, 6, '');      r14 = nth_wt(r14, 7, ' yrs')
npv_full = f'{SYM} {fmt(INVEST_COST)}' if NO_NPV else f'{SYM} {fmt(NPV_VALUE)}'
# TR14 nodes [14]='EUR 5,' [15]='802' are the split NPV value; [16]='Net Present Value*' label
r14 = nth_wt(r14, 14, npv_full); r14 = nth_wt(r14, 15, '')
xml = xml[:tr[14][0]] + r14 + xml[tr[14][1]:]
tr  = get_tr_pos(xml)

# TR 15: Investment + IRR (skip when IRR row was removed)
if not NO_NPV:
    r15 = strip_ids(xml[tr[15 + irr_adj][0]:tr[15 + irr_adj][1]])
    r15 = nth_wt(r15, 0, f'{SYM} {fmt(INVEST_COST)}')
    r15 = nth_wt(r15, 1, ''); r15 = nth_wt(r15, 2, ''); r15 = nth_wt(r15, 3, '')
    irr_str = fmtirr(TOP10_IRR)
    r15 = nth_wt(r15, 5, irr_str); r15 = nth_wt(r15, 6, '')
    r15 = nth_wt(r15, 7, irr_str); r15 = nth_wt(r15, 8, '')
    xml = xml[:tr[15 + irr_adj][0]] + r15 + xml[tr[15 + irr_adj][1]:]
    tr  = get_tr_pos(xml)

# TR 17: CO2 savings + BEV count
r17 = strip_ids(xml[tr[17 + irr_adj][0]:tr[17 + irr_adj][1]])
r17 = nth_wt(r17, 0, str(round(float(CO2_SAVINGS or 0) / 1000)))
r17 = nth_wt(r17, 5, str(round(float(BEV_COUNT or 0))))
xml = xml[:tr[17 + irr_adj][0]] + r17 + xml[tr[17 + irr_adj][1]:]
tr  = get_tr_pos(xml)

# Clear "Credit Voucher & Report via Take-back Program" cell from sustainability row
_r17s, _r17e = tr[17 + irr_adj]
_r17xml = xml[_r17s:_r17e]
_tcs17  = list(re.finditer(r'<w:tc>', _r17xml))
for _tm in reversed(_tcs17):
    _tc_s = _tm.start()
    _tc_e = _r17xml.find('</w:tc>', _tc_s) + len('</w:tc>')
    _tc_c = _r17xml[_tc_s:_tc_e]
    if 'take-back' in _tc_c.lower() or 'Credit Voucher' in _tc_c or 'takeback' in _tc_c.lower():
        _tcp_m = re.search(r'<w:tcPr>.*?</w:tcPr>', _tc_c, re.DOTALL)
        _tcp_s = _tcp_m.group(0) if _tcp_m else ''
        _blank = f'<w:tc>{_tcp_s}<w:p><w:r><w:t></w:t></w:r></w:p></w:tc>'
        _r17xml = _r17xml[:_tc_s] + _blank + _r17xml[_tc_e:]
        break
xml = xml[:_r17s] + _r17xml + xml[_r17e:]
tr  = get_tr_pos(xml)

# ── Step 5: Top-N data rows ───────────────────────────────────────────────────
# CEE table layout (14 cols in template):
#   #, App, E-Cons, E-Cost, CO2-Cons, E-Sav-kWh, Invest, Invest-CEE, E-Cost-Sav,
#   Payback(std), E-Sav-%, Payback-CEE, CO2-Sav, Takeback(✓)
# Remove ONLY col 13 (Takeback); restore col 9 (std payback) by injecting <w:t>:
#   → final: 13 cols, 13 <w:t> per data row, 12 <w:t> per total row

def remove_last_tc(row_xml):
    """Remove the last <w:tc>...</w:tc> from a table row."""
    p = row_xml.rfind('<w:tc>')
    if p == -1: return row_xml
    e = row_xml.find('</w:tc>', p) + len('</w:tc>')
    return row_xml[:p] + row_xml[e:]

print(f"Building Top-{NT} table ({NT} rows)...")
tr = get_tr_pos(xml)
_hdr_idx = 18 + irr_adj   # header row of top-10 table

# Helper: remove the Nth gridCol (0-indexed) from the tblGrid before _limit
def _remove_nth_gridcol(xml_str, tbl_start, limit, n):
    gcs = list(re.finditer(r'<w:gridCol[^/]*/>', xml_str[tbl_start:limit]))
    if n >= len(gcs): return xml_str
    gc = gcs[n]
    abs_s = tbl_start + gc.start()
    abs_e = tbl_start + gc.end()
    return xml_str[:abs_s] + xml_str[abs_e:]

# 1. Remove ONLY gridCol[13] (Takeback) — col 9 (std payback) is kept and filled
_tbl_s = xml.rfind('<w:tbl>', 0, tr[_hdr_idx][0])
xml = _remove_nth_gridcol(xml, _tbl_s, tr[_hdr_idx][0], 13)
tr  = get_tr_pos(xml)

# 2. Remove last <w:tc> (Takeback) from header row; keep cell[9] "Payback time (years)"
_hs, _he = tr[_hdr_idx]
_hdr_row = remove_last_tc(xml[_hs:_he])
xml = xml[:_hs] + _hdr_row + xml[_he:]
tr  = get_tr_pos(xml)

# 3. Read data-row template; strip Takeback (last); inject <w:t> into empty cell[9]
#    (Cell[9] "Payback time" has the paragraph structure but no run/text node.)
tmpl_t10 = xml[tr[19 + irr_adj][0]:tr[19 + irr_adj][1]]
tmpl_t10 = remove_last_tc(tmpl_t10)         # remove Takeback (was cell[13])
tmpl_t10 = _inject_wt_in_cell(tmpl_t10, 9)  # inject bare <w:t> so fill_row can fill it

def v_top10(a):
    """13 values for 13 <w:t> nodes."""
    is_p       = a['npv'] is not None and float(a['npv']) > 0
    sav        = float(a['e_sav_cost'] or 0)
    inv_g      = float(a['invest']     or 0)
    inv_cee    = float(a['invest_cee'] or 0)
    # Standard payback (gross investment, no CEE subsidy)
    std_pb     = a['payback']
    if std_pb is None and is_p and sav > 0:
        std_pb = inv_g / sav
    std_pb_str = fmtyrs(std_pb) if is_p else ''
    # CEE payback (net investment after subsidy)
    pb_str     = f"{inv_cee / sav:.1f} " if (is_p and sav > 0) else ''
    return [
        str(a['num']),             # node 0  → col 0  (#)
        a['tag'],                  # node 1  → col 1  (Application)
        fmt(a['e_cons']),          # node 2  → col 2  (Energy Cons)
        fmt(a['e_cost']),          # node 3  → col 3  (Energy Cost)
        fmt(a['co2_cons']),        # node 4  → col 4  (CO2 Cons)
        fmt(a['e_sav_kwh']),       # node 5  → col 5  (Energy Sav kWh)
        fmt(a['invest']),          # node 6  → col 6  (Investment gross)
        fmt(a['invest_cee']),      # node 7  → col 7  (Investment – CEE net)
        fmt(a['e_sav_cost']),      # node 8  → col 8  (Energy Cost Savings)
        std_pb_str,                # node 9  → col 9  (Payback time, std gross)
        fmtpct(a['e_sav_pct']),   # node 10 → col 10 (Energy Sav %)
        pb_str,                    # node 11 → col 11 (Payback Including CEE)
        f"{fmt(a['co2_sav'])} ",  # node 12 → col 12 (CO2 Sav)
    ]

def build_top10_row(a):
    """Fill one top-table row; white-shade NPV-negative assets (cells 0-9)."""
    row_xml    = fill_row(tmpl_t10, v_top10(a))
    is_npv_neg = (a['npv'] is None or float(a['npv'] or 0) <= 0) and not NO_NPV_MODE
    if is_npv_neg:
        # Cells 0-9 white/no-fill; cells 10+ (E-Sav%, Payback-CEE, CO2-Sav) keep original
        tc_starts = [m.start() for m in re.finditer(r'<w:tc>', row_xml)]
        if len(tc_starts) >= 10:
            cut     = tc_starts[10]
            before  = row_xml[:cut].replace('w:fill="E2EFD9"', 'w:fill="FFFFFF"')
            row_xml = before + row_xml[cut:]
    return row_xml

block = ''.join(build_top10_row(a) for a in top10)
xml   = xml[:tr[19 + irr_adj][0]] + block + xml[tr[28 + irr_adj][1]:]
tr    = get_tr_pos(xml)

# Dynamic index formulas
I_TR1       = 19 + NT + irr_adj   # "Total Energy Savings Assets – Top (N)" row
I_TR2       = 20 + NT + irr_adj   # "Total NPV Positive assets" data row
I_APP_FIRST = 22 + NT + irr_adj   # first App Details data row
I_APP_LAST  = 31 + NT + irr_adj   # last of 10 original App Details rows

# 4. Prepare total rows: strip Takeback (last cell); inject <w:t> into payback cell[8].
#    (Total rows have cell[0] gridSpan=2; std-payback is physical cell[8].)
for _tidx in [I_TR1, I_TR2]:
    _rs, _re = tr[_tidx]
    _rxml    = remove_last_tc(xml[_rs:_re])      # remove Takeback (last cell)
    _rxml    = _inject_wt_in_cell(_rxml, 8)       # inject <w:t> into std payback cell[8]
    xml = xml[:_rs] + _rxml + xml[_re:]
    tr  = get_tr_pos(xml)

# ── Step 6: Total rows ────────────────────────────────────────────────────────
# After removals both total rows have 12 cells and 12 <w:t> nodes:
#   node[0]=label (cell[0] span=2), node[1]=e_cons, node[2]=e_cost,
#   node[3]=co2_cons, node[4]=e_sav_kwh, node[5]=invest_g, node[6]=invest_cee,
#   node[7]=e_sav_cost, node[8]=std_payback, node[9]=e_sav_pct,
#   node[10]=payback_cee, node[11]=co2_sav
# Use I_TR2 as XML template for BOTH rows via fill_row.
print("Updating total rows...")
tmpl_total = strip_ids(xml[tr[I_TR2][0]:tr[I_TR2][1]])

def _fill_total(label, e_cons, e_cost, co2_c, sav_kwh,
                inv_g, inv_c, sav_c, std_payback, pct, payback_cee, co2):
    return fill_row(tmpl_total, [
        label,
        fmt(e_cons), fmt(e_cost), fmt(co2_c), fmt(sav_kwh),
        fmt(inv_g),  fmt(inv_c),  fmt(sav_c),
        fmtyrs(std_payback), fmtpct(pct), fmtyrs(payback_cee), fmt(co2),
    ])

# Row 1: always shown — Top-N totals
# Styling: medium green background (70AD47, matching Avoided CO2 / Payback column headers)
#          with automatic black text (matching template TR[29] original style)
r_tr1 = _fill_total(
    f'Total Energy Savings Assets \u2013 Top ({NT})',
    t10_e_cons, t10_e_cost, t10_co2_c, t10_sav_kwh,
    t10_invest_g, t10_invest_c, t10_sav_c,
    t10_std_payback, t10_pct, TOP10_PAYBACK, t10_co2_sav,
)
# tmpl_total is from TR[30] which uses 3E5F27 (dark green) + FFFFFF (white text).
# For I_TR1 we want 70AD47 (medium green, same as Avoided CO2 column header) + black text.
r_tr1 = r_tr1.replace('w:fill="3E5F27"', 'w:fill="70AD47"')
r_tr1 = re.sub(r'<w:color\b[^/]*/>', '<w:color w:val="000000"/>', r_tr1)
xml = xml[:tr[I_TR1][0]] + r_tr1 + xml[tr[I_TR1][1]:]
tr  = get_tr_pos(xml)

# Row 2: NPV+ totals when count differs from Top-N; removed otherwise
if NPV_POS_CNT != NT and NPV_POS_CNT > 0 and not NO_NPV_MODE:
    r_tr2 = _fill_total(
        f'Total NPV Positive Assets ({NPV_POS_CNT})',
        all_e_cons, all_e_cost, all_co2_c, all_sav_kwh,
        all_invest_g, all_invest_c, all_sav_c,
        all_std_payback, all_pct, all_payback, all_co2,
    )
    xml = xml[:tr[I_TR2][0]] + r_tr2 + xml[tr[I_TR2][1]:]
    npv_adj = 0
    print(f"  Total rows: Top-{NT} + NPV Positive ({NPV_POS_CNT})")
else:
    xml     = xml[:tr[I_TR2][0]] + xml[tr[I_TR2][1]:]
    npv_adj = -1
    print(f"  Total rows: Top-{NT} only (NPV count matches or no NPV)")
tr = get_tr_pos(xml)

# ── Step 7: Application Details motor table ───────────────────────────────────
print(f"Building Application Details ({NT} rows)...")
I_APP_FIRST_a = I_APP_FIRST + npv_adj
I_APP_LAST_a  = I_APP_LAST  + npv_adj

def v_app(a):
    return [str(a['num']), a['tag'], a['ie'], a['load'],
            a['flow_ctrl'], a['connection'],
            str(a['output_kw']), str(a['shaft_h']),
            str(a['run_hrs']), str(a['avg_val']),
            a['ess_motor'], a['ess_conn']]

tmpl_app = xml[tr[I_APP_FIRST_a][0]:tr[I_APP_FIRST_a][1]]
block    = ''.join(fill_row(tmpl_app, v_app(a)) for a in top10)
xml      = xml[:tr[I_APP_FIRST_a][0]] + block + xml[tr[I_APP_LAST_a][1]:]
tr       = get_tr_pos(xml)

# ── Step 8: Appendix — inject App Details + Det Rec when NA > 10 ──────────────
# CEE template does not have its own appendix tables.
# Borrow header/data row XML from EA_Report_Template_Standard.docx.
data_listed_pos     = xml.rfind('Data is listed as total annual')
if data_listed_pos == -1: data_listed_pos = 0
appendix_para_start = find_para_start(xml, '>Appendix<', data_listed_pos)
appendix_para_end   = find_para_end(xml, appendix_para_start)
calcmeth_para       = find_para_start(xml, '>Calculation Methodology<',
                                      appendix_para_start + 1)

def _strip_row_heights(t):
    return re.sub(r'<w:trHeight[^/]*/>', '', t)

if NA > 10 and appendix_para_end > 0 and calcmeth_para > appendix_para_start:
    try:
        with open(TEMPLATE_STD_PATH, 'rb') as _f:
            _stxml = zipfile.ZipFile(io.BytesIO(_f.read()), 'r') \
                            .read('word/document.xml').decode('utf-8')
        _sttr = get_tr_pos(_stxml)

        app_det_hdg   = _extract_para(_stxml, 'Application Details')
        notes_hdg     = _extract_para(_stxml, 'Notes on Payback')
        det_rec_hdg   = _extract_para(_stxml, 'Details of Recommendation')
        det_more_para = _extract_para(_stxml, 'More information about the various devices')
        det_info_para = _extract_para(_stxml, 'This section contains an overview')

        _notes_pos = _stxml.rfind('>Notes on Payback<')
        _notes_ps  = _stxml.rfind('<w:p ', 0, _notes_pos)
        _notes_pe  = _stxml.find('</w:p>', _notes_pos) + len('</w:p>')
        notes_body = _stxml[_notes_pe: _stxml.find('</w:p>', _notes_pe) + len('</w:p>')]

        apx_hdr_tmpl = strip_ids(_stxml[_sttr[42][0]:_sttr[42][1]])
        apx_dat_tmpl = strip_ids(_stxml[_sttr[43][0]:_sttr[43][1]])
        apx_hdr_tmpl = re.sub(r'>([A-Z]{3})<',
                               lambda m: f'>{xe(CURRENCY)}<', apx_hdr_tmpl)

        def v_app2(a):
            npv   = a['npv']
            is_p  = npv is not None and float(npv) > 0
            npv_s = (fmt(npv) + ' ') if is_p else (fmt_parens(npv) if npv is not None else '')
            co2_r = float(a['co2_sav'] or 0)
            co2_s = (fmt(co2_r) + ' ') if co2_r > 0 else fmt_parens(co2_r)
            sav   = float(a['e_sav_cost'] or 0)
            inv_c = float(a['invest_cee'] or 0)
            pb_s  = f"{inv_c / sav:.1f} " if (is_p and sav > 0) else ''
            return [str(a['num']), a['tag'],
                    fmt(a['e_cons']), fmt(a['e_cost']), fmt(a['co2_cons']),
                    fmt(a['e_sav_kwh']), fmt(a['invest_cee']), fmt(a['e_sav_cost']),
                    pb_s, co2_s, fmtpct(a['e_sav_pct']), npv_s]

        apx_data      = ''.join(fill_row(apx_dat_tmpl, v_app2(a)) for a in assets_by_num)
        apx_tbl_s, _  = _find_tbl(_stxml, _sttr[42][0])
        apx_tbl_xml   = _stxml[apx_tbl_s:_sttr[42][1]] + apx_data + '</w:tbl>'

        det_hdr_tmpl  = strip_ids(_stxml[_sttr[61][0]:_sttr[61][1]])
        det_dat_tmpl  = strip_ids(_stxml[_sttr[62][0]:_sttr[62][1]])
        det_data      = ''.join(fill_row(det_dat_tmpl, v_app(a)) for a in assets_by_num)
        det_tbl_s, _  = _find_tbl(_stxml, _sttr[61][0])
        det_tbl_xml   = _stxml[det_tbl_s:_sttr[61][1]] + det_data + '</w:tbl>'

        inject = (app_det_hdg + _strip_row_heights(apx_tbl_xml) +
                  '<w:p><w:r><w:br w:type="page"/></w:r></w:p>' +
                  notes_hdg + notes_body +
                  det_rec_hdg + det_more_para + det_info_para +
                  _strip_row_heights(det_tbl_xml))

        xml = xml[:appendix_para_end] + inject + xml[calcmeth_para:]
        print(f"  Appendix: injected App Details + Det Rec ({NA} rows each)")

    except Exception as e:
        print(f"  Warning: could not inject Appendix tables: {e}")

elif appendix_para_end > 0 and calcmeth_para > appendix_para_start:
    xml = xml[:appendix_para_end] + xml[calcmeth_para:]
    print(f"Appendix trimmed (NA={NA} \u2264 10): kept only Calculation Methodology + NPV Methodology")

# ── Save output ───────────────────────────────────────────────────────────────
safe_c   = re.sub(r'[\\/:*?"<>|]', '_', args[1])
safe_p   = re.sub(r'[\\/:*?"<>|]', '_', PLANT)
out_name = f'{safe_c}_{safe_p}_EA_Report_CEE.docx'
out_path = os.path.join(_script_dir, out_name)

all_files['word/document.xml'] = _finalize_xml(xml).encode('utf-8')
if 'word/settings.xml' in all_files:
    try:
        stg = all_files['word/settings.xml'].decode('utf-8')
        if 'updateFields' not in stg:
            stg = stg.replace('</w:settings>',
                              '<w:updateFields w:val="1"/></w:settings>')
        all_files['word/settings.xml'] = stg.encode('utf-8')
    except Exception: pass

buf = io.BytesIO()
with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
    for name, data in all_files.items():
        zout.writestr(name, data)
with open(out_path, 'wb') as f:
    f.write(buf.getvalue())

print()
print(f"  Done!  \u2192  {out_path}")
print(f"  {CUSTOMER} | {PLANT} | {CURRENCY} | {NA} assets | {NPV_POS_CNT} NPV+ | {NT} Top-N")
print(f"  Annual savings : {SYM} {fmt(ANNUAL_SAVINGS)}")
print(f"  Investment\u2013CEE : {SYM} {fmt(INVEST_COST)}")
print(f"  Payback        : {fmtyrs(PAYBACK_TIME)} yrs  (Top-{NT}: {fmtyrs(TOP10_PAYBACK)} yrs)")
print(f"  NPV            : {SYM} {fmt(NPV_VALUE)}")
print(f"  IRR Top-{NT}      : {fmtirr(TOP10_IRR)}")
if SENSITIVITY:
    _arrow = '\u2192'
    print(f"  Sensitivity    : {[f'{s[0]:+.0%}{_arrow}{s[1]:.1f}yrs' for s in SENSITIVITY]}")
