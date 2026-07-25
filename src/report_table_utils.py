#!/usr/bin/env python3
"""
Shared docx table-rendering engine for the EA report generators.

Every report template (Executive, Complete Asset, ...) has exactly one
Energy Savings table and one Application Details table. The only difference
between report types is how many ranked assets are populated into each
table (top 10 vs. all). This module provides a single reusable renderer:
locate the table following a section heading, validate its header row
against the field map, then clone its {{ROW_TEMPLATE}} row once per asset.
"""

import os
import re
import sys
import tempfile

import openpyxl

from excel_formula_engine import FormulaEngine, FormulaError


# ---------------------------------------------------------------------------
# Workbook recalculation
# ---------------------------------------------------------------------------
# openpyxl never evaluates formulas. A Saving Calculations workbook filled by
# fill_saving_calculations.py therefore carries stale/blank cached results for
# every formula cell (NPV, payback, IRR, KPI totals, the Top-10 rank column,
# ...) until those formulas are evaluated. The report generators must not
# work around that by recomputing the values themselves in a second,
# independent implementation (see generate_report_standard.py /
# generate_report_executive.py for why that silently diverges from the
# workbook's own formulas) -- they call recalculate_workbook() first instead,
# and fail loudly via verify_recalculated() if a cell still isn't populated.
#
# recalculate_workbook() evaluates the workbook's OWN formula text in-process
# via excel_formula_engine -- a small, generic Excel-function interpreter
# (SUM/IF/NPV/IRR/etc. with their standard, workbook-agnostic definitions),
# not a reimplementation of this workbook's financial model. No external
# application (LibreOffice, Excel) is involved. If the workbook's formulas
# change, this requires no code changes here -- it evaluates whatever formula
# text is actually in the cells.

_SC_SHEET_PREFIXES = ('saving_calculations', 'savings_calculations',
                      'saving_calculatios', 'savings_calculatios')


def _find_sc_sheet_for_recalc(wb):
    def is_sc(name):
        n = name.strip().lower().replace(' ', '_')
        return any(n.startswith(p) for p in _SC_SHEET_PREFIXES)

    candidates = [s for s in wb.sheetnames if is_sc(s)]
    if not candidates:
        raise RuntimeError(f"No Saving_Calculations sheet found. Sheets: {wb.sheetnames}")
    if len(candidates) > 1:
        versioned = [s for s in candidates if re.search(r'_v\d+$', s.strip().lower())]
        if versioned:
            return wb[versioned[0]]
    return wb[candidates[0]]


# Cells this engine evaluates and freezes as literal values. Column indices:
# C=3 D=4 E=5 F=6 G=7 H=8 I=9 J=10 K=11 L=12 M=13 Q=17 V=22 BF=58
_KPI_ROWS = (13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 25)
_GROUP_TOTAL_COLS = (5, 6, 7, 8, 9, 10, 11, 13, 17)
_PER_ASSET_COLS = (6, 7, 9, 10, 12, 13, 17)  # F,G,I,J,L,M,Q
_ASSET_START_ROW = 37
_ASSET_ROW_CEILING = 1100


def _is_formula(raw):
    return isinstance(raw, str) and raw.startswith('=')


def _freeze(ws, engine, row, col):
    """Evaluate ws cell (row, col) via the engine and overwrite it with the
    literal result, but only if it's actually a formula -- literal inputs
    (assumption cells, the C24 reference constant, ...) are left untouched."""
    cell = ws.cell(row=row, column=col)
    if _is_formula(cell.value):
        cell.value = engine.get(row, col)


def recalculate_workbook(xlsx_path):
    """
    Evaluate the Saving Calculations workbook's own formulas (NPV, payback,
    IRR, KPI totals, Top-10 ranking, ...) in-process and write the results
    as literal values into a temp copy, whose path is returned.

    Raises RuntimeError (never falls back to computing these values some
    other way) if a formula this engine doesn't support is encountered.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = _find_sc_sheet_for_recalc(wb)
    engine = FormulaEngine(ws, header_row=36, table_name="Saving_Table2410")

    # How many asset rows actually have data -- column B ('#') is a raw
    # literal, never a formula, so this needs no evaluation.
    num_assets = 0
    for i in range(_ASSET_ROW_CEILING):
        if ws.cell(row=_ASSET_START_ROW + i, column=2).value is None:
            break
        num_assets += 1

    try:
        for i in range(num_assets):
            r = _ASSET_START_ROW + i
            for col in _PER_ASSET_COLS:
                _freeze(ws, engine, r, col)
            _freeze(ws, engine, r, 58)  # BF: Payback Rank (Top 10) -- for audit visibility

        for row in _KPI_ROWS:
            for col in (3, 4):
                _freeze(ws, engine, row, col)
        for row in (33, 34):
            for col in _GROUP_TOTAL_COLS:
                _freeze(ws, engine, row, col)
        for row in range(19, 24):
            _freeze(ws, engine, row, 22)  # V: sensitivity payback
    except FormulaError as exc:
        raise RuntimeError(
            f"Could not recalculate {os.path.basename(xlsx_path)!r}: {exc}"
        ) from exc

    out_dir = tempfile.mkdtemp(prefix="recalc_")
    recalced_path = os.path.join(out_dir, os.path.basename(xlsx_path))
    wb.save(recalced_path)
    wb.close()
    return recalced_path


def _stale_cells(ws, checks):
    """(row, col, desc) entries among `checks` whose cell value is None."""
    return [
        (row, col, desc)
        for row, col, desc in checks
        if ws.cell(row=row, column=col).value is None
    ]


def cells_are_cached(ws, checks):
    """
    Return True if every (row, col, description) cell in `checks` already
    holds a non-None value -- i.e. the workbook's formulas were already
    evaluated by a real spreadsheet app (or a prior recalculation) and
    LibreOffice does not need to be invoked again.

    Use this to skip recalculate_workbook() entirely when it isn't needed --
    LibreOffice should only be a dependency for the (comparatively rare)
    case of a workbook whose formulas were never recalculated, not for
    every single report generation.
    """
    return not _stale_cells(ws, checks)


def verify_recalculated(ws, checks):
    """
    Fail loudly if recalculation did not actually populate formula cells.

    `checks` is a list of (row, col, description) tuples for cells that must
    hold a non-None value once real formulas have been evaluated (e.g. the
    workbook's own Total-assets count, NPV+ count, a sample per-asset NPV
    cell). Raises RuntimeError naming exactly which cells are still blank
    rather than letting the caller silently fall back to computing the value
    itself.
    """
    stale = [
        f"{desc} (row {row}, col {col})"
        for row, col, desc in _stale_cells(ws, checks)
    ]
    if stale:
        raise RuntimeError(
            "Workbook recalculation did not populate expected formula cells -- "
            "report generation cannot proceed without silently duplicating "
            "business calculations. Uncached cells:\n  " + "\n  ".join(stale)
        )


def xe(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def strip_ids(s):
    for attr in ['w14:paraId', 'w14:textId', 'w:rsidR', 'w:rsidRPr',
                 'w:rsidRDefault', 'w:rsidTr', 'w:rsidP', 'w:rsidDel']:
        s = re.sub(f' {re.escape(attr)}="[^"]*"', '', s)
    return s


def _para_text(para_xml):
    return ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', para_xml))


def _norm_dash(s):
    return re.sub(r'[–—‒\-]', '-', s)


def normalize_header(text):
    """Strip {{TOKEN}} markers and bare 2-5 letter currency codes, then lowercase."""
    text = re.sub(r'\{\{[A-Z0-9_]+\}\}', '', text)
    text = re.sub(r'\s*\([A-Z]{2,5}\)\s*', ' ', text)
    return re.sub(r'\s+', ' ', text).strip().lower()


def _get_rows(tbl_xml):
    """Return list of (start, end) char positions for every <w:tr> in tbl_xml."""
    rows, pos = [], 0
    while True:
        s = tbl_xml.find('<w:tr', pos)
        if s == -1:
            break
        e = tbl_xml.find('</w:tr>', s)
        if e == -1:
            break
        e += len('</w:tr>')
        rows.append((s, e))
        pos = e
    return rows


def _is_toc_para(para_xml):
    """True if this paragraph is a Table-of-Contents entry (w:pStyle TOC1/TOC2/...),
    not a real section heading — TOC entries repeat heading text and must be skipped."""
    return bool(re.search(r'<w:pStyle w:val="TOC\d*"', para_xml))


def locate_table(doc_xml, anchor_text):
    """Find <w:tbl> immediately following the first non-TOC paragraph containing
    anchor_text. Returns (start, end) char positions in doc_xml, or (None, None)."""
    target = _norm_dash(anchor_text).lower()
    for m in re.finditer(r'<w:p[ >].*?</w:p>', doc_xml, re.DOTALL):
        if _is_toc_para(m.group()):
            continue
        if target in _norm_dash(_para_text(m.group())).lower():
            ts = doc_xml.find('<w:tbl', m.end())
            if ts == -1:
                return None, None
            te = doc_xml.find('</w:tbl>', ts)
            if te == -1:
                return None, None
            return ts, te + len('</w:tbl>')
    return None, None


def build_header_map(tbl_xml):
    """Read header row (row 0) → {normalize_header(cell_text): col_index}."""
    rows = _get_rows(tbl_xml)
    if not rows:
        return {}
    rs, re_ = rows[0]
    cells = re.findall(r'<w:tc[ >].*?</w:tc>', tbl_xml[rs:re_], re.DOTALL)
    result = {}
    for i, cell in enumerate(cells):
        raw = ''.join(re.findall(r'<w:t[^>]*>([^<]*)</w:t>', cell))
        key = normalize_header(raw)
        if key:
            result[key] = i
    return result


def _set_cell_text(cell_xml, new_val):
    """Replace all <w:t> runs in a cell: first gets new_val, rest cleared."""
    done = [False]

    def rep(m):
        if not done[0]:
            done[0] = True
            return m.group(1) + xe(new_val) + m.group(3)
        return m.group(1) + m.group(3)

    return re.sub(r'(<w:t[^>]*>)([^<]*)(</w:t>)', rep, cell_xml)


def build_row(tmpl_xml, header_map, field_map, asset):
    """Clone the ROW_TEMPLATE row and fill cells by header-mapped column index."""
    row = strip_ids(tmpl_xml)
    col_vals = {header_map[k]: fn(asset) for k, fn in field_map.items() if k in header_map}
    spans = [(m.start(), m.end()) for m in re.finditer(r'<w:tc[ >].*?</w:tc>', row, re.DOTALL)]
    if not spans:
        return row
    out, prev = [], 0
    for i, (ts, te) in enumerate(spans):
        out.append(row[prev:ts])
        cell = row[ts:te]
        if i in col_vals:
            cell = _set_cell_text(cell, col_vals[i])
        out.append(cell)
        prev = te
    out.append(row[prev:])
    return ''.join(out)


def validate_headers(header_map, required, tname):
    missing = [h for h in required if h not in header_map]
    if missing:
        print(f"ERROR: Table '{tname}' is missing required header columns:")
        for h in missing:
            print(f"  '{h}'")
        print(f"  Detected headers: {sorted(header_map)}")
        sys.exit(1)


def rebuild_table(tbl_xml, asset_list, header_map, field_map, n_totals):
    """Remove ROW_TEMPLATE + sample rows, clone ROW_TEMPLATE once per asset.
    n_totals: number of trailing rows to preserve as totals (already scalar-filled).

    Templates that carry an explicit {{ROW_TEMPLATE}} marker row use that row as
    the clone source. Templates with no marker at all (e.g. a real, previously
    filled-in sample report used as-is, with no tokens anywhere) fall back to
    row index 1 - the first data row right after the header row - as the clone
    source instead. This fallback only engages when no marker is present, so it
    changes nothing for templates that already have one.
    """
    rows = _get_rows(tbl_xml)
    tmpl_idx = next(
        (i for i, (rs, re_) in enumerate(rows) if '{{ROW_TEMPLATE}}' in tbl_xml[rs:re_]),
        None
    )
    if tmpl_idx is None:
        tmpl_idx = 1 if len(rows) > 1 else None
    if tmpl_idx is None:
        return tbl_xml
    tmpl_xml = tbl_xml[rows[tmpl_idx][0]:rows[tmpl_idx][1]]
    before = tbl_xml[:rows[tmpl_idx][0]]
    if n_totals > 0 and len(rows) > n_totals:
        after = tbl_xml[rows[-n_totals][0]:]
    else:
        after = tbl_xml[rows[-1][1]:]
    new_rows = ''.join(build_row(tmpl_xml, header_map, field_map, a) for a in asset_list)
    return before + new_rows + after


def _norm_ws(s):
    return re.sub(r'\s+', ' ', s).strip()


def replace_literal_paragraphs(xml, replacements, warn_ambiguous=True):
    """Substitute scalar values in a template that has no {{TOKEN}} markers -
    i.e. a real, previously filled-in sample report used as-is.

    `replacements` is a list of (old_text, new_text) pairs. Each old_text is
    matched against the *whole* text of one <w:p>...</w:p> paragraph (all its
    runs concatenated, whitespace-normalized) - not a raw substring search -
    so a value that happens to also appear inside unrelated running text
    elsewhere in the document is not touched. When a paragraph matches, every
    <w:t> run inside it is replaced the same way build_row()/_set_cell_text()
    already do: the first run gets the new value, the rest are cleared. This
    keeps the surrounding run/paragraph XML (fonts, colors, text-box shapes,
    table-cell shading) completely untouched.

    Matches inside <w:txbxContent> (floating text boxes, e.g. a number
    overlaid on a donut-chart image) are handled identically, since they are
    just <w:p> elements at a different nesting depth.

    Prints a warning (does not raise) if an old_text is not found, or is found
    in more than one paragraph - callers should treat that as a signal to
    re-check the template rather than trust the substitution blindly.
    """
    for old_text, new_text in replacements:
        target = _norm_ws(old_text)
        matches = [m for m in re.finditer(r'<w:p[ >].*?</w:p>', xml, re.DOTALL)
                   if _norm_ws(_para_text(m.group())) == target]
        if not matches:
            if warn_ambiguous:
                print(f"  WARNING: literal value not found, left unchanged: {old_text!r}")
            continue
        if len(matches) > 1 and warn_ambiguous:
            print(f"  WARNING: literal value found in {len(matches)} places, "
                  f"replacing all: {old_text!r}")
        # Replace from the end so earlier match spans stay valid.
        for m in reversed(matches):
            new_para = _set_cell_text(m.group(), new_text)
            xml = xml[:m.start()] + new_para + xml[m.end():]
    return xml


def replace_literal_runs(xml, replacements, warn_ambiguous=True):
    """Like replace_literal_paragraphs, but matches a single <w:t>...</w:t>
    run's own exact text (not a whole paragraph). Use this for a cell/text-box
    that mixes a scalar value and static caption text in one paragraph across
    several runs (e.g. "21", "/", "45", " motors ", caption...) - replacing at
    the paragraph level would clobber the caption's own run; replacing at the
    run level touches only the run(s) that match.
    """
    for old_text, new_text in replacements:
        pattern = re.compile(r'(<w:t[^>]*>)' + re.escape(old_text) + r'(</w:t>)')
        matches = list(pattern.finditer(xml))
        if not matches:
            if warn_ambiguous:
                print(f"  WARNING: literal run not found, left unchanged: {old_text!r}")
            continue
        if len(matches) > 1 and warn_ambiguous:
            print(f"  WARNING: literal run found {len(matches)} times, "
                  f"replacing all: {old_text!r}")
        xml = pattern.sub(lambda m: m.group(1) + xe(new_text) + m.group(2), xml)
    return xml


def replace_within_row(xml, row_anchor_text, replacements, warn_ambiguous=True):
    """Apply replace_literal_paragraphs, but scoped to only the <w:tr>...
    </w:tr> row(s) whose text contains row_anchor_text (every matching row,
    e.g. the same totals-row label repeated in both a Top-10 table and an
    All-Assets appendix table).

    Once Phase 2 has replaced every per-asset data row with this customer's
    real numbers, a bare totals-row value like "2.9" or "23%" is no longer
    safe to match document-wide with replace_literal_paragraphs - some other
    asset's own row could coincidentally have the exact same value in one of
    its cells. Anchoring to a totals row's own (distinctive) label text before
    substituting keeps the replacement scoped to just that row.
    """
    rows = _get_rows(xml)
    match_positions = [(rs, re_) for (rs, re_) in rows if row_anchor_text in xml[rs:re_]]
    if not match_positions:
        if warn_ambiguous:
            print(f"  WARNING: row anchor not found, left unchanged: {row_anchor_text!r}")
        return xml
    for rs, re_ in reversed(match_positions):
        new_row = replace_literal_paragraphs(xml[rs:re_], replacements, warn_ambiguous=warn_ambiguous)
        xml = xml[:rs] + new_row + xml[re_:]
    return xml


def replace_within_textboxes(xml, replacements, warn_ambiguous=True):
    """Apply replace_literal_paragraphs, but scoped to only the content of
    <w:txbxContent>...</w:txbxContent> blocks (floating text boxes, e.g. a
    number overlaid on a donut-chart image).

    A bare value like "28%" is too generic to match document-wide once Phase 2
    has written this customer's own per-asset data - some other asset's own
    percentage column could coincidentally be the same value. Text boxes are
    never used for per-asset table cells in this template family, so scoping
    to just their content rules that collision out entirely.

    Each text box normally holds only one of several replacements, so "not
    found in this block" is expected for most (old_text, block) pairs - only
    a value that is missing from every block is worth warning about.
    """
    blocks = list(re.finditer(r'<w:txbxContent>.*?</w:txbxContent>', xml, re.DOTALL))
    found = {old for old, _ in replacements
             for m in blocks if _norm_ws(old) in _norm_ws(_para_text(m.group()))}
    if warn_ambiguous:
        for old, _ in replacements:
            if old not in found:
                print(f"  WARNING: literal value not found in any text box, left unchanged: {old!r}")
    for m in reversed(blocks):
        new_block = replace_literal_paragraphs(m.group(), replacements, warn_ambiguous=False)
        xml = xml[:m.start()] + new_block + xml[m.end():]
    return xml


def find_para_start(xml, search_text, after=0):
    """Return the start of the <w:p...> that contains search_text."""
    idx = xml.find(search_text, after)
    if idx == -1:
        return -1
    return xml.rfind('<w:p ', 0, idx)


def find_para_end(xml, para_start):
    """Return the char position just after the </w:p> that starts at para_start."""
    end = xml.find('</w:p>', para_start)
    return end + len('</w:p>') if end != -1 else -1


def trim_stale_appendix(doc_xml, before_anchor='Data is listed as total annual'):
    """Remove leftover duplicate Energy Savings / Application Details tables that
    some not-yet-cleaned templates still carry between the body 'Appendix' heading
    and 'Calculation Methodology'. New templates have no such duplicate and this
    is a no-op for them.
    """
    anchor_pos = doc_xml.rfind(before_anchor)
    if anchor_pos == -1:
        return doc_xml
    appendix_start = find_para_start(doc_xml, '>Appendix<', anchor_pos)
    appendix_end = find_para_end(doc_xml, appendix_start)
    if appendix_end <= 0:
        return doc_xml
    calcmeth_start = find_para_start(doc_xml, '>Calculation Methodology<', appendix_end)
    if calcmeth_start <= appendix_end:
        return doc_xml
    print(f"  Trimmed stale duplicate appendix tables ({calcmeth_start - appendix_end} chars)")
    return doc_xml[:appendix_end] + doc_xml[calcmeth_start:]


def render_table_section(doc_xml, name, anchor, asset_list, field_map, required, n_totals=0):
    """Single reusable renderer for a ranked-asset table.

    Locates the Energy Savings / Application Details section by heading text,
    takes the first table after that heading, validates its header row, and
    replaces its data rows with one row per asset in asset_list (top 10 or
    all — the caller decides by what it passes as asset_list).

    Returns the updated doc_xml.
    """
    ts, te = locate_table(doc_xml, anchor)
    if ts is None:
        print(f"  ERROR: Table '{name}' not found (anchor: '{anchor}')")
        sys.exit(1)

    tbl_xml = doc_xml[ts:te]
    hmap = build_header_map(tbl_xml)
    validate_headers(hmap, required, name)

    new_tbl = rebuild_table(tbl_xml, asset_list, hmap, field_map, n_totals)
    doc_xml = doc_xml[:ts] + new_tbl + doc_xml[te:]
    print(f"  {name}: {len(asset_list)} rows, {len(hmap)} columns mapped")
    return doc_xml
