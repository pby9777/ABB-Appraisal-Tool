"""
fill_saving_calculations.py
----------------------------
Fills one ABB Saving Calculations template from one or more zipped
EEA tool output folders. When a project has many assets the tool
produces multiple zip files — pass them all and the data is combined.

Usage:
    python fill_saving_calculations.py <template_xlsx> <zip1> [zip2 zip3 ...] [options]

Arguments:
    template_xlsx   The Saving Calculations template to fill
    zip(s)          One or more zipped tool output folders

Optional flags:
    --currency      Currency code, e.g. INR EUR USD NOK SAR SGD
    --tariff        Electricity tariff per kWh, e.g. 8 or 0.14
    --co2           CO2 intensity in kg/kWh, e.g. 0.54
    --tax           Corporate tax rate in %, e.g. 19  (stored as 0.19)
    --discount      Discount rate in %, e.g. 6.5  (stored as 0.065)

Output:
    Saved as <template_name>_filled.xlsx in the same folder.

Examples:
    Single zip:
        python fill_saving_calculations.py Poland_template.xlsx report.zip ^
            --currency EUR --tariff 0.14 --co2 0.75 --tax 19 --discount 6.5

    Multiple zips (large project):
        python fill_saving_calculations.py Poland_template.xlsx ^
            report-part1.zip report-part2.zip report-part3.zip ^
            --currency EUR --tariff 0.14 --co2 0.75 --tax 19 --discount 6.5
"""

import sys
import os
import argparse
import zipfile
import tempfile
import shutil

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed.  Run:  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Canonical header variants
# ---------------------------------------------------------------------------
CANONICAL_HEADERS = {
    "num":          ["#"],
    "equip_id":     ["Customer Equipment Id"],
    "application":  ["Application"],
    "energy_kwh":   ["Annual Energy. Cons (kWh)", "Annual Energy Cons (kWh)",
                     "Annual Energy Cons. (kWh)"],
    "savings_kwh":  ["Annual Energy Savings, kWh", "Annual Energy Savings (kWh)",
                     "Annual Energy Savings,kWh"],
    "investment":   ["Investment",
                     "Investment,",
                     "Investment, Only Drive"],
    "ie_class":     ["Ie Eff Class", "IE Eff Class", "Ie Eff. Class"],
    "dol_vsd":      ["Dol Vsd", "DOL/VSD", "Dol Vsd / Connection",
                     "Dol Vsd/Connection"],
    "flow_control": ["Flow Control"],
    "output_kw":    ["Output (kW)"],
    "shaft_height": ["Shaft height (Frame)", "Shaft Height (Frame)",
                     "Shaft height(Frame)"],
    "run_hours":    ["Annual Running Hours", "Running Hours",
                     "Annual Running Hours [h]"],
    "avg_loading":  ["Average Loading", "Average Loading ",
                     "Average flow", "Avg Loading", "Average Flow"],
    "avg_freq":     ["Average Freqency", "Average Frequency",
                     "Avg Freqency", "Avg Frequency"],
    "ess_motor":    ["Recommended ESS motor", "Recommended ESS Motor"],
    "ess_drive":    ["ESS connection", "ESS Connection"],
}

ASSUMPTION_LABELS = {
    "currency": ("Currency",          3),
    "tariff":   ("Electricity Price", 4),
    "co2":      ("Carbon Intensity",  4),
    "discount": ("Discount Rate",     4),
    "tax":      ("Tax Rate",          4),
}

SCAN_LIMIT = 1100


# ---------------------------------------------------------------------------
# Template helpers
# ---------------------------------------------------------------------------

def find_saving_sheets(wb):
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, 51):
            if ws.cell(row=r, column=2).value == "#":
                sheets.append(ws)
                break
    if not sheets:
        raise ValueError(
            f"No saving-calculations sheet found.\n"
            f"Available sheets: {wb.sheetnames}"
        )
    return sheets


def find_header_row(ws):
    for r in range(1, 51):
        if ws.cell(row=r, column=2).value == "#":
            return r
    raise ValueError(f"Header row not found in '{ws.title}'.")


def build_col_map(ws, header_row):
    header_lookup = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            header_lookup[str(cell.value).strip()] = cell.column

    col_map, missing = {}, []
    for field, variants in CANONICAL_HEADERS.items():
        for v in variants:
            if v.strip() in header_lookup:
                col_map[field] = header_lookup[v.strip()]
                break
        else:
            missing.append(field)

    if missing:
        print(f"      Skipped (column not in this template): {missing}")
    return col_map


def find_assumption_cells(ws):
    result = {}
    for r in range(1, 21):
        label = ws.cell(row=r, column=2).value
        if label is None:
            continue
        for key, (target, val_col) in ASSUMPTION_LABELS.items():
            if str(label).strip() == target:
                result[key] = (r, val_col)
    return result


def find_last_filled_row(ws, num_col, data_start):
    last = data_start - 1
    for r in range(data_start, SCAN_LIMIT + 1):
        if ws.cell(row=r, column=num_col).value is not None:
            last = r
        else:
            break
    return last


# ---------------------------------------------------------------------------
# Source data readers  (support multiple zips combined)
# ---------------------------------------------------------------------------

def find_file(directory, keyword):
    for root, _, files in os.walk(directory):
        for f in files:
            if keyword.lower() in f.lower() and f.lower().endswith(".xlsx"):
                return os.path.join(root, f)
    return None


def read_assessment_data(path):
    """Read assessment data xlsx. Returns list of asset dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        records.append({
            "num":         int(row[0]),
            "equip_id":    row[1],
            "energy_kwh":  row[2],
            "savings_kwh": row[6],
            "investment":  row[10],
            "ess_motor":   row[13],
            "ess_drive":   row[14],
        })
    return records


def read_input_assets(path):
    """Read input assets xlsx. Returns dict keyed by equip_id."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    assets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        shaft_raw    = row[5]
        shaft_height = int(shaft_raw) if shaft_raw and int(shaft_raw) > 0 else 0
        assets[row[1]] = {
            "application":  row[2],
            "dol_vsd":      row[3],
            "flow_control": row[4],
            "shaft_height": shaft_height,
            "output_kw":    row[6],
            "run_hours":    row[8],
            "ie_class":     row[9],
            "avg_loading":  row[15],
            "avg_freq":     row[16],
        }
    return assets


def load_all_zips(zip_paths):
    """
    Extract and combine data from one or more zip files.
    Returns (assessment_list, input_assets_dict).
    Assets are re-numbered sequentially across all zips.
    """
    all_assessment  = []
    all_input_assets = {}
    running_offset  = 0

    tmp_dirs = []
    try:
        for zip_path in zip_paths:
            tmp = tempfile.mkdtemp()
            tmp_dirs.append(tmp)

            print(f"  Extracting : {os.path.basename(zip_path)}")
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(tmp)

            assessment_path = find_file(tmp, "assessment data")
            input_path      = find_file(tmp, "input assets")

            if not assessment_path:
                print(f"  ERROR: 'assessment data' xlsx not found in {os.path.basename(zip_path)}")
                sys.exit(1)
            if not input_path:
                print(f"  ERROR: 'input assets' xlsx not found in {os.path.basename(zip_path)}")
                sys.exit(1)

            print(f"  Assessment : {os.path.basename(assessment_path)}")
            print(f"  Input data : {os.path.basename(input_path)}")

            records = read_assessment_data(assessment_path)
            assets  = read_input_assets(input_path)

            # Re-number sequentially if multiple zips
            for rec in records:
                rec["num"] = running_offset + rec["num"]
            running_offset += len(records)

            all_assessment.extend(records)
            all_input_assets.update(assets)
            print(f"  Assets in this zip: {len(records)}")

    finally:
        for td in tmp_dirs:
            shutil.rmtree(td, ignore_errors=True)

    return all_assessment, all_input_assets


# ---------------------------------------------------------------------------
# Fill one sheet
# ---------------------------------------------------------------------------

def fill_sheet(ws, assessment, input_assets, args):
    header_row  = find_header_row(ws)
    data_start  = header_row + 1
    col_map     = build_col_map(ws, header_row)
    assumptions = find_assumption_cells(ws)

    print(f"    Header row {header_row}, data from row {data_start}, "
          f"{len(col_map)}/{len(CANONICAL_HEADERS)} columns mapped")

    # Assumption cells
    def set_val(key, raw, divisor=1):
        if raw is None or key not in assumptions:
            return
        r, c = assumptions[key]
        ws.cell(row=r, column=c).value = round(raw / divisor, 6) if divisor != 1 else raw

    if args.currency is not None and "currency" in assumptions:
        r, c = assumptions["currency"]
        ws.cell(row=r, column=c).value = args.currency.upper()
    set_val("tariff",   args.tariff)
    set_val("co2",      args.co2)
    set_val("discount", args.discount, divisor=100)
    set_val("tax",      args.tax,      divisor=100)

    # Clear old data
    num_col     = col_map.get("num", 2)
    last_filled = find_last_filled_row(ws, num_col, data_start)
    clear_end   = max(last_filled, data_start + len(assessment) - 1)
    for r in range(data_start, clear_end + 1):
        for col in col_map.values():
            ws.cell(row=r, column=col).value = None

    # Write new data
    def w(row, field, value):
        col = col_map.get(field)
        if col is not None:
            ws.cell(row=row, column=col).value = value

    for i, asset in enumerate(assessment):
        r        = data_start + i
        equip_id = asset["equip_id"]
        inp      = input_assets.get(equip_id, {})

        w(r, "num",          asset["num"])
        w(r, "equip_id",     equip_id)
        w(r, "application",  inp.get("application", ""))
        w(r, "energy_kwh",   asset["energy_kwh"])
        w(r, "savings_kwh",  asset["savings_kwh"])
        w(r, "investment",   asset["investment"])
        w(r, "ie_class",     inp.get("ie_class", ""))
        w(r, "dol_vsd",      inp.get("dol_vsd", ""))
        w(r, "flow_control", inp.get("flow_control", ""))
        w(r, "output_kw",    inp.get("output_kw", ""))
        w(r, "shaft_height", inp.get("shaft_height", 0))
        w(r, "run_hours",    inp.get("run_hours", ""))
        w(r, "avg_loading",  inp.get("avg_loading", ""))
        w(r, "avg_freq",     inp.get("avg_freq", ""))
        w(r, "ess_motor",    asset["ess_motor"])
        w(r, "ess_drive",    asset["ess_drive"])

    return len(assessment)


# ---------------------------------------------------------------------------
# Fill all saving sheets in the workbook
# ---------------------------------------------------------------------------

def fill_template(template_path, assessment, input_assets, args):
    wb     = openpyxl.load_workbook(template_path)
    sheets = find_saving_sheets(wb)

    print(f"\n  Template : {os.path.basename(template_path)}")
    print(f"  Sheets   : {[ws.title for ws in sheets]}")

    for ws in sheets:
        print(f"\n  → '{ws.title}'")
        count = fill_sheet(ws, assessment, input_assets, args)
        print(f"    {count} assets written")

    if getattr(args, "output", None):
        # Use custom name; ensure .xlsx extension
        out_name = args.output
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"
        # Save in same folder as template
        template_dir = os.path.dirname(os.path.abspath(template_path))
        output_path  = os.path.join(template_dir, out_name)
    else:
        base, ext   = os.path.splitext(template_path)
        output_path = f"{base}_filled{ext}"
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Argument parser
# ---------------------------------------------------------------------------

def build_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Fill one ABB Saving Calculations template from one or more "
            "zipped EEA tool output folders."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Single zip:
    python fill_saving_calculations.py Poland_template.xlsx report.zip ^
        --currency EUR --tariff 0.14 --co2 0.75 --tax 19 --discount 6.5

  Multiple zips (large project split across several outputs):
    python fill_saving_calculations.py France_template.xlsx ^
        report-part1.zip report-part2.zip report-part3.zip ^
        --currency EUR --tariff 0.10 --co2 0.053 --tax 25 --discount 6.5
        """
    )
    parser.add_argument("template_xlsx",
                        help="Template Excel file to fill")
    parser.add_argument("zip_files",
                        nargs="+",
                        help="One or more zipped tool output folders")
    parser.add_argument("-o", "--output", metavar="FILENAME",
                        help="Custom output file name  (default: <template>_filled.xlsx)")
    parser.add_argument("--currency", metavar="CODE",
                        help="Currency code: INR EUR USD NOK SAR SGD ...")
    parser.add_argument("--tariff",   type=float, metavar="VALUE",
                        help="Electricity tariff per kWh")
    parser.add_argument("--co2",      type=float, metavar="VALUE",
                        help="CO2 intensity in kg/kWh")
    parser.add_argument("--tax",      type=float, metavar="PCT",
                        help="Corporate tax rate in %%  (e.g. 19 → stored 0.19)")
    parser.add_argument("--discount", type=float, metavar="PCT",
                        help="Discount rate in %%  (e.g. 6.5 → stored 0.065)")
    return parser


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = build_parser()
    args   = parser.parse_args()

    if not os.path.isfile(args.template_xlsx):
        print(f"ERROR: Template not found: {args.template_xlsx}")
        sys.exit(1)

    missing = [z for z in args.zip_files if not os.path.isfile(z)]
    if missing:
        for z in missing:
            print(f"ERROR: Zip file not found: {z}")
        sys.exit(1)

    print(f"\nTemplate     : {os.path.basename(args.template_xlsx)}")
    print(f"Zip file(s)  : {len(args.zip_files)}")

    print("\nAssumptions to apply:")
    print(f"  Currency      : {args.currency             if args.currency  is not None else '(unchanged)'}")
    print(f"  Tariff        : {args.tariff               if args.tariff    is not None else '(unchanged)'}")
    print(f"  CO2 intensity : {args.co2                  if args.co2       is not None else '(unchanged)'}")
    print(f"  Tax rate      : {str(args.tax)+'%'         if args.tax       is not None else '(unchanged)'}")
    print(f"  Discount rate : {str(args.discount)+'%'    if args.discount  is not None else '(unchanged)'}")

    print(f"\nLoading data from zip(s)...")
    assessment, input_assets = load_all_zips(args.zip_files)
    print(f"\nTotal assets combined: {len(assessment)}")

    output_path = fill_template(args.template_xlsx, assessment, input_assets, args)

    print(f"\n{'='*60}")
    print(f"Done.  Output saved: {output_path}")

    print(f"\n{'#':<4} {'Equipment ID':<50} {'kWh':>12} {'Savings kWh':>12} {'Investment':>14}")
    print("-" * 96)
    for a in assessment:
        print(
            f"{a['num']:<4} {str(a['equip_id'])[:49]:<50} "
            f"{a['energy_kwh']:>12,.0f} {a['savings_kwh']:>12,.0f} "
            f"{a['investment']:>14,.0f}"
        )


if __name__ == "__main__":
    main()


# ---------------------------------------------------------------------------
# Public API  (called by Flask web app)
# ---------------------------------------------------------------------------

class _Args:
    """Simple namespace so fill_sheet() works without argparse."""
    def __init__(self, currency=None, tariff=None, co2=None,
                 tax=None, discount=None, output=None):
        self.currency = currency
        self.tariff   = tariff
        self.co2      = co2
        self.tax      = tax
        self.discount = discount
        self.output   = output


def run_fill(template_path, zip_paths, output_path,
             currency=None, tariff=None, co2=None,
             tax=None, discount=None):
    """
    Fill template_path using data from zip_paths, save to output_path.
    Returns (asset_count, sheet_names).
    """
    args = _Args(currency=currency, tariff=tariff, co2=co2,
                 tax=tax, discount=discount, output=output_path)

    assessment, input_assets = load_all_zips(zip_paths)

    wb     = openpyxl.load_workbook(template_path)
    sheets = find_saving_sheets(wb)
    for ws in sheets:
        fill_sheet(ws, assessment, input_assets, args)
    wb.save(output_path)
    return len(assessment), [ws.title for ws in sheets]
