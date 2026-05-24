#!/usr/bin/env python3
"""
ABB Energy Appraisal – Poland (White Certificates) Report Generator
Template: EA_Report_Template_Poland.docx

Usage:
  python generate_report_poland.py  <saving_calc.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]

Requirements:
  pip install openpyxl

Poland-specific features vs Standard:
  - Summary payback = payback including White Certificates (WC)
  - Top-N table: 13 columns — WC Earnings + WC Payback (no Take-Back ✓)
  - Cover page text stored as split XML runs → handled with nth_wt()
  - WC Incentive shown on summary business case row
  - Sensitivity chart: WC-aware fallback formula when Excel cells are errors
  - Adaptive top-N padding: pads to min(10, NA) rows
  - Appendix/Det Rec removed when NA < 10
"""

import sys, os, re, math, io, zipfile
from datetime import datetime
import openpyxl

# ── CLI ───────────────────────────────────────────────────────────────────────
args = sys.argv[1:]
if len(args) < 3:
    print('Usage: python generate_report_poland.py  <excel.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]')
    sys.exit(1)

XLSX_PATH   = args[0]
CUSTOMER    = args[1].upper()
PLANT       = args[2]
RPT_DATE    = args[3] if len(args) > 3 else datetime.now().strftime("%m.%d.%Y")
DATA_SOURCE = args[4] if len(args) > 4 else "Customer Input"

_script_dir = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = next(
    (os.path.join(_script_dir, n) for n in ['EA_Report_Template_Poland.docx']
     if os.path.exists(os.path.join(_script_dir, n))), None)
if TEMPLATE_PATH is None:
    print("ERROR: EA_Report_Template_Poland.docx not found in script folder")
    sys.exit(1)

for path, label in [(XLSX_PATH, "Excel file"), (TEMPLATE_PATH, "Template")]:
    if not os.path.exists(path):
        print(f"ERROR: {label} not found: {path}")
        sys.exit(1)

def date_to_iso(d):
    """Convert any date string to YYYY-MM-DD."""
    for fmt in ('%m.%d.%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(d, fmt).strftime('%Y-%m-%d')
        except: pass
    return d  # fallback: return as-is

RPT_DATE_ISO = date_to_iso(RPT_DATE)

# ── Read Excel ────────────────────────────────────────────────────────────────
print(f"Reading Excel: {os.path.basename(XLSX_PATH)}")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Accept any sheet matching Saving(s)_Calculation(s) with any version suffix
# Also handles 'Saving_Calculatios' (common typo in Poland files)
def _is_sc_sheet(name):
    n = name.strip().lower().replace(' ', '_')
    prefixes = ('saving_calculations', 'savings_calculations',
                'saving_calculatios',  'savings_calculatios')
    return any(n.startswith(p) for p in prefixes)

_sc_sheet = next((s for s in wb.sheetnames if _is_sc_sheet(s)), None)
if _sc_sheet is None:
    print(f"ERROR: No Saving_Calculations sheet found. Sheets: {wb.sheetnames}")
    sys.exit(1)
_sc_candidates = [s for s in wb.sheetnames if _is_sc_sheet(s)]
if len(_sc_candidates) > 1:
    import re as _re
    _vn = [s for s in _sc_candidates if _re.search(r'_v\d+$', s.strip().lower())]
    if _vn:
        _sc_sheet = _vn[0]
if _sc_sheet not in ("Saving_Calculations", "Savings_Calculations"):
    print(f"  Note: using sheet '{_sc_sheet}'")
ws = wb[_sc_sheet]

_XL_ERRORS = frozenset({
    # Excel formula error strings
    '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!', '#ERROR!',
    # Common "no value" placeholder strings used in Excel templates
    '-', '\u2013', '\u2014', 'N/A', 'n/a', 'NA', 'na', 'n.a.', 'N.A.',
})

def cv(row, col):
    """Read a cell value; return None for Excel error strings (#DIV/0!, #N/A, etc.)."""
    v = ws.cell(row=row, column=col).value
    return None if isinstance(v, str) and v.strip() in _XL_ERRORS else v

# ── Detect column positions from header row 36 ────────────────────────────────
HEADER_ROW = 36
col_map = {}
for c in range(1, 70):
    h = cv(HEADER_ROW, c)
    if h:
        col_map[str(h).strip().lower()] = c

def find_col(names, default):
    for n in names:
        if n.lower() in col_map:
            return col_map[n.lower()]
    return default

COL_NUM      = find_col(['#', '# no.', 'no.'], 2)
COL_E_CONS   = find_col(['annual energy. cons (kwh)', 'annual energy cons (kwh)'], 5)
COL_E_COST   = find_col(['annual energy cost'], 6)
COL_CO2_CONS = find_col(['annual co2 cons.'], 7)
COL_SAV_KWH  = find_col(['annual energy savings, kwh'], 8)
COL_SAV_COST = find_col(['annual energy cost savings'], 9)
COL_SAV_PCT  = find_col(['annual energy savings (%)'], 10)
COL_INVEST   = find_col(['investment'], 11)
# v2-style Excel: col 13 = 'Investment - CEE' (net), col 14 = payback, col 15 = CO2 sav
_hdr13 = str(ws.cell(HEADER_ROW, 13).value or '').lower()
_hdr12 = str(ws.cell(HEADER_ROW, 12).value or '').lower()
_IS_V2_EXCEL = 'cee' in _hdr12 or 'cee' in _hdr13
if _IS_V2_EXCEL:
    COL_INVEST_NET = find_col(['investment - cee', 'investment-cee'], 13)
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 14)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 15)
    print(f"  Detected v2-style Excel (CEE columns) — using net investment col {COL_INVEST_NET}")
else:
    COL_INVEST_NET = COL_INVEST
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 12)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 13)
COL_NPV      = find_col(['npv'], 17)
COL_IE       = find_col(['ie eff class', 'ie'], 41)
COL_LOAD     = find_col(['driven load', 'application'], 42)
COL_FLOW     = find_col(['flow control', 'flow control method'], 43)
COL_CONN     = find_col(['dol vsd', 'connection'], 44)
COL_OUTPUT   = find_col(['output (kw)', 'rated power, kw'], 45)
COL_SHAFT    = find_col(['shaft height (frame)', 'shaft height'], 46)
# Poland: 'annual running hours' is an extra alias; default col 46
COL_RUNHRS   = find_col(['running hours', 'annual running hours', 'running time (hours)'], 46)
COL_AVG      = find_col(['average flow % / average frequency (hz)',
                          'average flow', 'average frequency', 'avg. frequency'], 48)
COL_ESS      = find_col(['recommended ess motor'], 50)
COL_ESSC     = find_col(['ess connection'], 51)
# Poland: White Certificate columns
COL_WC_EARN    = find_col(['potential white certificate earnings', 'white certificate earnings'], 14)
COL_WC_PAYBACK = find_col(['effective payback with white certificates', 'payback with white'], 15)

# ── KPI values ────────────────────────────────────────────────────────────────
CURRENCY       = str(cv(3, 3) or "PLN").strip()
ELEC_PRICE     = cv(5, 4) or 8
CO2_INTENSITY  = cv(6, 4) or 0.54
DISCOUNT_RATE  = float(cv(7, 4) or 0.065)  # Poland Excel row 7 = discount rate
TAX_RATE       = cv(8, 4) or 0.349
_NPV_YEARS     = 20                         # standard NPV horizon
TOTAL_ASSETS   = int(cv(13, 3) or 0)
NPV_POS_CNT    = int(cv(14, 3) or 0)
ANNUAL_SAVINGS = cv(15, 3)
CONSUMP_BEFORE = cv(16, 3)
SAVINGS_KWH    = cv(17, 3)
NPV_VALUE      = cv(19, 3)
INVEST_COST    = cv(20, 3)
PAYBACK_TIME   = cv(21, 3)
CO2_SAVINGS    = cv(22, 3)
IRR_VALUE      = cv(23, 3)
BEV_COUNT      = cv(25, 4)
TOP10_PAYBACK  = cv(21, 4)
TOP10_NPV      = cv(19, 4)
TOP10_INVEST   = cv(20, 4)
TOP10_IRR      = cv(23, 4)
# Poland: WC payback from Excel rows 33 (Top-10) and 34 (NPV+), col 15
# Use cv() so Excel error strings (#DIV/0! etc.) are treated as None → 0
_WC_PB_TOP10 = float(cv(33, 15) or 0)
_WC_PB_NPVP  = float(cv(34, 15) or 0)

# Poland: WC rate per kWh — used to compute per-asset WC earnings when formula
# cells are uncached. Try common cells in the KPI section (rows 9-12, cols 3-5)
# and the WC rate per toe summary cell (row 33-34, col 3-4).
# Formula: WC Earnings = Energy Savings (kWh) / 11,630 (kWh/toe) * WC rate (PLN/toe)
# 11,630 kWh = 1 toe (Polish WC regulation conversion factor)
_TOE_PER_KWH = 1.0 / 11630.0
_WC_RATE_PLN_TOE = 0.0  # PLN per toe; 0 = unknown → will skip WC earn fallback
for _wr, _wc in [(9, 4), (10, 4), (9, 3), (10, 3), (11, 4), (12, 4)]:
    _v = ws.cell(_wr, _wc).value
    if _v is not None:
        try:
            _f = float(_v)
            # WC rates in Poland are typically 50–3000 PLN/toe
            if 50 <= _f <= 5000:
                _WC_RATE_PLN_TOE = _f
                print(f"  WC rate found at row {_wr} col {_wc}: {_f} PLN/toe")
                break
        except: pass

# Payback sensitivity (rows 19-23, col 20 = price delta, col 22 = payback)
SENSITIVITY = []
for r in range(19, 24):
    delta   = cv(r, 20)
    payback = cv(r, 22)
    try:
        SENSITIVITY.append((float(delta), float(payback)))
    except (TypeError, ValueError):
        pass  # skip None, '#DIV/0!' and other non-numeric Excel errors
_SENSITIVITY_PENDING = len(SENSITIVITY) < 5  # will recompute with WC formula if incomplete

# ── Load assets ───────────────────────────────────────────────────────────────
ASSET_START_ROW = 37
assets = []
# Use max(TOTAL_ASSETS, 0) + 30 as upper bound so all assets are found even when
# TOTAL_ASSETS is a formula cell returning None (openpyxl data_only limitation).
# When TOTAL_ASSETS is a formula cell returning None (openpyxl data_only limitation),
# fall back to a generous scan window. The loop breaks naturally on the first empty row.
_scan_limit = max(int(TOTAL_ASSETS or 0), 200)
for r in range(ASSET_START_ROW, ASSET_START_ROW + _scan_limit):
    num = cv(r, COL_NUM)
    if num is None: break
    try: num = int(float(num))
    except: break
    if not (1 <= num <= 9999): break

    ess_motor = str(cv(r, COL_ESS) or "")

    # Shaft height: use Excel column value — Poland template has no shaft height column
    shaft_raw = cv(r, COL_SHAFT)
    try:
        shaft_val = str(int(float(shaft_raw))) if shaft_raw is not None else ""
    except:
        shaft_val = str(shaft_raw) if shaft_raw else ""

    # Avg Flow% / Frequency: use col 48 directly
    avg_raw = cv(r, COL_AVG)
    if avg_raw is not None:
        try:
            v = float(avg_raw)
            avg_val = f"{round(v*100)}" if 0 < v <= 1 else str(avg_raw).replace(' ', '')
        except:
            avg_val = str(avg_raw).replace(' ', '')
    else:
        fm = re.search(r'([\d.]+)\s*Hz', ess_motor)
        avg_val = fm.group(1) if fm else ""

    # Connection: DOL when blank/None
    conn_raw = cv(r, COL_CONN)
    conn_val = str(conn_raw) if conn_raw else "DOL"

    assets.append({
        "num":        num,
        "tag":        str(cv(r, 3) or ""),
        "load":       str(cv(r, COL_LOAD) or cv(r, 4) or ""),
        "e_cons":     cv(r, COL_E_CONS)    or 0,
        "e_cost":     cv(r, COL_E_COST)    or 0,
        "co2_cons":   cv(r, COL_CO2_CONS)  or 0,
        "e_sav_kwh":  cv(r, COL_SAV_KWH)   or 0,
        "e_sav_cost": cv(r, COL_SAV_COST)  or 0,
        "e_sav_pct":  cv(r, COL_SAV_PCT)   or 0,
        "invest":     cv(r, COL_INVEST)    or 0,
        "payback":    cv(r, COL_PAYBACK)
                      if not isinstance(cv(r, COL_PAYBACK), str)
                      or cv(r, COL_PAYBACK).strip() not in ('', '-') else None,
        "co2_sav":    cv(r, COL_CO2_SAV)   or 0,
        "npv":        cv(r, COL_NPV),
        "ie":         str(cv(r, COL_IE)    or ""),
        "flow_ctrl":  str(cv(r, COL_FLOW)  or "Information not available"),
        "connection": conn_val,
        "output_kw":  str(cv(r, COL_OUTPUT) or ""),
        "shaft_h":    shaft_val,
        "run_hrs":    cv(r, COL_RUNHRS),
        "avg_val":    avg_val,
        "ess_motor":  ess_motor,
        "ess_conn":   str(cv(r, COL_ESSC)  or ""),
        "wc_earn":    float(cv(r, COL_WC_EARN)    or 0),
        "wc_payback": float(cv(r, COL_WC_PAYBACK) or 0),
    })

if not assets:
    print("ERROR: No assets found. Check TOTAL_ASSETS in Excel row 13.")
    sys.exit(1)

# ── Per-asset fallback: compute formula-dependent values when Excel hasn't cached them ──
# openpyxl data_only=True can only read cached values. If the workbook was saved
# without recalculating (manual calc mode), all formula cells return None.
# We recompute from the raw input values that are always present.
_needs_fallback = any(a["e_sav_cost"] == 0 and a["e_sav_kwh"] and a["npv"] is None
                      for a in assets)
if _needs_fallback:
    print("  Note: Excel formula cells uncached — computing e_cost/NPV/payback from raw inputs")
for a in assets:
    ec   = float(a["e_cons"]    or 0)
    esav = float(a["e_sav_kwh"] or 0)
    inv  = float(a["invest"]    or 0)
    ep   = float(ELEC_PRICE     or 8)
    co2i = float(CO2_INTENSITY  or 0.054)

    # Energy costs and savings (formula: e_cons/e_sav_kwh * elec_price)
    if not a["e_cost"]     and ec   > 0: a["e_cost"]     = ec   * ep
    if not a["e_sav_cost"] and esav > 0: a["e_sav_cost"] = esav * ep
    if not a["e_sav_pct"]  and ec   > 0: a["e_sav_pct"]  = esav / ec
    if not a["co2_cons"]   and ec   > 0: a["co2_cons"]   = ec   * co2i
    if not a["co2_sav"]    and esav > 0: a["co2_sav"]    = esav * co2i

    # NPV using annuity formula: PV(after-tax annual savings, n years) – investment
    # Matches Excel: =NPV(discount_rate, year1_cf:year20_cf) + year0_cf
    if a["npv"] is None:
        ann_sav = float(a["e_sav_cost"] or 0)
        if ann_sav > 0:
            r         = DISCOUNT_RATE
            after_tax = ann_sav * (1 - float(TAX_RATE or 0.25))
            pv_factor = ((1 - (1 + r) ** (-_NPV_YEARS)) / r) if r > 0 else _NPV_YEARS
            a["npv"]  = round(after_tax * pv_factor - inv, 2)

    # Simple payback = investment / annual savings (only meaningful when NPV > 0)
    if a["payback"] is None:
        ann_sav = float(a["e_sav_cost"] or 0)
        npv_val = float(a["npv"]) if a["npv"] is not None else -1
        if ann_sav > 0 and npv_val > 0:
            a["payback"] = round(inv / ann_sav, 2)

    # White Certificate fallback: if WC earn is 0 and we have a WC rate, compute it
    if not a.get("wc_earn") and _WC_RATE_PLN_TOE > 0 and esav > 0:
        a["wc_earn"] = round(esav * _TOE_PER_KWH * _WC_RATE_PLN_TOE, 2)
    # WC payback = invest / (annual savings + WC earnings)
    if not a.get("wc_payback"):
        wce = float(a.get("wc_earn") or 0)
        ann_sav = float(a["e_sav_cost"] or 0)
        if inv > 0 and (ann_sav + wce) > 0:
            a["wc_payback"] = round(inv / (ann_sav + wce), 2)

NA = len(assets)
npv_pos = [a for a in assets if a["npv"] is not None and float(a["npv"]) > 0]

# Recount NPV_POS_CNT and TOTAL_ASSETS from actual assets when Excel cells were None
if TOTAL_ASSETS == 0:
    TOTAL_ASSETS = NA
if NPV_POS_CNT == 0 and npv_pos:
    NPV_POS_CNT = len(npv_pos)
    print(f"  Note: NPV_POS_CNT recomputed from assets: {NPV_POS_CNT}")

if npv_pos:
    # Sort NPV+ by payback ascending
    npv_sorted = sorted(
        npv_pos,
        key=lambda a: (float(a["payback"]) if a["payback"] else 9999,
                       -(float(a["e_sav_cost"] or 0))))[:10]

    # Adaptive padding: when NA <= 10 show ALL assets; otherwise pad to 10
    target = 10 if NA > 10 else NA
    if len(npv_sorted) < target:
        npv_neg = [a for a in assets if a["npv"] is None or float(a["npv"] or 0) <= 0]
        npv_neg_sorted = sorted(npv_neg, key=lambda a: -(float(a["e_sav_cost"] or 0)))
        slots = target - len(npv_sorted)
        top10 = npv_sorted + npv_neg_sorted[:slots]
        print(f"  Top table: {len(npv_sorted)} NPV+ + {min(slots, len(npv_neg_sorted))} NPV- (white-shaded)")
    else:
        top10 = npv_sorted

    NO_NPV_MODE = False
else:
    # No NPV+ assets → use ALL assets sorted by energy cost savings descending
    top10 = sorted(assets, key=lambda a: -(float(a["e_sav_cost"] or 0)))[:10]
    NO_NPV_MODE = True
    print("  NOTE: No NPV-positive assets — using all assets sorted by energy savings")

NT = len(top10)
NA = len(assets)

# Sort ALL assets by serial number for appendix + details tables
assets_by_num = sorted(assets, key=lambda a: a["num"])

# Appendix logic
# NA > 10 → App Details + Details of Recommendation included in Appendix
# NA <= 10 → Appendix trimmed to Calculation Methodology + NPV only
SKIP_APP_DET = (NA <= 10)

# ── NO_NPV_MODE: re-derive summary KPIs from asset rows ──────────────────────
# The Excel "NPV+" summary cells (rows 15-25 cols 3-4) are zero/blank when ALL
# assets are NPV-negative. Compute directly from asset data so the summary page
# correctly shows cost savings, CO2 avoided, energy chart, and Vehicles.
# NOTE: Sensitivity fallback is handled separately below via _SENSITIVITY_PENDING.
if NO_NPV_MODE:
    ANNUAL_SAVINGS = sum(float(a["e_sav_cost"] or 0) for a in assets)
    CO2_SAVINGS    = sum(float(a["co2_sav"]    or 0) for a in assets)
    CONSUMP_BEFORE = sum(float(a["e_cons"]     or 0) for a in assets)
    SAVINGS_KWH    = sum(float(a["e_sav_kwh"]  or 0) for a in assets)
    BEV_COUNT      = round(float(SAVINGS_KWH) / 3500) if SAVINGS_KWH else 0
    print("  NO_NPV_MODE: summary KPIs computed from all assets (Excel NPV+ cells empty)")

# ── General summary KPI fallback (uncached Excel formula cells) ───────────────
# When Excel was saved without recalculating, the KPI summary rows (15-25) return
# None even when assets ARE NPV-positive. Recompute from asset data in that case.
# For NPV-positive mode we aggregate from npv_pos; for NO_NPV_MODE we use all assets
# (already filled above, so we only overwrite if still None).
_kpi_src = assets if NO_NPV_MODE else npv_pos
if ANNUAL_SAVINGS is None or float(ANNUAL_SAVINGS or 0) == 0:
    ANNUAL_SAVINGS = sum(float(a["e_sav_cost"] or 0) for a in _kpi_src)
    print(f"  KPI fallback: ANNUAL_SAVINGS = {ANNUAL_SAVINGS:,.0f}")
if CONSUMP_BEFORE is None or float(CONSUMP_BEFORE or 0) == 0:
    CONSUMP_BEFORE = sum(float(a["e_cons"] or 0) for a in _kpi_src)
if SAVINGS_KWH is None or float(SAVINGS_KWH or 0) == 0:
    SAVINGS_KWH = sum(float(a["e_sav_kwh"] or 0) for a in _kpi_src)
if CO2_SAVINGS is None or float(CO2_SAVINGS or 0) == 0:
    CO2_SAVINGS = sum(float(a["co2_sav"] or 0) for a in _kpi_src)
if INVEST_COST is None or float(INVEST_COST or 0) == 0:
    INVEST_COST = sum(float(a["invest"] or 0) for a in _kpi_src)
if BEV_COUNT is None or float(BEV_COUNT or 0) == 0:
    BEV_COUNT = round(float(SAVINGS_KWH or 0) / 3500)
if PAYBACK_TIME is None or float(PAYBACK_TIME or 0) == 0:
    _sav = float(ANNUAL_SAVINGS or 0)
    PAYBACK_TIME = round(float(INVEST_COST or 0) / _sav, 2) if _sav > 0 else None
if NPV_VALUE is None or float(NPV_VALUE or 0) == 0:
    NPV_VALUE = sum(float(a["npv"] or 0) for a in _kpi_src if a["npv"] is not None)
# Top-10 KPI fallback (col 4 in Excel)
if TOP10_NPV is None or float(TOP10_NPV or 0) == 0:
    TOP10_NPV = sum(float(a["npv"] or 0) for a in top10
                    if a["npv"] is not None and float(a["npv"]) > 0)
if TOP10_INVEST is None or float(TOP10_INVEST or 0) == 0:
    TOP10_INVEST = sum(float(a["invest"] or 0) for a in top10)
if TOP10_PAYBACK is None or float(TOP10_PAYBACK or 0) == 0:
    _t10s = sum(float(a["e_sav_cost"] or 0) for a in top10)
    TOP10_PAYBACK = round(float(TOP10_INVEST) / _t10s, 2) if _t10s > 0 else None

# ── IRR fallback (Newton-Raphson on flat-annuity cash flow) ──────────────────
def _compute_irr(invest, annual_cf, n_years):
    """Return IRR for cash flows: -invest at t=0, annual_cf at t=1..n_years.
    Uses Newton-Raphson iteration.  Returns None if solution not found."""
    if invest <= 0 or annual_cf <= 0 or n_years <= 0:
        return None
    r = annual_cf / invest  # initial guess ≈ 1/payback
    for _ in range(200):
        try:
            if abs(r) < 1e-10:
                break
            pow_neg_n  = (1 + r) ** (-n_years)
            pow_neg_n1 = (1 + r) ** (-n_years - 1)
            annuity    = (1 - pow_neg_n) / r
            npv_val    = -invest + annual_cf * annuity
            # d(annuity)/dr = [n*(1+r)^(-n-1)*r + (1+r)^-n - 1] / r²
            d_annuity  = (n_years * pow_neg_n1 * r + pow_neg_n - 1) / (r * r)
            d_npv      = annual_cf * d_annuity
            if abs(d_npv) < 1e-12:
                break
            r_new = r - npv_val / d_npv
            if abs(r_new - r) < 1e-9:
                r = r_new
                break
            r = r_new
        except (ValueError, ZeroDivisionError, OverflowError):
            break
    return round(r, 6) if 0 < r < 10 else None  # sanity: IRR 0%–1000%

if TOP10_IRR is None or float(TOP10_IRR or 0) == 0:
    _t10i = float(TOP10_INVEST or 0)
    _t10s = sum(float(a["e_sav_cost"] or 0) for a in top10)
    # Poland: include WC earnings in annual benefit for IRR calculation
    _t10wc = sum(float(a.get("wc_earn") or 0) for a in top10)
    _irr   = _compute_irr(_t10i, _t10s + _t10wc, _NPV_YEARS)
    if _irr is not None:
        TOP10_IRR = _irr
        print(f"  IRR fallback (Top-{NT}): {_irr:.1%}")

if IRR_VALUE is None or float(IRR_VALUE or 0) == 0:
    _src_irr = assets if NO_NPV_MODE else npv_pos
    _all_s   = sum(float(a["e_sav_cost"] or 0) for a in _src_irr)
    _all_wc  = sum(float(a.get("wc_earn") or 0) for a in _src_irr)
    _all_i   = float(INVEST_COST or 0)
    _irr_all = _compute_irr(_all_i, _all_s + _all_wc, _NPV_YEARS)
    if _irr_all is not None:
        IRR_VALUE = _irr_all

# ── WC-aware sensitivity fallback ────────────────────────────────────────────
# If fewer than 5 sensitivity rows loaded from Excel, recompute using WC formula:
# payback_with_wc(delta) = invest / (savings*(1+delta) + wc_earnings)
if _SENSITIVITY_PENDING and npv_pos:
    _ns  = sum(float(a.get('e_sav_cost') or 0) for a in npv_pos)
    _ni  = sum(float(a.get('invest')     or 0) for a in npv_pos)
    _nwc = sum(a.get('wc_earn', 0)             for a in npv_pos)
    if _ns > 0:
        SENSITIVITY = [
            (_d, round(_ni / ((_ns * (1 + _d)) + _nwc), 4))
            for _d in [-0.2, -0.1, 0.0, 0.1, 0.2]
            if (_ns * (1 + _d)) + _nwc > 0
        ]
        print(f"  Sensitivity (WC-aware fallback): {[round(s[1],2) for s in SENSITIVITY]}")
elif _SENSITIVITY_PENDING and NO_NPV_MODE:
    # All NPV-negative: standard fallback without WC
    _total_invest = sum(float(a["invest"] or 0) for a in assets)
    _base_savings = float(ANNUAL_SAVINGS or 0)
    if _base_savings > 0:
        SENSITIVITY = [
            (_d, _total_invest / (_base_savings * (1 + _d)))
            for _d in [-0.2, -0.1, 0.0, 0.1, 0.2]
            if _base_savings * (1 + _d) > 0
        ]
        print(f"  Sensitivity (NO_NPV fallback): {[round(s[1],2) for s in SENSITIVITY]}")

print(f"  {NA} assets, {len(npv_pos)} NPV+, {NT} in Top-{NT} "
      f"{'(no NPV+ mode)' if NO_NPV_MODE else '(sorted by payback)'}")

# ── Formatting helpers ────────────────────────────────────────────────────────
SYM = CURRENCY

def fmt(n, dp=0):
    if n is None or (isinstance(n, float) and math.isnan(n)): return "\u2013"
    try:
        n = float(n)
        s = f"{abs(n):,.{dp}f}"
        return s if n >= 0 else f"-{s}"
    except: return str(n)

def fmt_parens(n):
    if n is None: return "\u2013"
    try:
        n = float(n)
        return f"({fmt(abs(n))})" if n < 0 else fmt(n)
    except: return str(n)

def fmtpct(n):
    if n is None: return "\u2013"
    v = round(float(n) * 100)
    return f"{v}%" if v >= 0 else f"-{abs(v)}%"

def fmtyrs(n):
    if n is None: return ""
    try: return f"{float(n):.1f}"
    except: return ""

def pct_str(n):
    if n is None: return "\u2013"
    return f"{round(float(n)*100, 1)}%"

def fmtirr(n):
    """Format IRR safely — Excel may store it as '-' string when undefined."""
    if n is None: return "\u2013"
    try: return f"{round(float(n)*100)}%"
    except: return str(n)

def _fmtK(v):
    """Compact K/M format for summary page large numbers."""
    v = float(v or 0)
    if abs(v) >= 1_000_000: return f"{v/1_000_000:.1f}M"
    if abs(v) >= 1_000:     return f"{v/1_000:.1f}K"
    return f"{v:.0f}"

# ── Top-N totals ──────────────────────────────────────────────────────────────
t10_e_cons  = sum(float(a["e_cons"]     or 0) for a in top10)
t10_e_cost  = sum(float(a["e_cost"]     or 0) for a in top10)
t10_co2_c   = sum(float(a["co2_cons"]   or 0) for a in top10)
t10_sav_kwh = sum(float(a["e_sav_kwh"]  or 0) for a in top10)
t10_invest  = sum(float(a["invest"]     or 0) for a in top10)
t10_sav_c   = sum(float(a["e_sav_cost"] or 0) for a in top10)
t10_co2_sav = sum(float(a["co2_sav"]    or 0) for a in top10)
t10_pct     = t10_sav_kwh / t10_e_cons if t10_e_cons else 0

t10_simple_payback = t10_invest / t10_sav_c if t10_sav_c else 0
# Only sum NPV+ assets — NPV- padding must not drag total negative and
# falsely trigger "no NPV" layout on the summary page
t10_npv = sum(float(a["npv"] or 0) for a in top10
              if a["npv"] is not None and float(a["npv"]) > 0)

# TOP10_PAYBACK for top-table total row; PAYBACK_TIME for summary page
TOP10_PAYBACK = t10_simple_payback if t10_simple_payback > 0 else TOP10_PAYBACK
TOP10_NPV     = t10_npv
TOP10_INVEST  = t10_invest

# ── XML helpers ───────────────────────────────────────────────────────────────
def xe(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

def get_tr_pos(xml):
    pos = []
    for m in re.finditer(r'<w:tr[ >]', xml):
        s = m.start()
        e = xml.find('</w:tr>', s)
        if e == -1: continue
        pos.append((s, e + len('</w:tr>')))
    return pos

def nth_wt(xml_str, n, new_val):
    ms = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', xml_str))
    if n >= len(ms): return xml_str
    m = ms[n]
    return xml_str[:m.start()] + m.group(1) + xe(new_val) + m.group(3) + xml_str[m.end():]

def strip_ids(s):
    for attr in ['w14:paraId', 'w14:textId', 'w:rsidR', 'w:rsidRPr',
                 'w:rsidRDefault', 'w:rsidTr', 'w:rsidP', 'w:rsidDel']:
        s = re.sub(f' {re.escape(attr)}="[^"]*"', '', s)
    return s

def fill_row(tmpl, vals):
    row  = strip_ids(tmpl)
    ms   = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', row))
    vals = list(vals)[:len(ms)]
    for i, v in reversed(list(enumerate(vals))):
        m   = ms[i]
        row = row[:m.start()] + m.group(1) + xe(v) + m.group(3) + row[m.end():]
    return row

def repl(xml, old, new):
    o, n = xe(old), xe(new)
    xml = xml.replace(f'<w:t>{o}</w:t>', f'<w:t>{n}</w:t>')
    xml = xml.replace(f'<w:t xml:space="preserve">{o}</w:t>',
                      f'<w:t xml:space="preserve">{n}</w:t>')
    return xml

def _get_tc_pos(row_xml):
    """Return list of (start, end) positions for each <w:tc>...</w:tc> in a row."""
    positions = []
    for m in re.finditer(r'<w:tc(?:\s[^>]*)?>',  row_xml):
        close = row_xml.find('</w:tc>', m.end())
        if close == -1: break
        positions.append((m.start(), close + 7))  # 7 = len('</w:tc>')
    return positions

def _insert_cell_after(row_xml, after_idx, header_text=None):
    """Clone the cell at after_idx, replace its <w:t> text with header_text (or
    empty string), and insert the clone immediately after the source cell.
    """
    tcs = _get_tc_pos(row_xml)
    if after_idx >= len(tcs):
        return row_xml
    src_s, src_e = tcs[after_idx]
    cell = strip_ids(row_xml[src_s:src_e])
    # Collapse all <w:t>...</w:t> nodes to a single placeholder, then set text
    # Step 1: remove every <w:t> node
    cell_no_t = re.sub(r'<w:t[^>]*>[^<]*</w:t>', '', cell)
    # Step 2: find the first <w:r> closing tag and inject a single <w:t> before it
    new_text = xe(header_text) if header_text is not None else ''
    cell = cell_no_t.replace('</w:r>', f'<w:t>{new_text}</w:t></w:r>', 1)
    insert_pos = tcs[after_idx][1]
    return row_xml[:insert_pos] + cell + row_xml[insert_pos:]

def repl_in_file(content, old, new):
    return content.replace(xe(old), xe(new)).replace(old, new)

def find_para_start(xml, search_text, after=0):
    idx = xml.find(search_text, after)
    if idx == -1: return -1
    return xml.rfind('<w:p ', 0, idx)

def find_para_end(xml, para_start):
    end = xml.find('</w:p>', para_start)
    return end + len('</w:p>') if end != -1 else -1

# ── Shared helper: finalize XML before saving ─────────────────────────────────
def _finalize_xml(xml_str):
    """Strip row heights, renumber bookmark IDs, close unmatched bookmarkStarts."""
    xml_str = re.sub(r'<w:trHeight[^/]*/>', '', xml_str)
    _c = [0]
    def _nxt(m): _c[0] += 1; return m.group(1) + str(_c[0]) + m.group(3)
    xml_str = re.sub(r'(<w:bookmarkStart[^>]*\bw:id=")(\d+)(")', _nxt, xml_str)
    _c[0] = 0
    xml_str = re.sub(r'(<w:bookmarkEnd[^>]*\bw:id=")(\d+)(")',   _nxt, xml_str)
    _starts = re.findall(r'<w:bookmarkStart[^>]*\bw:id="(\d+)"', xml_str)
    _ends   = re.findall(r'<w:bookmarkEnd[^>]*\bw:id="(\d+)"',   xml_str)
    _miss   = [i for i in _starts if i not in _ends]
    if _miss:
        xml_str = xml_str.replace('</w:body>',
            ''.join(f'<w:bookmarkEnd w:id="{i}"/>' for i in _miss) + '</w:body>', 1)
    return xml_str

# ── Load template ─────────────────────────────────────────────────────────────
print(f"Reading template: {os.path.basename(TEMPLATE_PATH)}")
with open(TEMPLATE_PATH, 'rb') as f:
    tmpl_bytes = f.read()
zin = zipfile.ZipFile(io.BytesIO(tmpl_bytes), 'r')
all_files = {name: zin.read(name) for name in zin.namelist()}
zin.close()
try:    xml = all_files['word/document.xml'].decode('utf-8')
except: xml = all_files['word/document.xml'].decode('latin-1')

# ── Step 1: Scalar replacements ───────────────────────────────────────────────
print("Updating scalar values...")
print("  Template: Poland / White Certificates")

# Simple repl()-based substitutions (single XML run nodes)
for old, new in [
    ('Dyckerhoff',     CUSTOMER),
    ('Customer Input', DATA_SOURCE),
]:
    xml = repl(xml, old, new)

# Poland Cover page: text is split across multiple <w:t> runs, so use nth_wt().

# Plant name TR2: template stores it as split nodes e.g. 'P','ol','and'
tr = get_tr_pos(xml)
_r2 = strip_ids(xml[tr[2][0]:tr[2][1]])
_r2 = nth_wt(_r2, 1, PLANT); _r2 = nth_wt(_r2, 2, ''); _r2 = nth_wt(_r2, 3, '')
xml = xml[:tr[2][0]] + _r2 + xml[tr[2][1]:]

# Report date TR4: template stores as split nodes e.g. '06','.','16','.20','25'
tr = get_tr_pos(xml)
_r4 = strip_ids(xml[tr[4][0]:tr[4][1]])
_r4 = nth_wt(_r4, 1, RPT_DATE)
for _i in range(2, 6): _r4 = nth_wt(_r4, _i, '')
xml = xml[:tr[4][0]] + _r4 + xml[tr[4][1]:]

# Number of assets TR5: "NPV_POS_CNT / NA" in split nodes
tr = get_tr_pos(xml)
_r5 = strip_ids(xml[tr[5][0]:tr[5][1]])
_r5 = nth_wt(_r5, 1, str(NPV_POS_CNT)); _r5 = nth_wt(_r5, 2, '')
_r5 = nth_wt(_r5, 3, ' / ');            _r5 = nth_wt(_r5, 4, str(NA)); _r5 = nth_wt(_r5, 5, '')
xml = xml[:tr[5][0]] + _r5 + xml[tr[5][1]:]

# Electricity cost TR6
tr = get_tr_pos(xml)
_r6 = strip_ids(xml[tr[6][0]:tr[6][1]])
_r6 = nth_wt(_r6, 1, f'{float(ELEC_PRICE):.2f} {CURRENCY}/kWh')
for _i in range(2, 6): _r6 = nth_wt(_r6, _i, '')
xml = xml[:tr[6][0]] + _r6 + xml[tr[6][1]:]

# Carbon intensity TR7
tr = get_tr_pos(xml)
_r7 = strip_ids(xml[tr[7][0]:tr[7][1]])
_r7 = nth_wt(_r7, 1, f'{CO2_INTENSITY} kg CO2/kWh')
for _i in range(2, 10): _r7 = nth_wt(_r7, _i, '')
xml = xml[:tr[7][0]] + _r7 + xml[tr[7][1]:]

# Tax rate TR8
tr = get_tr_pos(xml)
_r8 = strip_ids(xml[tr[8][0]:tr[8][1]])
_r8 = nth_wt(_r8, 1, pct_str(TAX_RATE)); _r8 = nth_wt(_r8, 2, ''); _r8 = nth_wt(_r8, 3, '')
xml = xml[:tr[8][0]] + _r8 + xml[tr[8][1]:]

# Summary TR12: annual savings + NPV+/total count
tr = get_tr_pos(xml)
_r12 = strip_ids(xml[tr[12][0]:tr[12][1]])
_r12 = nth_wt(_r12, 0, f'{SYM} {_fmtK(ANNUAL_SAVINGS)}')
for _i in range(1, 6): _r12 = nth_wt(_r12, _i, '')
_r12 = nth_wt(_r12, 8, str(NPV_POS_CNT)); _r12 = nth_wt(_r12, 9, '')
_r12 = nth_wt(_r12, 11, str(NA));         _r12 = nth_wt(_r12, 12, '')
xml = xml[:tr[12][0]] + _r12 + xml[tr[12][1]:]

# WC Payback fallback: if Excel rows 33-34 col 15 are uncached formula cells (=0),
# recompute WC-adjusted payback = invest / (savings + wc_earnings).
if _WC_PB_TOP10 == 0 and top10:
    _t10s_wc = sum(float(a.get('e_sav_cost') or 0) for a in top10)
    _t10wce   = sum(a.get('wc_earn', 0) for a in top10)
    _t10inv   = sum(float(a.get('invest') or 0) for a in top10)
    _denom    = _t10s_wc + _t10wce
    _WC_PB_TOP10 = round(_t10inv / _denom, 2) if _denom > 0 else 0
if _WC_PB_NPVP == 0 and npv_pos:
    _nps_wc  = sum(float(a.get('e_sav_cost') or 0) for a in npv_pos)
    _npwce   = sum(a.get('wc_earn', 0) for a in npv_pos)
    _npinv   = sum(float(a.get('invest') or 0) for a in npv_pos)
    _denom   = _nps_wc + _npwce
    _WC_PB_NPVP = round(_npinv / _denom, 2) if _denom > 0 else 0

# Business Case TR14: payback INCLUDING White Certificates + NPV/Investment
# Summary payback selection:
#   NO_NPV_MODE      → simple payback (no WC data available)
#   NPV+ count < NT  → WC payback for NPV+ subset  (Excel row 34 col 15)
#   NPV+ count >= NT → WC payback for Top-10 set   (Excel row 33 col 15)
if NO_NPV_MODE:
    _PAY = PAYBACK_TIME or TOP10_PAYBACK
elif NPV_POS_CNT < NT:
    _PAY = _WC_PB_NPVP    # NPV+ assets WC payback
else:
    _PAY = _WC_PB_TOP10   # Top-10 assets WC payback
_NPV_VAL = t10_invest if float(TOP10_NPV or 0) <= 0 else TOP10_NPV
_NPV_LBL = 'Investment cost*' if float(TOP10_NPV or 0) <= 0 else 'Net Present Value*'
tr = get_tr_pos(xml)
_r14 = strip_ids(xml[tr[14][0]:tr[14][1]])
_r14 = nth_wt(_r14, 0, fmtyrs(_PAY)); _r14 = nth_wt(_r14, 1, ''); _r14 = nth_wt(_r14, 2, ' yrs')
_r14 = nth_wt(_r14, 3, fmtyrs(_PAY)); _r14 = nth_wt(_r14, 4, ''); _r14 = nth_wt(_r14, 5, ' yrs')
if NA <= 10: _r14 = nth_wt(_r14, 9, '')  # remove "Top 10" label when all assets shown
_r14 = nth_wt(_r14, 13, SYM);           _r14 = nth_wt(_r14, 14, '')
_r14 = nth_wt(_r14, 15, _fmtK(_NPV_VAL)); _r14 = nth_wt(_r14, 16, ''); _r14 = nth_wt(_r14, 17, _NPV_LBL)
xml = xml[:tr[14][0]] + _r14 + xml[tr[14][1]:]

# Investment / WC Incentive / IRR TR15
_WC_TOTAL = sum(a.get('wc_earn', 0) for a in npv_pos)
tr = get_tr_pos(xml)
_r15 = strip_ids(xml[tr[15][0]:tr[15][1]])
_r15 = nth_wt(_r15, 0, f'{SYM} {_fmtK(INVEST_COST)}')
for _i in range(1, 6): _r15 = nth_wt(_r15, _i, '')
_r15 = nth_wt(_r15, 7, f'{SYM} {_fmtK(_WC_TOTAL)}'); _r15 = nth_wt(_r15, 8, '')
_r15 = nth_wt(_r15, 10, fmtirr(TOP10_IRR));           _r15 = nth_wt(_r15, 11, '')
_r15 = nth_wt(_r15, 12, fmtirr(TOP10_IRR));           _r15 = nth_wt(_r15, 13, '')
if NA <= 10: _r15 = nth_wt(_r15, 15, '')  # remove "– Top 10" label
xml = xml[:tr[15][0]] + _r15 + xml[tr[15][1]:]

# Sustainability TR17: CO2 savings + EV equivalents
tr = get_tr_pos(xml)
_r17 = strip_ids(xml[tr[17][0]:tr[17][1]])
_r17 = nth_wt(_r17, 0, str(round(float(CO2_SAVINGS or 0) / 1000)))
_r17 = nth_wt(_r17, 1, ''); _r17 = nth_wt(_r17, 2, ' tCO2')
_r17 = nth_wt(_r17, 5, str(round(float(BEV_COUNT or 0))))
xml = xml[:tr[17][0]] + _r17 + xml[tr[17][1]:]

# Energy savings section heading: clear stale '– Top 10' suffix node, then rewrite
xml = repl(xml, '\u2013 Top 10', '')
xml = repl(xml,
    'Energy savings with ABB premium efficiency solutions',
    f'Energy savings with ABB premium efficiency solutions \u2013 Top {NT}')

# Currency in top-N table column headers (Poland template uses EUR as placeholder)
for old, new in [
    ('Energy Cost (EUR)',         f'Energy Cost ({CURRENCY})'),
    ('Investment (EUR)',          f'Investment ({CURRENCY})'),
    ('Energy Cost Savings (EUR)', f'Energy Cost Savings ({CURRENCY})'),
]:
    xml = repl(xml, old, new)
# Also replace standalone EUR nodes in appendix column headers
xml = re.sub(r'(<w:t[^>]*>)(EUR)(</w:t>)',
             lambda m: m.group(1) + xe(CURRENCY) + m.group(3), xml)

# ── Step 2: Footer / header dates ─────────────────────────────────────────────
print(f"Updating footer dates: 2025-09-19 → {RPT_DATE_ISO}")
for fname in ['word/footer1.xml', 'word/footer2.xml', 'word/footer3.xml',
              'word/header1.xml', 'word/header2.xml', 'word/header3.xml']:
    if fname in all_files:
        content = all_files[fname]
        try:    txt = content.decode('utf-8')
        except: txt = content.decode('latin-1')
        txt = txt.replace('2025-09-19', RPT_DATE_ISO)
        all_files[fname] = txt.encode('utf-8')

# ── Step 3: Charts (editable literal values) ──────────────────────────────────
# Replace <c:numRef>..formula ref..</c:numRef> with <c:numLit> so the chart
# data is self-contained and always editable in Word without the source Excel.

cons_before = float(CONSUMP_BEFORE or 0)
cons_after  = cons_before - float(SAVINGS_KWH or 0)

def numlit_1pt(fmt_code, value):
    return (f'<c:numLit><c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="1"/>'
            f'<c:pt idx="0"><c:v>{value:.10f}</c:v></c:pt></c:numLit>')

def numlit_5pt(fmt_code, values):
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v:.10f}</c:v></c:pt>'
                  for i, v in enumerate(values))
    return (f'<c:numLit><c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="5"/>{pts}</c:numLit>')

# Chart 1: Energy Consumption bar — auto-select kWh / MWh / GWh unit
if cons_before >= 1_000_000:
    unit_label    = 'GWh'
    datalabel_fmt = r'#.0,,\ &quot;GWh&quot;'
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'
    disp_before   = f'{cons_before/1e6:.1f} GWh'
    disp_after    = f'{cons_after/1e6:.1f} GWh'
elif cons_before >= 1_000:
    unit_label    = 'MWh'
    datalabel_fmt = r'#.0,\ &quot;MWh&quot;'
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'
    disp_before   = f'{cons_before/1e3:.1f} MWh'
    disp_after    = f'{cons_after/1e3:.1f} MWh'
else:
    unit_label    = 'kWh'
    datalabel_fmt = r'#,##0\ &quot;kWh&quot;'
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'
    disp_before   = f'{cons_before:.0f} kWh'
    disp_after    = f'{cons_after:.0f} kWh'

if 'word/charts/chart1.xml' in all_files:
    ctxt = all_files['word/charts/chart1.xml'].decode('utf-8')
    ctxt = re.sub(
        r'<c:numRef><c:f>Saving_Calculations!\$C\$16</c:f>.*?</c:numRef>',
        numlit_1pt(numlit_fmt, cons_before), ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:numRef><c:f>Saving_Calculations!\$C\$18</c:f>.*?</c:numRef>',
        numlit_1pt(numlit_fmt, cons_after),  ctxt, flags=re.DOTALL)
    ctxt = ctxt.replace('#.0,,\\ &quot;GWh&quot;', datalabel_fmt)
    all_files['word/charts/chart1.xml'] = ctxt.encode('utf-8')
    print(f"  Chart 1 (Energy Consumption): {disp_before} → {disp_after}  [{unit_label}, editable]")

# Chart 2: NPV positive pie chart — update embedded Excel workbook
if 'word/embeddings/Microsoft_Excel_Worksheet.xlsx' in all_files:
    emb_bytes = all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx']
    ewb = openpyxl.load_workbook(io.BytesIO(emb_bytes))
    ews = ewb.active
    ews['B2'] = NPV_POS_CNT
    ews['B3'] = NA - NPV_POS_CNT
    ews['A2'] = 'NPV Positive'
    ews['A3'] = 'Remaining Assets'
    buf2 = io.BytesIO()
    ewb.save(buf2)
    all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx'] = buf2.getvalue()
    if 'word/charts/chart2.xml' in all_files:
        c2txt = all_files['word/charts/chart2.xml'].decode('utf-8')
        c2txt = c2txt.replace('<c:v>192</c:v>', f'<c:v>{NPV_POS_CNT}</c:v>')
        c2txt = c2txt.replace('<c:v>6</c:v>',   f'<c:v>{NA - NPV_POS_CNT}</c:v>')
        c2txt = c2txt.replace('<c:v>2nd Qtr</c:v>', '<c:v>Remaining Assets</c:v>')
        all_files['word/charts/chart2.xml'] = c2txt.encode('utf-8')
    print(f"  Chart 2 (NPV pie): {NPV_POS_CNT} NPV+ / {NA - NPV_POS_CNT} remaining  [editable]")

# Chart 3: Payback sensitivity — dual-pass update for robustness
# Pass 1: replace standard formula refs with numLit literals
# Pass 2: update numCache values directly (handles non-standard formula ref paths)
if 'word/charts/chart3.xml' in all_files and SENSITIVITY:
    ctxt = all_files['word/charts/chart3.xml'].decode('utf-8')
    # Pass 1: formula ref → numLit (flexible sheet-name match)
    ctxt = re.sub(
        r'<c:numRef><c:f>[^<]*\$T\$19:\$T\$23</c:f>.*?</c:numRef>',
        numlit_5pt('0%',   [s[0] for s in SENSITIVITY]), ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:numRef><c:f>[^<]*\$V\$19:\$V\$23</c:f>.*?</c:numRef>',
        numlit_5pt('0.00', [s[1] for s in SENSITIVITY]), ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:strRef><c:f>[^<]*\$S\$18[^<]*</c:f>.*?</c:strRef>',
        '<c:v>Payback time sensitivity to electricity price</c:v>',
        ctxt, flags=re.DOTALL)
    # Pass 2: update numCache in the second <c:numRef> block directly
    _refs = list(re.finditer(r'<c:numRef>', ctxt))
    if len(_refs) >= 2:
        _rs     = _refs[1].start()
        _re_end = ctxt.find('</c:numRef>', _rs) + len('</c:numRef>')
        _seg    = ctxt[_rs:_re_end]
        _new_pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>'
                           for i, (_, v) in enumerate(SENSITIVITY))
        _seg = re.sub(r'<c:ptCount[^/]*/>', f'<c:ptCount val="{len(SENSITIVITY)}"/>', _seg)
        _seg = re.sub(r'(<c:ptCount[^/]*/>)(.*?)(?=</c:numCache>)',
                      lambda m: m.group(1) + _new_pts, _seg, flags=re.DOTALL)
        ctxt = ctxt[:_rs] + _seg + ctxt[_re_end:]
    # Remove external data link — prevents Word from refreshing chart on open
    ctxt = re.sub(r'<c:externalData\b[^>]*/>', '', ctxt)
    ctxt = re.sub(r'<c:externalData\b.*?</c:externalData>', '', ctxt, flags=re.DOTALL)
    all_files['word/charts/chart3.xml'] = ctxt.encode('utf-8')
    _rels_name = 'word/charts/_rels/chart3.xml.rels'
    if _rels_name in all_files:
        _rels = all_files[_rels_name].decode('utf-8')
        _rels = re.sub(r'<Relationship\b[^>]*Id="rId4"[^/]*/>', '', _rels)
        all_files[_rels_name] = _rels.encode('utf-8')
    print(f"  Chart 3 (Payback sensitivity WC): {[round(s[1],2) for s in SENSITIVITY]}  [editable]")

# ── Step 5a: Remove IRR row when NPV ≤ 0 ─────────────────────────────────────
NO_NPV = float(TOP10_NPV or 0) <= 0
if NO_NPV:
    tr_pre = get_tr_pos(xml)
    xml = xml[:tr_pre[15][0]] + xml[tr_pre[15][1]:]
    print("  Summary: NPV ≤ 0 — Investment cost shown, IRR row removed")

# irr_adj: -1 when IRR row (TR 15) was removed, shifting all subsequent TR indices
irr_adj = -1 if NO_NPV else 0

# ── Step 5: Top-N data rows ───────────────────────────────────────────────────
print(f"Building Top-{NT} table ({NT} rows)...")
tr = get_tr_pos(xml)
tmpl_t10 = xml[tr[19 + irr_adj][0]:tr[19 + irr_adj][1]]

def v_top10(a):
    """Poland: 13 columns — WC Earnings + WC Payback (no Take-Back ✓).
    Column order: #, Tag, E.Cons, E.Cost, CO2.Cons, E.Sav.kWh,
                  E.Sav.Cost, Sav%, Invest, Payback, CO2.Sav, WC.Earn, WC.Payback
    """
    is_npv = a["npv"] is not None and float(a["npv"] or 0) > 0
    wce    = float(a.get("wc_earn") or 0)
    wcp    = float(a.get("wc_payback") or 0)
    return [str(a["num"]), a["tag"],
            fmt(a["e_cons"]), fmt(a["e_cost"]), fmt(a["co2_cons"]),
            fmt(a["e_sav_kwh"]), fmt(a["e_sav_cost"]),
            fmtpct(a["e_sav_pct"]), fmt(a["invest"]),
            f"{fmtyrs(a['payback'])} " if a["payback"] else "",
            f"{fmt(a['co2_sav'])} ",
            f"{fmt(wce)} " if wce else "\u2013",
            f"{fmtyrs(wcp)} " if (is_npv and wcp) else ("\u2013" if is_npv else "")]

def build_top10_row(a):
    """Fill a top-table row; apply white shading to NPV-negative padded assets."""
    row_xml = fill_row(tmpl_t10, v_top10(a))
    is_npv_neg = (a["npv"] is None or float(a["npv"] or 0) <= 0) and not NO_NPV_MODE
    if is_npv_neg:
        # Cells 0-9 → white; cells 10-12 (CO2 Sav, WC Earn, WC Payback) keep green
        tc_starts = [m.start() for m in re.finditer(r'<w:tc>', row_xml)]
        if len(tc_starts) >= 11:
            cut    = tc_starts[10]
            before = row_xml[:cut].replace('w:fill="E2EFD9"', 'w:fill="FFFFFF"')
            row_xml = before + row_xml[cut:]
    return row_xml

block = "".join(build_top10_row(a) for a in top10)
# Replace original 10 template data rows with NT actual rows
xml = xml[:tr[19 + irr_adj][0]] + block + xml[tr[28 + irr_adj][1]:]
tr  = get_tr_pos(xml)

# Dynamic index formulas — relative to NT and irr_adj
I_TR1       = 19 + NT + irr_adj   # "Total with Top N assets" total row
I_TR2       = 20 + NT + irr_adj   # "Total NPV Positive Assets" row
I_APP_FIRST = 22 + NT + irr_adj   # first App Details data row
I_APP_LAST  = 31 + NT + irr_adj   # last of 10 original App Details rows

# ── Step 6: Total rows ────────────────────────────────────────────────────────
# Poland total row: label at node 0, data starts at node 1 (different from Standard)
# Nodes: [0]=label [1]=E.Cons [2]=E.Cost [3]=CO2.Cons [4]=E.Sav.kWh
#        [5]=E.Sav.Cost [6]=Sav% [7]=Invest [8]=Payback [9]=CO2.Sav
#        [10]=WC.Earn [11]=WC.Payback
_t10_wc_earn = sum(a.get('wc_earn', 0) for a in top10)
r29 = strip_ids(xml[tr[I_TR1][0]:tr[I_TR1][1]])
r29 = nth_wt(r29, 0, f"Total with Top {NT} assets")
r29 = nth_wt(r29, 1, fmt(t10_e_cons))
r29 = nth_wt(r29, 2, fmt(t10_e_cost))
r29 = nth_wt(r29, 3, fmt(t10_co2_c))
r29 = nth_wt(r29, 4, fmt(t10_sav_kwh))
r29 = nth_wt(r29, 5, fmt(t10_sav_c))
r29 = nth_wt(r29, 6, fmtpct(t10_pct))
r29 = nth_wt(r29, 7, fmt(t10_invest))
r29 = nth_wt(r29, 8, fmtyrs(TOP10_PAYBACK))
r29 = nth_wt(r29, 9, fmt(t10_co2_sav))
r29 = nth_wt(r29, 10, fmt(_t10_wc_earn))
r29 = nth_wt(r29, 11, fmtyrs(_WC_PB_TOP10))
xml = xml[:tr[I_TR1][0]] + r29 + xml[tr[I_TR1][1]:]
tr  = get_tr_pos(xml)

nc = float(CONSUMP_BEFORE or 0)
# Show "Total NPV Positive Assets" row when NPV+ count differs from top-N count
if NPV_POS_CNT != NT and NPV_POS_CNT > 0:
    _npv_wc_earn = sum(a.get('wc_earn', 0) for a in npv_pos)
    r30 = strip_ids(xml[tr[I_TR2][0]:tr[I_TR2][1]])
    r30 = nth_wt(r30, 0,  f"Total NPV Positive Assets ({NPV_POS_CNT})")
    r30 = nth_wt(r30, 1,  fmt(nc))
    r30 = nth_wt(r30, 2,  fmt(nc * float(ELEC_PRICE or 8)))
    r30 = nth_wt(r30, 3,  fmt(nc * float(CO2_INTENSITY or 0.54)))
    r30 = nth_wt(r30, 4,  fmt(float(SAVINGS_KWH or 0)))
    r30 = nth_wt(r30, 5,  fmt(float(ANNUAL_SAVINGS or 0)))
    r30 = nth_wt(r30, 6,  fmtpct(float(SAVINGS_KWH or 0) / nc if nc else 0))
    r30 = nth_wt(r30, 7,  fmt(float(INVEST_COST or 0)))
    r30 = nth_wt(r30, 8,  fmtyrs(PAYBACK_TIME))
    r30 = nth_wt(r30, 9,  fmt(float(CO2_SAVINGS or 0)))
    r30 = nth_wt(r30, 10, fmt(_npv_wc_earn))
    r30 = nth_wt(r30, 11, fmtyrs(_WC_PB_NPVP))
    xml = xml[:tr[I_TR2][0]] + r30 + xml[tr[I_TR2][1]:]
    npv_adj = 0
else:
    xml = xml[:tr[I_TR2][0]] + xml[tr[I_TR2][1]:]
    npv_adj = -1  # NPV row removed → subsequent indices shift by -1
tr  = get_tr_pos(xml)

# ── Step 7: Application Details top-N ────────────────────────────────────────
def v_app(a):
    """Poland: shaft height column is always blank (not in Poland Excel layout).
    run_hrs formatted as integer (no decimal).
    """
    run_hrs_raw = a["run_hrs"]
    try:
        run_hrs_str = str(int(float(run_hrs_raw))) if run_hrs_raw else ''
    except:
        run_hrs_str = str(run_hrs_raw) if run_hrs_raw else ''
    return [str(a["num"]), a["tag"], a["ie"], a["load"],
            a["flow_ctrl"], a["connection"],
            str(a["output_kw"]), '',          # shaft height always blank
            run_hrs_str, str(a["avg_val"]),
            a["ess_motor"], a["ess_conn"]]

I_APP_FIRST_a = I_APP_FIRST + npv_adj
I_APP_LAST_a  = I_APP_LAST  + npv_adj

print(f"Building Application Details ({NT} rows)...")
tmpl_app = xml[tr[I_APP_FIRST_a][0]:tr[I_APP_FIRST_a][1]]
block = "".join(fill_row(tmpl_app, v_app(a)) for a in top10)
xml = xml[:tr[I_APP_FIRST_a][0]] + block + xml[tr[I_APP_LAST_a][1]:]
tr  = get_tr_pos(xml)

# ── Step 8: Appendix all-assets table ────────────────────────────────────────
I_APP2_FIRST = 23 + 2*NT + irr_adj + npv_adj   # first data row of appendix table
I_APP2_LAST  = I_APP2_FIRST + 16               # last of 17 template data rows
I_APP2_HDR   = I_APP2_FIRST - 1               # appendix table header row

def v_app2(a):
    """Appendix all-assets table — 13 columns (WC Earnings inserted after E.Cost.Sav).
    Column order: #, Tag, E.Cons, E.Cost, CO2.Cons, E.Sav.kWh, Invest,
                  E.Cost.Sav, WC.Earn, Payback, CO2.Sav, E.Sav%, NPV
    """
    npv   = a["npv"]
    npv_s = (fmt(npv) + " ") if (npv is not None and float(npv) > 0) else fmt_parens(npv) if npv is not None else ""
    co2_r = float(a["co2_sav"] or 0)
    co2_s = (fmt(co2_r) + " ") if co2_r > 0 else fmt_parens(co2_r)
    sav_e = fmt(a["e_sav_kwh"])  if float(a["e_sav_kwh"]  or 0) >= 0 else f"-{fmt(abs(float(a['e_sav_kwh'])))}"
    sav_c = fmt(a["e_sav_cost"]) if float(a["e_sav_cost"] or 0) >= 0 else f"-{fmt(abs(float(a['e_sav_cost'])))}"
    wce   = float(a.get("wc_earn") or 0)
    wce_s = fmt(wce) if wce else "\u2013"
    return [str(a["num"]), a["tag"],
            fmt(a["e_cons"]), fmt(a["e_cost"]), fmt(a["co2_cons"]),
            sav_e, fmt(a["invest"]), sav_c,
            wce_s,                                               # [8] WC Earnings
            f"{fmtyrs(a['payback'])} " if a["payback"] else "",  # [9] Payback
            co2_s, fmtpct(a["e_sav_pct"]), npv_s]               # [10][11][12]

if NA > 10:
    print(f"Building Appendix ({NA} rows)...")
    # ── Insert WC Earnings column (after cell 7 = Energy Cost Savings) ────────
    # Modify the header row: add "WC Earnings (SYM)" header cell after cell 7
    hdr_row = xml[tr[I_APP2_HDR][0]:tr[I_APP2_HDR][1]]
    hdr_row = _insert_cell_after(hdr_row, 7, f"WC Earnings ({SYM})")
    xml = xml[:tr[I_APP2_HDR][0]] + hdr_row + xml[tr[I_APP2_HDR][1]:]
    tr  = get_tr_pos(xml)
    # Recompute indices after header modification
    I_APP2_FIRST = 23 + 2*NT + irr_adj + npv_adj
    I_APP2_LAST  = I_APP2_FIRST + 16
    # Modify the template data row: insert an empty cell after cell 7
    tmpl_ap2 = xml[tr[I_APP2_FIRST][0]:tr[I_APP2_FIRST][1]]
    tmpl_ap2 = _insert_cell_after(tmpl_ap2, 7, '')
    # Expand all 17 template rows with the extra cell, then fill with data
    block = "".join(fill_row(tmpl_ap2, v_app2(a)) for a in assets_by_num)
    xml = xml[:tr[I_APP2_FIRST][0]] + block + xml[tr[I_APP2_LAST][1]:]
    print(f"  Appendix: WC Earnings column inserted")
else:
    # NA < 10: remove entire appendix numerical table (header + 17 data rows)
    xml = xml[:tr[I_APP2_HDR][0]] + xml[tr[I_APP2_LAST][1]:]
    print(f"Appendix numerical table: removed (NA={NA} < 10)")
tr  = get_tr_pos(xml)

# ── Step 9: Details of Recommendation ────────────────────────────────────────
# Use dynamic header detection (more robust than fixed index arithmetic).
_det_hdr_nodes_target = ['#', 'Application', 'IE', 'Driven load', 'Flow Control']
I_DET_HDR = next(
    (i for i, (s, e) in enumerate(tr)
     if re.findall(r'<w:t[^>]*>([^<]*)</w:t>', xml[s:e])[:5] == _det_hdr_nodes_target),
    None)

if I_DET_HDR is None:
    print("WARNING: Det Rec header not found — skipping Details of Recommendation")
else:
    I_DET_FIRST = I_DET_HDR + 1
    I_DET_LAST  = I_DET_HDR + 17   # Poland template has 17 Det Rec data rows
    if NA > 10:
        print(f"Building Details of Recommendation ({NA} rows)...")
        tmpl_det  = xml[tr[I_DET_FIRST][0]:tr[I_DET_FIRST][1]]
        block_det = "".join(fill_row(tmpl_det, v_app(a)) for a in assets_by_num)
        xml = xml[:tr[I_DET_FIRST][0]] + block_det + xml[tr[I_DET_LAST][1]:]
    else:
        # NA <= 10: remove entire Det Rec table (header + 17 rows)
        xml = xml[:tr[I_DET_HDR][0]] + xml[tr[I_DET_LAST][1]:]
        print(f"Details of Recommendation: removed (NA={NA} < 10)")
tr = get_tr_pos(xml)

# ── Step 10: Trim Appendix section when NA < 10 ───────────────────────────────
# Remove all content between the "Appendix" heading paragraph and the
# Calculation Methodology table. In Poland the Calc Methodology is inside a
# table cell (not a standalone heading paragraph), so we anchor on 'Method A'
# and find the enclosing <w:tbl>.
data_listed_pos    = xml.rfind('Data is listed as total annual')
if data_listed_pos == -1: data_listed_pos = 0
appendix_para_start = find_para_start(xml, '>Appendix<', data_listed_pos)
appendix_para_end   = find_para_end(xml, appendix_para_start)

if NA <= 10 and appendix_para_end > 0:
    _method_pos = xml.find('Method A', appendix_para_end)
    if _method_pos > 0:
        _calc_tbl = xml.rfind('<w:tbl>', 0, _method_pos)
        if _calc_tbl > appendix_para_end:
            xml = xml[:appendix_para_end] + xml[_calc_tbl:]
            print(f"Appendix trimmed (NA={NA} <= 10): kept only Calculation Methodology + NPV Methodology")
        else:
            print("WARNING: Could not trim appendix — Calculation Methodology table not found")
    else:
        print("WARNING: Could not find 'Method A' for appendix trim")
elif NA > 10:
    print(f"Appendix kept full (NA={NA} > 10): App Det + Det Rec tables present")

# ── Save output docx ──────────────────────────────────────────────────────────
safe_c   = re.sub(r'[\\/:*?"<>|]', '_', args[1])
safe_p   = re.sub(r'[\\/:*?"<>|]', '_', PLANT)
out_name = f"{safe_c}_{safe_p}_EA_Report_Poland.docx"
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), out_name)

# TOC auto-update
all_files['word/document.xml'] = _finalize_xml(xml).encode('utf-8')
if 'word/settings.xml' in all_files:
    try:
        stg = all_files['word/settings.xml'].decode('utf-8')
        if 'updateFields' not in stg:
            stg = stg.replace('</w:settings>',
                '<w:updateFields w:val="1"/></w:settings>')
        all_files['word/settings.xml'] = stg.encode('utf-8')
    except Exception: pass

buf = io.BytesIO()
with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
    for name, data in all_files.items():
        zout.writestr(name, data)
with open(out_path, 'wb') as f:
    f.write(buf.getvalue())

print()
print(f"  Done!  →  {out_path}")
print(f"  {CUSTOMER} | {PLANT} | {CURRENCY} | {NA} assets | {NPV_POS_CNT} NPV+ | {NT} Top-N")
print(f"  Annual savings : {SYM} {fmt(ANNUAL_SAVINGS)}")
print(f"  Investment     : {SYM} {fmt(INVEST_COST)}")
print(f"  WC Incentive   : {SYM} {fmt(_WC_TOTAL)}")
print(f"  Payback (WC)   : {fmtyrs(_PAY)} yrs")
print(f"  NPV            : {SYM} {fmt(NPV_VALUE)}")
print(f"  IRR Top-{NT}      : {fmtirr(TOP10_IRR)}")
if SENSITIVITY:
    print(f"  Payback sensitivity: {[f'{s[0]:+.0%}→{s[1]:.1f}yrs' for s in SENSITIVITY]}")
