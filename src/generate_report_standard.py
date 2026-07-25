#!/usr/bin/env python3
"""
ABB Energy Appraisal – Standard Report Generator
Template: EA_Report_Template_Standard.docx
Regions:  India, Norway, Australia, and all standard markets

Usage:
  python generate_report_standard.py  <saving_calc.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]
                                       [--template <filename>] [--suffix <suffix>]

  --template  Word template filename to use, resolved under report_templates/.
              Defaults to ea_report_template_standard.docx (the Complete /
              All-Assets report). Pass ea_report_template_standard_top10.docx
              for the Executive / Top-10 report. The template determines the
              report structure; this script's rendering logic is unchanged
              regardless of which template is selected.
  --suffix    Optional suffix inserted before .docx in the output filename
              (e.g. "_Executive"), so two variants for the same customer/plant
              don't overwrite each other.

Requirements:
  pip install openpyxl
"""

import sys, os, re, math, io, zipfile
from datetime import datetime
import openpyxl

# ── CLI ───────────────────────────────────────────────────────────────────────
_raw_args = sys.argv[1:]

TEMPLATE_FILENAME = "ea_report_template_standard.docx"
OUTPUT_SUFFIX = ""

# Pull out optional --template/--suffix flags (order-independent), leaving only
# the original positional args (xlsx, customer, plant, date, data_source) behind.
args = []
_i = 0
while _i < len(_raw_args):
    if _raw_args[_i] == "--template" and _i + 1 < len(_raw_args):
        TEMPLATE_FILENAME = _raw_args[_i + 1]
        _i += 2
    elif _raw_args[_i] == "--suffix" and _i + 1 < len(_raw_args):
        OUTPUT_SUFFIX = _raw_args[_i + 1]
        _i += 2
    else:
        args.append(_raw_args[_i])
        _i += 1

if len(args) < 3:
    print('Usage: python generate_report_standard.py  <excel.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]  [--template <filename>] [--suffix <suffix>]')
    sys.exit(1)

XLSX_PATH   = args[0]
CUSTOMER    = args[1].upper()
PLANT       = args[2]
RPT_DATE    = args[3] if len(args) > 3 else datetime.now().strftime("%m.%d.%Y")
DATA_SOURCE = args[4] if len(args) > 4 else "Customer Input"

# Template file is selected via --template (defaults to the Complete report)
_script_dir   = os.path.dirname(os.path.abspath(__file__))
_template_dir = os.path.join(_script_dir, "report_templates")
TEMPLATE_PATH = os.path.join(_template_dir, TEMPLATE_FILENAME)
if not os.path.exists(TEMPLATE_PATH):
    print(f"ERROR: {TEMPLATE_FILENAME} not found in report_templates/")
    sys.exit(1)
TEMPLATE_V1_PATH = TEMPLATE_PATH  # same for this script

for path, label in [(XLSX_PATH, "Excel file"), (TEMPLATE_PATH, "Template")]:
    if not os.path.exists(path):
        print(f"ERROR: {label} not found: {path}")
        sys.exit(1)

def date_to_iso(d):
    """Convert any date string to YYYY-MM-DD."""
    for fmt in ('%m.%d.%Y','%d.%m.%Y','%Y-%m-%d','%d/%m/%Y','%m/%d/%Y'):
        try:
            return datetime.strptime(d, fmt).strftime('%Y-%m-%d')
        except: pass
    return d  # fallback: return as-is

RPT_DATE_ISO = date_to_iso(RPT_DATE)

# ── Read Excel ────────────────────────────────────────────────────────────────
print(f"Reading Excel: {os.path.basename(XLSX_PATH)}")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Accept any sheet matching Saving(s)_Calculation(s) with any version suffix
# Also handles 'Saving_Calculatios' (common typo in Poland files)
def _is_sc_sheet(name):
    n = name.strip().lower().replace(' ', '_')
    prefixes = ('saving_calculations', 'savings_calculations',
                'saving_calculatios',  'savings_calculatios')  # typo variants
    return any(n.startswith(p) for p in prefixes)

_sc_sheet = next((s for s in wb.sheetnames if _is_sc_sheet(s)), None)
if _sc_sheet is None:
    print(f"ERROR: No Saving_Calculations sheet found. Sheets: {wb.sheetnames}")
    sys.exit(1)
# If multiple matching sheets exist, prefer the one ending with _V<n> (version sheet)
# over Option sheets (e.g. _Option1_USD, _Option2)
_sc_candidates = [s for s in wb.sheetnames if _is_sc_sheet(s)]
if len(_sc_candidates) > 1:
    import re as _re
    _vn = [s for s in _sc_candidates if _re.search(r'_v\d+$', s.strip().lower())]
    if _vn:
        _sc_sheet = _vn[0]   # pick first version sheet
if _sc_sheet not in ("Saving_Calculations", "Savings_Calculations"):
    print(f"  Note: using sheet '{_sc_sheet}'")
ws = wb[_sc_sheet]

_XL_ERRORS = frozenset({
    # Excel formula error strings
    '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!', '#ERROR!',
    # Common "no value" placeholder strings used in Excel templates
    '-', '\u2013', '\u2014', 'N/A', 'n/a', 'NA', 'na', 'n.a.', 'N.A.',
})

def cv(row, col):
    """Read a cell value; return None for Excel error strings (#DIV/0!, #N/A, etc.)."""
    v = ws.cell(row=row, column=col).value
    return None if isinstance(v, str) and v.strip() in _XL_ERRORS else v

# ── Detect column positions from header row 36 ────────────────────────────────
HEADER_ROW = 36
col_map = {}
for c in range(1, 70):
    h = cv(HEADER_ROW, c)
    if h:
        col_map[str(h).strip().lower()] = c

def find_col(names, default):
    for n in names:
        if n.lower() in col_map:
            return col_map[n.lower()]
    return default

COL_NUM      = find_col(['#', '# no.', 'no.'], 2)
COL_E_CONS   = find_col(['annual energy. cons (kwh)', 'annual energy cons (kwh)'], 5)
COL_E_COST   = find_col(['annual energy cost'], 6)
COL_CO2_CONS = find_col(['annual co2 cons.'], 7)
COL_SAV_KWH  = find_col(['annual energy savings, kwh'], 8)
COL_SAV_COST = find_col(['annual energy cost savings'], 9)
COL_SAV_PCT  = find_col(['annual energy savings (%)'], 10)
COL_INVEST   = find_col(['investment'], 11)
# If v2-style Excel: col 13 = 'Investment - CEE' (net), col 14 = payback, col 15 = CO2 sav
# Detect by checking if col 13 header contains 'cee'
_hdr13 = str(ws.cell(HEADER_ROW, 13).value or '').lower()
_hdr12 = str(ws.cell(HEADER_ROW, 12).value or '').lower()
_IS_V2_EXCEL = 'cee' in _hdr12 or 'cee' in _hdr13
if _IS_V2_EXCEL:
    # Use net-of-CEE investment, and shift payback/CO2 columns accordingly
    COL_INVEST_NET = find_col(['investment - cee', 'investment-cee'], 13)
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 14)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 15)
    print(f"  Detected v2-style Excel (CEE columns) — using net investment col {COL_INVEST_NET}")
else:
    COL_INVEST_NET = COL_INVEST   # same column — no CEE in this Excel
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 12)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 13)
COL_NPV      = find_col(['npv'], 17)
COL_IE       = find_col(['ie eff class', 'ie'], 41)
COL_LOAD     = find_col(['driven load', 'application'], 42)
COL_FLOW     = find_col(['flow control', 'flow control method'], 43)
COL_CONN     = find_col(['dol vsd', 'connection'], 44)
COL_OUTPUT   = find_col(['output (kw)', 'rated power, kw'], 45)
COL_SHAFT    = find_col(['shaft height (frame)', 'shaft height'], 46)
COL_RUNHRS   = find_col(['running hours', 'running time (hours)'], 47)
COL_AVG      = find_col(['average flow % / average frequency (hz)',
                          'average flow', 'average frequency', 'avg. frequency'], 48)
COL_ESS      = find_col(['recommended ess motor'], 50)
COL_ESSC     = find_col(['ess connection'], 51)

# ── KPI values ────────────────────────────────────────────────────────────────
CURRENCY       = str(cv(3, 3) or "INR").strip()
ELEC_PRICE     = cv(5, 4) or 8
CO2_INTENSITY  = cv(6, 4) or 0.54
TAX_RATE       = cv(8, 4) or 0.349
DISCOUNT_RATE  = float(cv(7, 4) or 0.065)  # discount rate (row 7)
_NPV_YEARS     = 20
TOTAL_ASSETS   = int(cv(13, 3) or 0)
NPV_POS_CNT    = int(cv(14, 3) or 0)
ANNUAL_SAVINGS = cv(15, 3)
CONSUMP_BEFORE = cv(16, 3)
SAVINGS_KWH    = cv(17, 3)
NPV_VALUE      = cv(19, 3)
INVEST_COST    = cv(20, 3)
PAYBACK_TIME   = cv(21, 3)
CO2_SAVINGS    = cv(22, 3)
IRR_VALUE      = cv(23, 3)
BEV_COUNT      = cv(25, 4)
TOP10_PAYBACK  = cv(21, 4)
TOP10_NPV      = cv(19, 4)
TOP10_INVEST   = cv(20, 4)
TOP10_IRR      = cv(23, 4)

# Payback sensitivity data (rows 19-23, col 20=price delta, col 22=payback)
SENSITIVITY = []
for r in range(19, 24):
    delta   = cv(r, 20)
    payback = cv(r, 22)
    try:
        SENSITIVITY.append((float(delta), float(payback)))
    except (TypeError, ValueError):
        pass  # skip rows with None, '#DIV/0!' or other non-numeric Excel errors
_SENSITIVITY_PENDING = len(SENSITIVITY) < 5  # will recompute from assets if incomplete

# ── Load assets ───────────────────────────────────────────────────────────────
ASSET_START_ROW = 37
assets = []
_scan_limit = max(int(TOTAL_ASSETS or 0), 200)
for r in range(ASSET_START_ROW, ASSET_START_ROW + _scan_limit):
    num = cv(r, COL_NUM)
    if num is None: break
    try: num = int(float(num))
    except: break
    if not (1 <= num <= 9999): break

    ess_motor = str(cv(r, COL_ESS) or "")

    # Shaft height: use Excel column value only — 0 stays 0, never read from motor name
    shaft_raw = cv(r, COL_SHAFT)
    try:
        shaft_val = str(int(float(shaft_raw))) if shaft_raw is not None else ""
    except:
        shaft_val = str(shaft_raw) if shaft_raw else ""

    # Avg Flow% / Frequency: use col 48 directly (most accurate)
    avg_raw = cv(r, COL_AVG)
    if avg_raw is not None:
        try:
            v = float(avg_raw)
            # If it is a 0–1 fraction (e.g. 0.9) treat as percentage → "90" (no % sign)
            avg_val = f"{round(v*100)}" if 0 < v <= 1 else str(avg_raw).replace(' ', '')
        except:
            avg_val = str(avg_raw).replace(' ', '')
    else:
        # Fallback: extract Hz from motor name
        fm = re.search(r'([\d.]+)\s*Hz', ess_motor)
        avg_val = fm.group(1) if fm else ""

    # Connection: DOL when blank/None
    conn_raw = cv(r, COL_CONN)
    conn_val = str(conn_raw) if conn_raw else "DOL"

    assets.append({
        "num":       num,
        "tag":       str(cv(r, 3) or ""),
        "load":      str(cv(r, COL_LOAD) or cv(r, 4) or ""),
        "e_cons":    cv(r, COL_E_CONS)   or 0,
        "e_cost":    cv(r, COL_E_COST)   or 0,
        "co2_cons":  cv(r, COL_CO2_CONS) or 0,
        "e_sav_kwh": cv(r, COL_SAV_KWH)  or 0,
        "e_sav_cost":cv(r, COL_SAV_COST) or 0,
        "e_sav_pct": cv(r, COL_SAV_PCT)  or 0,
        "invest":    cv(r, COL_INVEST)   or 0,
        "payback":   cv(r, COL_PAYBACK) if not isinstance(cv(r, COL_PAYBACK), str) or cv(r, COL_PAYBACK).strip() not in ('', '-') else None,
        "co2_sav":   cv(r, COL_CO2_SAV)  or 0,
        "npv":       cv(r, COL_NPV),
        "ie":        str(cv(r, COL_IE)   or ""),
        "flow_ctrl": str(cv(r, COL_FLOW) or "Information not available"),
        "connection":conn_val,
        "output_kw": str(cv(r, COL_OUTPUT) or ""),
        "shaft_h":   shaft_val,
        "run_hrs":   str(cv(r, COL_RUNHRS) or ""),
        "avg_val":   avg_val,
        "ess_motor": ess_motor,
        "ess_conn":  str(cv(r, COL_ESSC) or ""),
    })

if not assets:
    print("ERROR: No assets found. Check TOTAL_ASSETS in Excel row 13.")
    sys.exit(1)

# ── Per-asset fallback: compute formula-dependent values when Excel hasn't cached them ──
# openpyxl data_only=True only reads cached values. If the workbook was saved
# without recalculating (manual calc mode), all formula cells return None.
# Recompute from raw input values (e_cons, e_sav_kwh, invest) which are always present.
_needs_fallback = any(a["e_sav_cost"] == 0 and a["e_sav_kwh"] and a["npv"] is None
                      for a in assets)
if _needs_fallback:
    print("  Note: Excel formula cells uncached — computing e_cost/NPV/payback from raw inputs")
for a in assets:
    ec   = float(a["e_cons"]    or 0)
    esav = float(a["e_sav_kwh"] or 0)
    inv  = float(a["invest"]    or 0)
    ep   = float(ELEC_PRICE     or 8)
    co2i = float(CO2_INTENSITY  or 0.054)

    if not a["e_cost"]     and ec   > 0: a["e_cost"]     = ec   * ep
    if not a["e_sav_cost"] and esav > 0: a["e_sav_cost"] = esav * ep
    if not a["e_sav_pct"]  and ec   > 0: a["e_sav_pct"]  = esav / ec
    if not a["co2_cons"]   and ec   > 0: a["co2_cons"]   = ec   * co2i
    if not a["co2_sav"]    and esav > 0: a["co2_sav"]    = esav * co2i

    # NPV via annuity formula: PV(after-tax savings, n years) − investment
    if a["npv"] is None:
        ann_sav = float(a["e_sav_cost"] or 0)
        if ann_sav > 0:
            r         = DISCOUNT_RATE
            after_tax = ann_sav * (1 - float(TAX_RATE or 0.25))
            pv_factor = ((1 - (1 + r) ** (-_NPV_YEARS)) / r) if r > 0 else _NPV_YEARS
            a["npv"]  = round(after_tax * pv_factor - inv, 2)

    # Simple payback = investment / annual savings (only when NPV > 0)
    if a["payback"] is None:
        ann_sav = float(a["e_sav_cost"] or 0)
        npv_val = float(a["npv"]) if a["npv"] is not None else -1
        if ann_sav > 0 and npv_val > 0:
            a["payback"] = round(inv / ann_sav, 2)

NA = len(assets)  # total asset count
npv_pos = [a for a in assets if a["npv"] is not None and float(a["npv"]) > 0]

# Recount when Excel KPI cells were uncached formula cells returning None/0
if TOTAL_ASSETS == 0:
    TOTAL_ASSETS = NA
if NPV_POS_CNT == 0 and npv_pos:
    NPV_POS_CNT = len(npv_pos)
    print(f"  Note: NPV_POS_CNT recomputed from assets: {NPV_POS_CNT}")

if npv_pos:
    # Sort NPV+ by payback ascending
    npv_sorted = sorted(
        npv_pos,
        key=lambda a: (float(a["payback"]) if a["payback"] else 9999,
                       -(float(a["e_sav_cost"] or 0))))[:10]

    # Change 2: When NPV+ < 10, pad with NPV- assets (white-shaded)
    # so the top table always shows all assets (up to 10 rows)
    if len(npv_sorted) < 10:
        npv_neg = [a for a in assets if a["npv"] is None or float(a["npv"] or 0) <= 0]
        npv_neg_sorted = sorted(npv_neg, key=lambda a: -(float(a["e_sav_cost"] or 0)))
        slots = 10 - len(npv_sorted)
        top10 = npv_sorted + npv_neg_sorted[:slots]
        print(f"  Top table: {len(npv_sorted)} NPV+ + {min(slots, len(npv_neg_sorted))} NPV- (white-shaded)")
    else:
        top10 = npv_sorted

    NO_NPV_MODE = False
else:
    # No NPV+ assets → use ALL assets sorted by energy cost savings descending
    top10 = sorted(assets, key=lambda a: -(float(a["e_sav_cost"] or 0)))[:10]
    NO_NPV_MODE = True
    print("  NOTE: No NPV-positive assets — using all assets sorted by energy savings")

NT = len(top10)   # actual top-N count (≤ 10)
NA = len(assets)  # total asset count

# Sort ALL assets by serial number (Sr. No.) for appendix + details tables
assets_by_num = sorted(assets, key=lambda a: a["num"])

# Change 3: Appendix logic
# NA > 10 → App Details + Details of Recommendation included in Appendix
# NA <= 10 → Appendix trimmed to Calculation Methodology + NPV only
SKIP_APP_DET = (NA <= 10)

# ── NO_NPV_MODE: re-derive summary KPIs from asset rows ──────────────────────
# The Excel "NPV+" summary cells (rows 15-25 cols 3-4) are zero/blank when ALL
# assets are NPV-negative. Compute directly from asset data so the summary page
# correctly shows: cost savings, energy consumption chart, CO2 avoided, Vehicles.
if NO_NPV_MODE:
    ANNUAL_SAVINGS = sum(float(a["e_sav_cost"] or 0) for a in assets)
    CO2_SAVINGS    = sum(float(a["co2_sav"]    or 0) for a in assets)
    CONSUMP_BEFORE = sum(float(a["e_cons"]     or 0) for a in assets)
    SAVINGS_KWH    = sum(float(a["e_sav_kwh"]  or 0) for a in assets)
    BEV_COUNT      = round(float(SAVINGS_KWH) / 3500) if SAVINGS_KWH else 0
    print("  NO_NPV_MODE: summary KPIs computed from all assets (Excel NPV+ cells empty)")

    # Compute payback sensitivity from assets when Excel payback cells are errors
    # Try to get price deltas from Excel col 20 (usually valid even in all-NPV- case)
    _deltas = []
    for r in range(19, 24):
        try:
            _deltas.append(float(cv(r, 20)))
        except (TypeError, ValueError):
            pass
    if not _deltas:
        _deltas = [-0.2, -0.1, 0.0, 0.1, 0.2]  # standard ±20% variation
    _total_invest  = sum(float(a["invest"]      or 0) for a in assets)
    _base_savings  = float(ANNUAL_SAVINGS or 0)
    SENSITIVITY = []
    for d in _deltas:
        adj_savings = _base_savings * (1 + d)
        if adj_savings > 0:
            SENSITIVITY.append((d, _total_invest / adj_savings))
    if SENSITIVITY:
        print(f"  NO_NPV_MODE: payback sensitivity computed from assets: {[f'{s[0]:+.0%}→{s[1]:.1f}yrs' for s in SENSITIVITY]}")

# ── General summary KPI fallback (uncached Excel formula cells) ───────────────
# When Excel was saved without recalculating, the KPI summary rows return None
# even when assets ARE NPV-positive. Recompute from asset data in that case.
_kpi_src = assets if NO_NPV_MODE else npv_pos
if ANNUAL_SAVINGS is None or float(ANNUAL_SAVINGS or 0) == 0:
    ANNUAL_SAVINGS = sum(float(a["e_sav_cost"] or 0) for a in _kpi_src)
    print(f"  KPI fallback: ANNUAL_SAVINGS = {ANNUAL_SAVINGS:,.0f}")
if CONSUMP_BEFORE is None or float(CONSUMP_BEFORE or 0) == 0:
    CONSUMP_BEFORE = sum(float(a["e_cons"] or 0) for a in _kpi_src)
if SAVINGS_KWH is None or float(SAVINGS_KWH or 0) == 0:
    SAVINGS_KWH = sum(float(a["e_sav_kwh"] or 0) for a in _kpi_src)
if CO2_SAVINGS is None or float(CO2_SAVINGS or 0) == 0:
    CO2_SAVINGS = sum(float(a["co2_sav"] or 0) for a in _kpi_src)
if INVEST_COST is None or float(INVEST_COST or 0) == 0:
    INVEST_COST = sum(float(a["invest"] or 0) for a in _kpi_src)
if BEV_COUNT is None or float(BEV_COUNT or 0) == 0:
    BEV_COUNT = round(float(SAVINGS_KWH or 0) / 3500)
if PAYBACK_TIME is None or float(PAYBACK_TIME or 0) == 0:
    _sav = float(ANNUAL_SAVINGS or 0)
    PAYBACK_TIME = round(float(INVEST_COST or 0) / _sav, 2) if _sav > 0 else None
if NPV_VALUE is None or float(NPV_VALUE or 0) == 0:
    NPV_VALUE = sum(float(a["npv"] or 0) for a in _kpi_src if a["npv"] is not None)
if TOP10_NPV is None or float(TOP10_NPV or 0) == 0:
    TOP10_NPV = sum(float(a["npv"] or 0) for a in top10
                    if a["npv"] is not None and float(a["npv"]) > 0)
if TOP10_INVEST is None or float(TOP10_INVEST or 0) == 0:
    TOP10_INVEST = sum(float(a["invest"] or 0) for a in top10)
if TOP10_PAYBACK is None or float(TOP10_PAYBACK or 0) == 0:
    _t10s = sum(float(a["e_sav_cost"] or 0) for a in top10)
    TOP10_PAYBACK = round(float(TOP10_INVEST) / _t10s, 2) if _t10s > 0 else None

# ── IRR fallback (Newton-Raphson on flat-annuity cash flow) ──────────────────
def _compute_irr(invest, annual_cf, n_years):
    """Return IRR for cash flows: -invest at t=0, annual_cf at t=1..n_years.
    Uses Newton-Raphson iteration.  Returns None if solution not found."""
    if invest <= 0 or annual_cf <= 0 or n_years <= 0:
        return None
    r = annual_cf / invest  # initial guess ≈ 1/payback
    for _ in range(200):
        try:
            if abs(r) < 1e-10:
                break
            pow_neg_n  = (1 + r) ** (-n_years)
            pow_neg_n1 = (1 + r) ** (-n_years - 1)
            annuity    = (1 - pow_neg_n) / r
            npv_val    = -invest + annual_cf * annuity
            # d(annuity)/dr = [n*(1+r)^(-n-1)*r + (1+r)^-n - 1] / r²
            d_annuity  = (n_years * pow_neg_n1 * r + pow_neg_n - 1) / (r * r)
            d_npv      = annual_cf * d_annuity
            if abs(d_npv) < 1e-12:
                break
            r_new = r - npv_val / d_npv
            if abs(r_new - r) < 1e-9:
                r = r_new
                break
            r = r_new
        except (ValueError, ZeroDivisionError, OverflowError):
            break
    return round(r, 6) if 0 < r < 10 else None  # sanity: IRR 0%–1000%

if TOP10_IRR is None or float(TOP10_IRR or 0) == 0:
    _t10i = float(TOP10_INVEST or 0)
    _t10s = sum(float(a["e_sav_cost"] or 0) for a in top10)
    _irr  = _compute_irr(_t10i, _t10s, _NPV_YEARS)
    if _irr is not None:
        TOP10_IRR = _irr
        print(f"  IRR fallback (Top-{NT}): {_irr:.1%}")

if IRR_VALUE is None or float(IRR_VALUE or 0) == 0:
    _src_irr = assets if NO_NPV_MODE else npv_pos
    _all_s   = sum(float(a["e_sav_cost"] or 0) for a in _src_irr)
    _all_i   = float(INVEST_COST or 0)
    _irr_all = _compute_irr(_all_i, _all_s, _NPV_YEARS)
    if _irr_all is not None:
        IRR_VALUE = _irr_all

# ── Sensitivity fallback for NPV+ mode (Excel cells uncached/error) ───────────
# NO_NPV_MODE already computes SENSITIVITY from assets above.
# This handles the case where there ARE NPV+ assets but Excel sensitivity rows
# returned None or error strings — leaving SENSITIVITY empty after the Excel read.
if _SENSITIVITY_PENDING and npv_pos:
    _ns = sum(float(a.get('e_sav_cost') or 0) for a in npv_pos)
    _ni = sum(float(a.get('invest')     or 0) for a in npv_pos)
    if _ns > 0:
        SENSITIVITY = [
            (_d, round(_ni / (_ns * (1 + _d)), 4))
            for _d in [-0.2, -0.1, 0.0, 0.1, 0.2]
            if _ns * (1 + _d) > 0
        ]
        print(f"  Sensitivity fallback (NPV+ assets): {[f'{s[0]:+.0%}→{s[1]:.2f}yrs' for s in SENSITIVITY]}")

print(f"  {NA} assets, {len(npv_pos)} NPV+, {NT} in Top-{NT} {'(no NPV+ mode)' if NO_NPV_MODE else '(sorted by payback)'}")

# ── Formatting helpers ────────────────────────────────────────────────────────
SYM = CURRENCY

def fmt(n, dp=0):
    if n is None or (isinstance(n, float) and math.isnan(n)): return "\u2013"
    try:
        n = float(n)
        s = f"{abs(n):,.{dp}f}"
        return s if n >= 0 else f"-{s}"
    except: return str(n)

def fmt_parens(n):
    if n is None: return "\u2013"
    try:
        n = float(n)
        return f"({fmt(abs(n))})" if n < 0 else fmt(n)
    except: return str(n)

def fmtpct(n):
    if n is None: return "\u2013"
    v = round(float(n) * 100)
    return f"{v}%" if v >= 0 else f"-{abs(v)}%"

def fmtyrs(n):
    if n is None: return ""
    try: return f"{float(n):.1f}"
    except: return ""  # handles '-' or any non-numeric string

def pct_str(n):
    if n is None: return "\u2013"
    return f"{round(float(n)*100, 1)}%"

def fmtirr(n):
    """Format IRR safely — Excel may store it as '-' string when undefined."""
    if n is None: return "\u2013"
    try: return f"{round(float(n)*100)}%"
    except: return str(n)   # e.g. returns '-' unchanged

# ── Top-N totals (from payback-sorted selection) ──────────────────────────────
t10_e_cons  = sum(float(a["e_cons"]    or 0) for a in top10)
t10_e_cost  = sum(float(a["e_cost"]    or 0) for a in top10)
t10_co2_c   = sum(float(a["co2_cons"]  or 0) for a in top10)
t10_sav_kwh = sum(float(a["e_sav_kwh"] or 0) for a in top10)
t10_invest  = sum(float(a["invest"]    or 0) for a in top10)
t10_sav_c   = sum(float(a["e_sav_cost"]or 0) for a in top10)
t10_co2_sav = sum(float(a["co2_sav"]  or 0) for a in top10)
t10_pct     = t10_sav_kwh / t10_e_cons if t10_e_cons else 0

# Recompute Top-N summary KPIs from our payback-sorted selection
# (payback and NPV are sum-based so we recalculate; IRR comes directly from Excel)
t10_simple_payback = t10_invest / t10_sav_c if t10_sav_c else 0
# Only sum NPV+ assets for t10_npv — NPV- padding must not affect the summary page
# (a padded NPV- asset with a large negative NPV would incorrectly trigger NO_NPV logic)
t10_npv = sum(float(a["npv"] or 0) for a in top10
              if a["npv"] is not None and float(a["npv"]) > 0)

# TOP10_PAYBACK is used only for the top table total row (includes any padded NPV- rows)
# PAYBACK_TIME (from Excel) is used for the summary page — always NPV+ assets only
TOP10_PAYBACK = t10_simple_payback if t10_simple_payback > 0 else TOP10_PAYBACK
TOP10_NPV     = t10_npv
TOP10_INVEST  = t10_invest

# ── XML helpers ───────────────────────────────────────────────────────────────
def xe(s):
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def get_tr_pos(xml):
    pos = []
    for m in re.finditer(r'<w:tr[ >]', xml):
        s = m.start()
        e = xml.find('</w:tr>', s)
        if e == -1: continue
        pos.append((s, e + len('</w:tr>')))
    return pos

def nth_wt(xml_str, n, new_val):
    ms = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', xml_str))
    if n >= len(ms): return xml_str
    m = ms[n]
    return xml_str[:m.start()] + m.group(1) + xe(new_val) + m.group(3) + xml_str[m.end():]

def strip_ids(s):
    for attr in ['w14:paraId','w14:textId','w:rsidR','w:rsidRPr',
                 'w:rsidRDefault','w:rsidTr','w:rsidP','w:rsidDel']:
        s = re.sub(f' {re.escape(attr)}="[^"]*"', '', s)
    return s

def fill_row(tmpl, vals):
    row = strip_ids(tmpl)
    ms  = list(re.finditer(r'(<w:t[^>]*>)([^<]*)(</w:t>)', row))
    vals = list(vals)[:len(ms)]
    for i, v in reversed(list(enumerate(vals))):
        m = ms[i]
        row = row[:m.start()] + m.group(1) + xe(v) + m.group(3) + row[m.end():]
    return row

def repl(xml, old, new):
    o, n = xe(old), xe(new)
    xml = xml.replace(f'<w:t>{o}</w:t>', f'<w:t>{n}</w:t>')
    xml = xml.replace(f'<w:t xml:space="preserve">{o}</w:t>',
                      f'<w:t xml:space="preserve">{n}</w:t>')
    return xml

def repl_in_file(content, old, new):
    """Replace in any file content (e.g. chart XML)."""
    return content.replace(xe(old), xe(new)).replace(old, new)

def find_para_start(xml, search_text, after=0):
    """Return the start of the <w:p...> that contains search_text."""
    idx = xml.find(search_text, after)
    if idx == -1: return -1
    return xml.rfind('<w:p ', 0, idx)

def find_para_end(xml, para_start):
    """Return the char position just after the </w:p> that starts at para_start."""
    end = xml.find('</w:p>', para_start)
    return end + len('</w:p>') if end != -1 else -1

# ── Load template ─────────────────────────────────────────────────────────────
print(f"Reading template: {os.path.basename(TEMPLATE_PATH)}")
with open(TEMPLATE_PATH, 'rb') as f:
    tmpl_bytes = f.read()
zin = zipfile.ZipFile(io.BytesIO(tmpl_bytes), 'r')
all_files = {name: zin.read(name) for name in zin.namelist()}
zin.close()

try:    xml = all_files['word/document.xml'].decode('utf-8')
except: xml = all_files['word/document.xml'].decode('latin-1')

# ── Shared helper: strip explicit row heights before saving ──────────────────
def _finalize_xml(xml_str):
    """Remove explicit <w:trHeight> elements — lets Word auto-size rows."""
    return re.sub(r'<w:trHeight[^/]*/>', '', xml_str)


# ── This is the Standard pipeline ────────────────────────────────────────────
print("  Template: Standard")
print("Updating scalar values...")
for old, new in [
    ('ION EXCHANGE',              CUSTOMER),
    ('India',                     PLANT),
    ('Customer Input',            DATA_SOURCE),
    ('09.19.2025',                RPT_DATE),
    ('11 / 18',                   f"{NPV_POS_CNT} / {NA}"),
    ('8.00 INR/kWh',              f"{float(ELEC_PRICE):.2f} {CURRENCY}/kWh"),
    ('0.54 kg CO2/kWh',           f"{CO2_INTENSITY} kg CO2/kWh"),
    ('34.9%',                     pct_str(TAX_RATE)),
    ('INR 3,312,551',             f"{SYM} {fmt(ANNUAL_SAVINGS)}"),
    ('11/18',                     f"{NPV_POS_CNT}/{NA}"),
    ('4.7 yrs',                   f"{fmtyrs(PAYBACK_TIME if (not NO_NPV_MODE and NPV_POS_CNT < NT) else TOP10_PAYBACK)} yrs"),
    # When NPV ≤ 0: show Investment cost in the NPV cell and rename label
    ('INR 12,015,351',            f"{SYM} {fmt(t10_invest)}" if float(TOP10_NPV or 0) <= 0 else f"{SYM} {fmt(TOP10_NPV)}"),
    ('Net Present Value*',        'Investment cost*' if float(TOP10_NPV or 0) <= 0 else 'Net Present Value*'),
    ('INR 15,679,593',            f"{SYM} {fmt(INVEST_COST)}"),
    ('16%',                       fmtirr(TOP10_IRR)),
    ('224 tCO2',                  f"{round(float(CO2_SAVINGS or 0)/1000)} tCO2"),
    ('118 Vehicles\xa0',          f"{round(float(BEV_COUNT or 0))} Vehicles\xa0"),
    ('Energy Cost (INR)',          f"Energy Cost ({CURRENCY})"),
    ('Investment (INR)',           f"Investment ({CURRENCY})"),
    ('Energy Cost Savings (INR)',  f"Energy Cost Savings ({CURRENCY})"),
]:
    xml = repl(xml, old, new)

# Replace standalone INR in appendix column headers
xml = re.sub(r'(<w:t[^>]*>)(INR)(</w:t>)',
             lambda m: m.group(1) + xe(CURRENCY) + m.group(3), xml)

# Update "Energy savings with ABB premium efficiency solutions – Top 10" heading
xml = repl(xml,
    'Energy savings with ABB premium efficiency solutions \u2013 Top 10',
    f'Energy savings with ABB premium efficiency solutions \u2013 Top {NT}')

# ── Change 1: Remove "Top 10" from summary payback/IRR labels when NA <= 10 ──
# Misleading to say "Top 10" when there are <= 10 total assets
if NA <= 10:
    xml = repl(xml, 'Payback time \u2013 Top 10*', 'Payback time*')
    xml = repl(xml, 'Internal rate of return \u2013 Top 10*', 'Internal rate of return*')
    print(f"  Summary: 'Top 10' removed from labels (NA={NA} \u2264 10)")
else:
    print(f"  Summary: 'Top 10' kept in labels (NA={NA} > 10)")

# ── Step 2: Footer dates (BUG FIX #1) ────────────────────────────────────────
print(f"Updating footer dates: 2025-09-19 → {RPT_DATE_ISO}")
for fname in ['word/footer1.xml', 'word/footer2.xml',
              'word/footer3.xml', 'word/header1.xml',
              'word/header2.xml', 'word/header3.xml']:
    if fname in all_files:
        content = all_files[fname]
        try:    txt = content.decode('utf-8')
        except: txt = content.decode('latin-1')
        txt = txt.replace('2025-09-19', RPT_DATE_ISO)
        all_files[fname] = txt.encode('utf-8')

# ── Step 3: Charts – NPV-positive data, fully editable (FIX #2 & #3) ─────────
# Strategy: replace <c:numRef><c:f>ExternalSheet!$X$Y</c:f><c:numCache>...</c:numCache></c:numRef>
# with     <c:numLit>...</c:numLit>  — self-contained literal values, always editable in Word.

cons_before = float(CONSUMP_BEFORE or 0)     # row 16 col 3 = "Consumption before with NPV+ve"
cons_after  = cons_before - float(SAVINGS_KWH or 0)  # NPV+ consumption after upgrading

def numlit_1pt(fmt_code, value):
    """Build a <c:numLit> block for a single data point."""
    return (f'<c:numLit>'
            f'<c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="1"/>'
            f'<c:pt idx="0"><c:v>{value:.10f}</c:v></c:pt>'
            f'</c:numLit>')

def numlit_5pt(fmt_code, values):
    """Build a <c:numLit> block for 5 data points."""
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v:.10f}</c:v></c:pt>' for i, v in enumerate(values))
    return (f'<c:numLit>'
            f'<c:formatCode>{fmt_code}</c:formatCode>'
            f'<c:ptCount val="5"/>'
            f'{pts}'
            f'</c:numLit>')

# Chart 1: Energy Consumption bar chart — pick unit based on magnitude
# GWh (≥ 1 GWh = 1,000,000 kWh): formatCode="#.0,,\ "GWh""  (,,  = ÷1,000,000)
# MWh (≥ 1 MWh = 1,000 kWh):    formatCode="#.0,\ "MWh""   (,   = ÷1,000)
# kWh (< 1,000 kWh):             formatCode="#,##0\ "kWh""  (no division)
if cons_before >= 1_000_000:
    unit_label = 'GWh'
    datalabel_fmt = r'#.0,,\ &quot;GWh&quot;'   # format code stored in XML (& escaped)
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'    # raw kWh stored in numLit; chart divides
    disp_before   = f'{cons_before/1e6:.1f} GWh'
    disp_after    = f'{cons_after/1e6:.1f} GWh'
elif cons_before >= 1_000:
    unit_label = 'MWh'
    datalabel_fmt = r'#.0,\ &quot;MWh&quot;'
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'
    disp_before   = f'{cons_before/1e3:.1f} MWh'
    disp_after    = f'{cons_after/1e3:.1f} MWh'
else:
    unit_label = 'kWh'
    datalabel_fmt = r'#,##0\ &quot;kWh&quot;'
    numlit_fmt    = r'#,##0_);[Red]\(#,##0\)'
    disp_before   = f'{cons_before:.0f} kWh'
    disp_after    = f'{cons_after:.0f} kWh'

if 'word/charts/chart1.xml' in all_files:
    ctxt = all_files['word/charts/chart1.xml'].decode('utf-8')
    # 1. Replace data values (numRef → numLit, raw kWh)
    ctxt = re.sub(
        r'<c:numRef><c:f>Saving_Calculations!\$C\$16</c:f>.*?</c:numRef>',
        numlit_1pt(numlit_fmt, cons_before),
        ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:numRef><c:f>Saving_Calculations!\$C\$18</c:f>.*?</c:numRef>',
        numlit_1pt(numlit_fmt, cons_after),
        ctxt, flags=re.DOTALL)
    # 2. Update both data-label format codes (GWh → chosen unit)
    ctxt = ctxt.replace(
        '#.0,,\\ &quot;GWh&quot;',
        datalabel_fmt)
    all_files['word/charts/chart1.xml'] = ctxt.encode('utf-8')
    print(f"  Chart 1 (Energy Consumption): {disp_before} → {disp_after}  [{unit_label}, editable]")

# Chart 2: NPV positive pie chart — update embedded Excel workbook
if 'word/embeddings/Microsoft_Excel_Worksheet.xlsx' in all_files:
    import openpyxl
    emb_bytes = all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx']
    ewb = openpyxl.load_workbook(io.BytesIO(emb_bytes))
    ews = ewb.active
    ews['B2'] = NPV_POS_CNT                # NPV positive count
    ews['B3'] = NA - NPV_POS_CNT           # non-NPV+ count
    ews['A2'] = 'NPV Positive'
    ews['A3'] = 'Remaining Assets'
    buf2 = io.BytesIO()
    ewb.save(buf2)
    all_files['word/embeddings/Microsoft_Excel_Worksheet.xlsx'] = buf2.getvalue()
    # Also update the cached values in chart2.xml
    if 'word/charts/chart2.xml' in all_files:
        c2txt = all_files['word/charts/chart2.xml'].decode('utf-8')
        c2txt = c2txt.replace('<c:v>192</c:v>', f'<c:v>{NPV_POS_CNT}</c:v>')
        c2txt = c2txt.replace('<c:v>6</c:v>',   f'<c:v>{NA - NPV_POS_CNT}</c:v>')
        c2txt = c2txt.replace('<c:v>2nd Qtr</c:v>', '<c:v>Remaining Assets</c:v>')
        all_files['word/charts/chart2.xml'] = c2txt.encode('utf-8')
    print(f"  Chart 2 (NPV pie): {NPV_POS_CNT} NPV+ / {NA - NPV_POS_CNT} remaining  [editable]")

# Chart 3: Payback sensitivity bar chart (NPV+ based) — dual-pass update
# Pass 1: replace known formula refs with numLit literals (handles standard templates)
# Pass 2: update numCache values directly (handles templates with different/no formula refs)
if 'word/charts/chart3.xml' in all_files and SENSITIVITY:
    ctxt = all_files['word/charts/chart3.xml'].decode('utf-8')
    # Pass 1: formula ref → numLit (catches any sheet-name variant)
    ctxt = re.sub(
        r'<c:numRef><c:f>[^<]*\$T\$19:\$T\$23</c:f>.*?</c:numRef>',
        numlit_5pt('0%', [s[0] for s in SENSITIVITY]),
        ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:numRef><c:f>[^<]*\$V\$19:\$V\$23</c:f>.*?</c:numRef>',
        numlit_5pt('0.00', [s[1] for s in SENSITIVITY]),
        ctxt, flags=re.DOTALL)
    ctxt = re.sub(
        r'<c:strRef><c:f>[^<]*\$S\$18[^<]*</c:f>.*?</c:strRef>',
        '<c:v>Payback time sensitivity to electricity price</c:v>',
        ctxt, flags=re.DOTALL)
    # Pass 2: directly overwrite numCache in every <c:numRef> block so Word
    # renders the correct values even when Pass 1 regex didn't match
    _refs = list(re.finditer(r'<c:numRef>', ctxt))
    if len(_refs) >= 2:
        # 2nd numRef = Y-axis (payback values)
        _rs     = _refs[1].start()
        _re_end = ctxt.find('</c:numRef>', _rs) + len('</c:numRef>')
        _seg    = ctxt[_rs:_re_end]
        _new_pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>'
                           for i, (_, v) in enumerate(SENSITIVITY))
        _seg = re.sub(r'<c:ptCount[^/]*/>', f'<c:ptCount val="{len(SENSITIVITY)}"/>', _seg)
        _seg = re.sub(r'(<c:ptCount[^/]*/>)(.*?)(?=</c:numCache>)',
                      lambda m: m.group(1) + _new_pts, _seg, flags=re.DOTALL)
        ctxt = ctxt[:_rs] + _seg + ctxt[_re_end:]
    # Remove external data link — prevents Word from refreshing chart on open
    ctxt = re.sub(r'<c:externalData\b[^>]*/>', '', ctxt)
    ctxt = re.sub(r'<c:externalData\b.*?</c:externalData>', '', ctxt, flags=re.DOTALL)
    all_files['word/charts/chart3.xml'] = ctxt.encode('utf-8')
    _rels_name = 'word/charts/_rels/chart3.xml.rels'
    if _rels_name in all_files:
        _rels = all_files[_rels_name].decode('utf-8')
        _rels = re.sub(r'<Relationship\b[^>]*Id="rId4"[^/]*/>', '', _rels)
        all_files[_rels_name] = _rels.encode('utf-8')
    print(f"  Chart 3 (Payback sensitivity): {[round(s[1],2) for s in SENSITIVITY]}  [editable]")

# ── Step 5a: Summary page — remove IRR row when NPV ≤ 0 ─────────────────────
# NPV/Investment swap is handled above in the scalar map.
# IRR row (TR 15) is always at index 15 (before top-N table); remove when NPV ≤ 0.
NO_NPV = float(TOP10_NPV or 0) <= 0
if NO_NPV:
    tr_pre = get_tr_pos(xml)
    xml = xml[:tr_pre[15][0]] + xml[tr_pre[15][1]:]
    print(f"  Summary: NPV ≤ 0 — Investment cost shown, IRR row removed")

# irr_adj: -1 when IRR row (TR 15) was removed, shifting all subsequent TR indices
irr_adj = -1 if NO_NPV else 0

# ── Step 5: Top-N data rows ───────────────────────────────────────────────────
print(f"Building Top-{NT} table ({NT} rows)...")
tr = get_tr_pos(xml)
tmpl_t10 = xml[tr[19 + irr_adj][0]:tr[19 + irr_adj][1]]

def v_top10(a):
    row = [str(a["num"]), a["tag"],
           fmt(a["e_cons"]), fmt(a["e_cost"]), fmt(a["co2_cons"]),
           fmt(a["e_sav_kwh"]), fmt(a["invest"]), fmt(a["e_sav_cost"]),
           fmtpct(a["e_sav_pct"]),
           f"{fmtyrs(a['payback'])} " if a["payback"] else "",
           f"{fmt(a['co2_sav'])} "]
    row.append("\u2713")   # Take-Back ✓
    return row

def build_top10_row(a):
    """Fill a top-table row; apply white shading to NPV-negative assets."""
    row_xml = fill_row(tmpl_t10, v_top10(a))
    is_npv_neg = (a["npv"] is None or float(a["npv"] or 0) <= 0) and not NO_NPV_MODE
    if is_npv_neg:
        # Cells 1-10: white; cells 11-12 (CO2 + Take-Back ✓): keep C5E0B3
        tc_starts = [m.start() for m in re.finditer(r'<w:tc>', row_xml)]
        if len(tc_starts) >= 11:
            cut = tc_starts[10]   # start of 11th cell (0-indexed)
            before = row_xml[:cut].replace('w:fill="E2EFD9"', 'w:fill="FFFFFF"')
            row_xml = before + row_xml[cut:]
    return row_xml

block = "".join(build_top10_row(a) for a in top10)
# Replace original 10 data rows with NT rows
xml = xml[:tr[19 + irr_adj][0]] + block + xml[tr[28 + irr_adj][1]:]
tr  = get_tr_pos(xml)

# Dynamic index formulas — all relative to NT and irr_adj
I_TR1       = 19 + NT + irr_adj   # "Total Energy- Top N" row
I_TR2       = 20 + NT + irr_adj   # "Total NPV Positive Assets" row
I_APP_FIRST = 22 + NT + irr_adj   # first App Details data row
I_APP_LAST  = 31 + NT + irr_adj   # last of 10 original App Details rows

# ── Step 6: Total rows (dynamic indices) ─────────────────────────────────────
r29 = strip_ids(xml[tr[I_TR1][0]:tr[I_TR1][1]])
r29 = nth_wt(r29, 0, f"Total Energy- Top {NT}")
# node 1 = "Saving Assets" (keep as-is)
r29 = nth_wt(r29, 2,  fmt(t10_e_cons))
r29 = nth_wt(r29, 3,  fmt(t10_e_cost))
r29 = nth_wt(r29, 4,  fmt(t10_co2_c))
r29 = nth_wt(r29, 5,  fmt(t10_sav_kwh))
r29 = nth_wt(r29, 6,  fmt(t10_invest))
r29 = nth_wt(r29, 7,  fmt(t10_sav_c))
r29 = nth_wt(r29, 8,  fmtpct(t10_pct))
r29 = nth_wt(r29, 9,  fmtyrs(TOP10_PAYBACK))
r29 = nth_wt(r29, 10, fmt(t10_co2_sav))
xml = xml[:tr[I_TR1][0]] + r29 + xml[tr[I_TR1][1]:]
tr  = get_tr_pos(xml)

nc  = float(CONSUMP_BEFORE or 0)
# Show "Total NPV Positive Assets" row whenever its count differs from top-N count
# This covers: NPV+ > NT (more assets than shown) AND NPV+ < NT (padded with NPV-)
if NPV_POS_CNT != NT and NPV_POS_CNT > 0:
    r30 = strip_ids(xml[tr[I_TR2][0]:tr[I_TR2][1]])
    r30 = nth_wt(r30, 0, f"Total NPV Positive Assets ({NPV_POS_CNT})")
    r30 = nth_wt(r30, 1, fmt(nc))
    r30 = nth_wt(r30, 2, fmt(nc * float(ELEC_PRICE or 8)))
    r30 = nth_wt(r30, 3, fmt(nc * float(CO2_INTENSITY or 0.54)))
    r30 = nth_wt(r30, 4, fmt(float(SAVINGS_KWH or 0)))
    r30 = nth_wt(r30, 5, fmt(float(INVEST_COST or 0)))
    r30 = nth_wt(r30, 6, fmt(float(ANNUAL_SAVINGS or 0)))
    r30 = nth_wt(r30, 7, fmtpct(float(SAVINGS_KWH or 0)/nc if nc else 0))
    r30 = nth_wt(r30, 8, fmtyrs(PAYBACK_TIME))
    r30 = nth_wt(r30, 9, fmt(float(CO2_SAVINGS or 0)))
    xml = xml[:tr[I_TR2][0]] + r30 + xml[tr[I_TR2][1]:]
    npv_adj = 0   # NPV row kept → no index shift
else:
    xml = xml[:tr[I_TR2][0]] + xml[tr[I_TR2][1]:]
    npv_adj = -1  # NPV row removed → all subsequent indices shift by -1
tr  = get_tr_pos(xml)

# ── Step 7: Application Details top-N ────────────────────────────────────────
def v_app(a):
    return [str(a["num"]), a["tag"], a["ie"], a["load"],
            a["flow_ctrl"], a["connection"],
            str(a["output_kw"]), str(a["shaft_h"]),
            str(a["run_hrs"]), str(a["avg_val"]),
            a["ess_motor"], a["ess_conn"]]

# Adjusted indices: I_APP_FIRST/LAST were computed before step 6;
# npv_adj accounts for NPV row removal, irr_adj for IRR row removal
I_APP_FIRST_a = I_APP_FIRST + npv_adj
I_APP_LAST_a  = I_APP_LAST  + npv_adj

# Always build App Details motor table for top-N assets
print(f"Building Application Details ({NT} rows)...")
tmpl_app = xml[tr[I_APP_FIRST_a][0]:tr[I_APP_FIRST_a][1]]
block = "".join(fill_row(tmpl_app, v_app(a)) for a in top10)
xml = xml[:tr[I_APP_FIRST_a][0]] + block + xml[tr[I_APP_LAST_a][1]:]
tr  = get_tr_pos(xml)

# ── Step 8: Appendix all-assets ───────────────────────────────────────────────
# Always use normal formula (App Details was always built)
I_APP2_FIRST = 23 + 2*NT + irr_adj + npv_adj
I_APP2_LAST  = I_APP2_FIRST + 17

print(f"Building Appendix ({NA} rows)...")
tmpl_ap2 = xml[tr[I_APP2_FIRST][0]:tr[I_APP2_FIRST][1]]

def v_app2(a):
    npv = a["npv"]
    if npv is not None:
        npv_s = (fmt(npv)+" ") if float(npv) > 0 else fmt_parens(npv)
    else:
        npv_s = ""
    co2_r = float(a["co2_sav"] or 0)
    co2_s = (fmt(co2_r)+" ") if co2_r > 0 else fmt_parens(co2_r)
    sav_e = fmt(a["e_sav_kwh"])  if float(a["e_sav_kwh"]  or 0) >= 0 else f"-{fmt(abs(float(a['e_sav_kwh'])))}"
    sav_c = fmt(a["e_sav_cost"]) if float(a["e_sav_cost"] or 0) >= 0 else f"-{fmt(abs(float(a['e_sav_cost'])))}"
    return [str(a["num"]), a["tag"],
            fmt(a["e_cons"]), fmt(a["e_cost"]), fmt(a["co2_cons"]),
            sav_e, fmt(a["invest"]), sav_c,
            f"{fmtyrs(a['payback'])} " if a["payback"] else "",
            co2_s, fmtpct(a["e_sav_pct"]), npv_s]

block = "".join(fill_row(tmpl_ap2, v_app2(a)) for a in assets_by_num)
xml = xml[:tr[I_APP2_FIRST][0]] + block + xml[tr[I_APP2_LAST][1]:]
tr  = get_tr_pos(xml)

# ── Step 9: Details of Recommendation — always build ─────────────────────────
I_DET_HDR   = 23 + 2*NT + irr_adj + npv_adj + NA
I_DET_FIRST = I_DET_HDR + 1
I_DET_LAST  = I_DET_HDR + 18
print(f"Building Details of Recommendation ({NA} rows, TR {I_DET_FIRST}-{I_DET_FIRST+NA-1})...")
tmpl_det = xml[tr[I_DET_FIRST][0]:tr[I_DET_FIRST][1]]
block_det = "".join(fill_row(tmpl_det, v_app(a)) for a in assets_by_num)
xml = xml[:tr[I_DET_FIRST][0]] + block_det + xml[tr[I_DET_LAST][1]:]

# ── Step 10: Trim Appendix — keep only Calculation Methodology + NPV text ────
# Remove: "Application Details" heading + table + "Details of Recommendation"
#          heading + table (everything between Appendix heading and Calc Methodology)
# Keep:   Calculation Methodology table + NPV Methodology paragraphs
#
# Strategy: find end of "Appendix" heading paragraph, then cut forward to the
# <w:p> containing "Calculation Methodology" (search after the Appendix para).
# This removes ALL content between them regardless of what was left by steps 7-9.

def find_para_start(xml, search_text, after=0):
    """Return the start of the <w:p...> that contains search_text."""
    idx = xml.find(search_text, after)
    if idx == -1: return -1
    return xml.rfind('<w:p ', 0, idx)

def find_para_end(xml, para_start):
    """Return the char position just after the </w:p> that starts at para_start."""
    end = xml.find('</w:p>', para_start)
    return end + len('</w:p>') if end != -1 else -1

# Find the "Appendix" heading paragraph in the BODY (not the TOC).
# Anchor: "*Data is listed as total annual figures" always appears just before
# the Appendix heading in the body, giving us a reliable body-only anchor.
data_listed_pos = xml.rfind('Data is listed as total annual')
if data_listed_pos == -1:
    data_listed_pos = 0  # fallback
appendix_para_start = find_para_start(xml, '>Appendix<', data_listed_pos)
appendix_para_end   = find_para_end(xml, appendix_para_start)

# Find "Calculation Methodology" paragraph AFTER the body Appendix heading
calcmeth_para = find_para_start(xml, '>Calculation Methodology<', appendix_para_end)

if NA <= 10 and appendix_para_end > 0 and calcmeth_para > appendix_para_end:
    # 10 or fewer assets: cut App Det + Det Rec tables from appendix body
    xml = xml[:appendix_para_end] + xml[calcmeth_para:]
    print(f"Appendix trimmed (NA={NA} <= 10): kept only Calculation Methodology + NPV Methodology")
elif NA > 10:
    print(f"Appendix kept full (NA={NA} > 10): App Det + Det Rec tables present")
else:
    print(f"WARNING: Could not trim appendix (apx_end={appendix_para_end}, calc={calcmeth_para})")

# ── Save output docx ──────────────────────────────────────────────────────────
safe_c   = re.sub(r'[\\/:*?"<>|]', '_', args[1])
safe_p   = re.sub(r'[\\/:*?"<>|]', '_', PLANT)
out_name = f"{safe_c}_{safe_p}_EA_Report{OUTPUT_SUFFIX}.docx"
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), out_name)

# TOC auto-update: add <w:updateFields> to settings.xml
all_files['word/document.xml'] = _finalize_xml(xml).encode('utf-8')
if 'word/settings.xml' in all_files:
    try:
        stg = all_files['word/settings.xml'].decode('utf-8')
        if 'updateFields' not in stg:
            stg = stg.replace('</w:settings>',
                '<w:updateFields w:val="1"/></w:settings>')
        all_files['word/settings.xml'] = stg.encode('utf-8')
    except Exception: pass

buf = io.BytesIO()
with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
    for name, data in all_files.items():
        zout.writestr(name, data)
with open(out_path, 'wb') as f:
    f.write(buf.getvalue())

print()
print(f"  Done!  →  {out_path}")
print(f"  {CUSTOMER} | {PLANT} | {CURRENCY} | {NA} assets | {NPV_POS_CNT} NPV+ | {NT} Top-N")
print(f"  Annual savings : {SYM} {fmt(ANNUAL_SAVINGS)}")
print(f"  Investment     : {SYM} {fmt(INVEST_COST)}")
print(f"  Payback        : {fmtyrs(PAYBACK_TIME)} yrs (Top-{NT}: {fmtyrs(TOP10_PAYBACK)} yrs)")
print(f"  NPV            : {SYM} {fmt(NPV_VALUE)}")
print(f"  IRR Top-{NT}      : {fmtirr(TOP10_IRR)}")
if SENSITIVITY:
    print(f"  Payback sensitivity: {[f'{s[0]:+.0%}→{s[1]:.1f}yrs' for s in SENSITIVITY]}")
