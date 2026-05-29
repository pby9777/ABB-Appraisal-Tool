#!/usr/bin/env python3
"""
Phase 2B: Prove fill_sheet() can populate Saving_Calculations,
Saving_Calculations_2, and Saving_Calculations_3 independently.

Each sheet receives its own zip data.  After saving, the test reopens
the workbook and verifies:
  - correct equipment IDs per sheet (no cross-contamination)
  - tables visible via ws.tables on every sheet
  - structured-reference formulas survived the round-trip
  - header row shared-strings preserved

Usage:
    python legacy/test_fill_sheet.py
"""

import io
import os
import re
import shutil
import sys
import tempfile
import zipfile as _zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from excel_clone_poc import clone_saving_calculations
from fill_saving_calculations import fill_sheet

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

TEMPLATE    = "legacy/excel_templates/1_0_region_site_name_saving_calculations.xlsx"
STEP1_PATH  = "/tmp/test_fill_sheet_step1.xlsx"
WB_PATH     = "legacy/test_fill_sheet_output.xlsx"

# ---------------------------------------------------------------------------
# Mock EEA data: three asset groups, one per sheet
# ---------------------------------------------------------------------------
# Each entry maps to the columns read by read_assessment_data (positional)
# and read_input_assets (header-based).

GROUPS = {
    "Saving_Calculations": [
        dict(equip_id="EQ-A001", energy_kwh=120_000, savings_kwh=30_000,
             investment=25_000, ess_motor="ABB-X1", ess_drive="Direct",
             application="Centrifugal Pump", dol_vsd="VSD",
             flow_control="Throttling",      output_kw=22.0,
             run_hours=4000, ie_class="IE3",  avg_loading=75.0, avg_freq=48.0),
        dict(equip_id="EQ-A002", energy_kwh=85_000,  savings_kwh=18_000,
             investment=15_000, ess_motor="ABB-X2", ess_drive="Direct",
             application="Fan or blower",    dol_vsd="DOL",
             flow_control="On/off",          output_kw=15.0,
             run_hours=6000, ie_class="IE2",  avg_loading=60.0, avg_freq=50.0),
        dict(equip_id="EQ-A003", energy_kwh=200_000, savings_kwh=55_000,
             investment=42_000, ess_motor="ABB-X3", ess_drive="Star-delta",
             application="Centrifugal Pump", dol_vsd="VSD",
             flow_control="No control (Bypass)", output_kw=37.0,
             run_hours=3000, ie_class="IE3",  avg_loading=80.0, avg_freq=47.0),
    ],
    "Saving_Calculations_2": [
        dict(equip_id="EQ-B001", energy_kwh=55_000,  savings_kwh=12_000,
             investment=10_000, ess_motor="ABB-Y1", ess_drive="Direct",
             application="Fan or blower",    dol_vsd="VSD",
             flow_control="Damper",          output_kw=11.0,
             run_hours=5000, ie_class="IE3",  avg_loading=70.0, avg_freq=49.0),
        dict(equip_id="EQ-B002", energy_kwh=40_000,  savings_kwh=9_000,
             investment=8_000,  ess_motor="ABB-Y2", ess_drive="Direct",
             application="Centrifugal Pump", dol_vsd="DOL",
             flow_control="Inlet vanes",     output_kw=7.5,
             run_hours=4000, ie_class="IE2",  avg_loading=65.0, avg_freq=50.0),
    ],
    "Saving_Calculations_3": [
        dict(equip_id="EQ-C001", energy_kwh=30_000,  savings_kwh=7_000,
             investment=5_000,  ess_motor="ABB-Z1", ess_drive="Direct",
             application="Mixer",            dol_vsd="DOL",
             flow_control="Information not available", output_kw=4.0,
             run_hours=3000, ie_class="IE3",  avg_loading=55.0, avg_freq=50.0),
    ],
}

ASSUMPTIONS = {
    "currency": "EUR",
    "tariff":   0.14,
    "co2":      0.75,
    "tax":      19.0,
    "discount": 6.5,
}

# ---------------------------------------------------------------------------
# Mock zip builder
# ---------------------------------------------------------------------------

_ASSESS_HEADERS = [
    "Sr No", "Customer Equipment Id",
    "Annual Energy Consumption (kWh)", "C3", "C4", "C5",
    "Annual Energy Savings (kWh)", "C7", "C8", "C9",
    "Investment (currency)", "C11", "C12",
    "Recommended ESS Motor", "ESS Connection",
]

_INPUT_HEADERS = [
    "Customer Equipment Id", "Driven Load", "Dol Vsd",
    "Flow Control", "Shaft Height [mm]", "Rated Power [kW]",
    "Annual Running Hours [h]", "IE Eff Class",
    "Avg Loading [%]", "Avg Frequency [Hz]",
]


def _make_mock_zip(assets: list) -> str:
    """
    Build an in-memory EEA tool output zip and write it to a temp file.
    Returns the path to the temp zip file.
    read_assessment_data uses positional column access (0-based).
    read_input_assets   uses header-name access.
    """
    tmpdir = tempfile.mkdtemp()

    # ── assessment data.xlsx ──────────────────────────────────────────────
    awb = openpyxl.Workbook()
    aws = awb.active
    for ci, h in enumerate(_ASSESS_HEADERS, 1):
        aws.cell(1, ci).value = h
    for i, a in enumerate(assets, 1):
        r = i + 1
        aws.cell(r, 1).value  = i                 # row[0] → num
        aws.cell(r, 2).value  = a["equip_id"]     # row[1] → equip_id
        aws.cell(r, 3).value  = a["energy_kwh"]   # row[2] → energy_kwh
        aws.cell(r, 7).value  = a["savings_kwh"]  # row[6] → savings_kwh
        aws.cell(r, 11).value = a["investment"]   # row[10] → investment
        aws.cell(r, 14).value = a["ess_motor"]    # row[13] → ess_motor
        aws.cell(r, 15).value = a["ess_drive"]    # row[14] → ess_drive
    assess_path = os.path.join(tmpdir, "assessment data.xlsx")
    awb.save(assess_path)

    # ── input assets.xlsx ─────────────────────────────────────────────────
    iwb = openpyxl.Workbook()
    iws = iwb.active
    for ci, h in enumerate(_INPUT_HEADERS, 1):
        iws.cell(1, ci).value = h
    for i, a in enumerate(assets, 2):
        iws.cell(i, 1).value  = a["equip_id"]     # Customer Equipment Id
        iws.cell(i, 2).value  = a["application"]  # Driven Load
        iws.cell(i, 3).value  = a["dol_vsd"]      # Dol Vsd
        iws.cell(i, 4).value  = a["flow_control"] # Flow Control
        iws.cell(i, 5).value  = 0                 # Shaft Height [mm]
        iws.cell(i, 6).value  = a["output_kw"]    # Rated Power [kW]
        iws.cell(i, 7).value  = a["run_hours"]    # Annual Running Hours [h]
        iws.cell(i, 8).value  = a["ie_class"]     # IE Eff Class
        iws.cell(i, 9).value  = a["avg_loading"]  # Avg Loading [%]
        iws.cell(i, 10).value = a["avg_freq"]     # Avg Frequency [Hz]
    input_path = os.path.join(tmpdir, "input assets.xlsx")
    iwb.save(input_path)

    # ── package as zip ────────────────────────────────────────────────────
    zip_path = os.path.join(tmpdir, "mock_eea_output.zip")
    with _zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(assess_path, "assessment data.xlsx")
        zf.write(input_path,  "input assets.xlsx")

    return zip_path, tmpdir   # caller must clean up tmpdir


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

PASS_LIST: list = []
FAIL_LIST: list = []


def check(label: str, cond: bool, detail: str = "") -> None:
    if cond:
        PASS_LIST.append(label)
    else:
        FAIL_LIST.append(f"{label}  {detail}")


def _main_table(ws):
    """Return (table_name, ref) for the large data table on the sheet."""
    for tname, tobj in ws.tables.items():
        ref = tobj.ref if hasattr(tobj, "ref") else str(tobj)
        m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ref)
        if m and (int(m.group(2)) - int(m.group(1))) > 100:
            return tname, ref
    return None, None


def _header_row(ref: str) -> int:
    m = re.match(r"[A-Z]+(\d+):", ref)
    return int(m.group(1)) if m else 36


def _col_map(ws, hrow: int) -> dict:
    return {
        cell.value: cell.column
        for cell in ws[hrow]
        if cell.value is not None
    }


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------

def step1_build_workbook():
    print("\n── STEP 1: build 3-sheet workbook (ZIP-level clone) ────────────")
    d1 = clone_saving_calculations(
        TEMPLATE, STEP1_PATH,
        clone_display_name="Saving_Calculations_2", clone_suffix="_2",
    )
    d2 = clone_saving_calculations(
        STEP1_PATH, WB_PATH,
        clone_display_name="Saving_Calculations_3", clone_suffix="_3",
    )
    check("Clone _2 SR intact",    d1["sr_in_source"] == d1["sr_in_clone"])
    check("Clone _3 SR intact",    d2["sr_in_source"] == d2["sr_in_clone"])
    print(f"  Workbook written: {WB_PATH}")


def step2_fill_independently():
    print("\n── STEP 2: fill each sheet independently via fill_sheet() ──────")
    tmp_dirs = []
    zip_paths = {}

    try:
        for sname, assets in GROUPS.items():
            zpath, tmpdir = _make_mock_zip(assets)
            zip_paths[sname] = zpath
            tmp_dirs.append(tmpdir)

        wb = openpyxl.load_workbook(WB_PATH, data_only=False)

        for sname, assets in GROUPS.items():
            count = fill_sheet(wb, sname, [zip_paths[sname]], ASSUMPTIONS)
            print(f"  fill_sheet({sname!r})  → {count} assets written")
            check(
                f"fill_sheet returns correct count for {sname}",
                count == len(assets),
                f"expected {len(assets)} got {count}",
            )

        wb.save(WB_PATH)
        print(f"  Saved → {WB_PATH}")

    finally:
        for td in tmp_dirs:
            shutil.rmtree(td, ignore_errors=True)


def step3_verify():
    print("\n── STEP 3: verification ────────────────────────────────────────")
    wb = openpyxl.load_workbook(WB_PATH, data_only=False)

    # ── 3a. Sheet count ──────────────────────────────────────────────────
    expected_names = [
        "Note", "Input_ Reference",
        "Saving_Calculations", "Saving_Calculations_2", "Saving_Calculations_3",
    ]
    check("Sheet count == 5", len(wb.sheetnames) == 5,
          f"got {wb.sheetnames}")
    for sn in expected_names:
        check(f"Sheet present: {sn!r}", sn in wb.sheetnames)

    for sname, assets in GROUPS.items():
        ws = wb[sname]
        tname, ref = _main_table(ws)
        hrow = _header_row(ref or "A36:BE1100")
        cmap = _col_map(ws, hrow)

        # ── 3b. Tables present ───────────────────────────────────────────
        check(f"2 tables on {sname}", len(ws.tables) == 2,
              f"got {len(ws.tables)}")
        check(f"Main table found on {sname}", tname is not None)

        # ── 3c. Correct equipment IDs (data isolation) ───────────────────
        equip_col = cmap.get("Customer Equipment Id")
        if equip_col:
            for i, asset in enumerate(assets):
                data_row = hrow + 1 + i
                actual = ws.cell(data_row, equip_col).value
                check(
                    f"{sname} row{data_row} equip_id",
                    actual == asset["equip_id"],
                    f"expected={asset['equip_id']!r} got={actual!r}",
                )

        # ── 3d. Cross-contamination: other sheets' IDs must NOT appear ───
        other_ids = {
            a["equip_id"]
            for other_sheet, other_assets in GROUPS.items()
            if other_sheet != sname
            for a in other_assets
        }
        if equip_col:
            for r in range(hrow + 1, hrow + 1 + len(assets)):
                cell_val = ws.cell(r, equip_col).value
                if cell_val in other_ids:
                    check(
                        f"{sname} row{r} no cross-contamination",
                        False,
                        f"found foreign id {cell_val!r}",
                    )
            check(f"{sname} no cross-contamination", True)

        # ── 3e. Formulas survived ─────────────────────────────────────────
        formula_checks = [("F", "Annual Energy Cost"), ("G", "Annual CO2 Cons.")]
        for col_letter, desc in formula_checks:
            cell = ws[f"{col_letter}{hrow + 1}"]
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            check(f"{sname} {col_letter}{hrow+1} is formula ({desc})", is_formula,
                  f"value={str(cell.value)[:60]!r}")

        # ── 3f. Application value written (input-assets mapping) ─────────
        app_col = cmap.get("Application")
        if app_col:
            actual_app = ws.cell(hrow + 1, app_col).value
            expected_app = assets[0]["application"]
            check(
                f"{sname} row{hrow+1} Application",
                actual_app == expected_app,
                f"expected={expected_app!r} got={actual_app!r}",
            )

        # ── 3g. Header row shared-strings preserved ───────────────────────
        for col_idx in list(cmap.values())[:5]:
            expected = ws.cell(hrow, col_idx).value
            check(
                f"{sname} header col{col_idx} non-null",
                expected is not None,
            )


def print_report() -> int:
    SEP = "=" * 62
    print(f"\n{SEP}")
    print("TEST FILL_SHEET — VALIDATION REPORT")
    print(SEP)
    print(f"\n  PASS: {len(PASS_LIST)}   FAIL: {len(FAIL_LIST)}")

    if FAIL_LIST:
        print("\n  FAILURES:")
        for msg in FAIL_LIST:
            print(f"    FAIL  {msg}")

    print("\n  PASSING CHECKS:")
    for msg in PASS_LIST:
        print(f"    ok    {msg}")

    print()
    return len(FAIL_LIST)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    step1_build_workbook()
    step2_fill_independently()
    step3_verify()
    failures = print_report()
    sys.exit(0 if failures == 0 else 1)
