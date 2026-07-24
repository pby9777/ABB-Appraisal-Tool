"""
excel_formula_engine.py
------------------------
A small, purpose-built Excel formula evaluator, scoped to the formula
vocabulary actually used by the Saving_Calculations workbook: arithmetic,
cell/range references, Table structured references ([[#This Row],[Col]]),
and a fixed set of generic Excel functions (SUM, COUNT, COUNTIF, SUMIF,
SUMPRODUCT, IF, IFERROR, ISNUMBER, RANK.EQ, NPV, IRR, CONCATENATE).

This is NOT a general Excel-compatible engine, and it is NOT a
reimplementation of the workbook's financial model: it parses and executes
whatever formula text is actually stored in each cell, using generic,
function-name-level semantics identical to Excel's own definitions of
SUM/IF/NPV/etc. If the workbook's formulas change -- a different discount
period, a different cell reference, an added column -- this module needs
no changes; it evaluates whatever the new formula text says. The workbook's
cells remain the only place the business logic (which columns feed which
calculations, what assumptions apply) is expressed.

Every formula string read via openpyxl is already the fully-resolved,
concrete text for that specific cell (openpyxl resolves Excel's internal
"shared formula" compression on load), so there is no relative/absolute
reference-shifting to implement here -- $ signs are purely cosmetic and are
stripped during tokenizing.

Every token is a 2-tuple (kind, payload); payload is itself a tuple for
multi-part tokens (RANGE, TABLEREF).
"""

import re
from openpyxl.utils import column_index_from_string


# ---------------------------------------------------------------------------
# Tokenizer
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r'^[A-Z]{1,3}\d+$')


def _strip_dollars(word):
    return word.replace('$', '')


def tokenize(text):
    text = text.strip()
    if text.startswith('='):
        text = text[1:]
    tokens = []
    i, n = 0, len(text)
    while i < n:
        c = text[i]
        if c.isspace():
            i += 1
            continue
        if c == '"':
            j = i + 1
            buf = []
            while j < n:
                if text[j] == '"':
                    if j + 1 < n and text[j + 1] == '"':
                        buf.append('"')
                        j += 2
                        continue
                    j += 1
                    break
                buf.append(text[j])
                j += 1
            tokens.append(('STR', ''.join(buf)))
            i = j
            continue
        if c.isdigit() or (c == '.' and i + 1 < n and text[i + 1].isdigit()):
            j = i
            while j < n and (text[j].isdigit() or text[j] == '.'):
                j += 1
            tokens.append(('NUM', float(text[i:j])))
            i = j
            continue
        if c in '+-*/^&(),':
            tokens.append(('OP', c))
            i += 1
            continue
        if c == '<':
            if i + 1 < n and text[i + 1] in '=>':
                tokens.append(('OP', text[i:i + 2]))
                i += 2
            else:
                tokens.append(('OP', '<'))
                i += 1
            continue
        if c == '>':
            if i + 1 < n and text[i + 1] == '=':
                tokens.append(('OP', '>='))
                i += 2
            else:
                tokens.append(('OP', '>'))
                i += 1
            continue
        if c == '=':
            tokens.append(('OP', '='))
            i += 1
            continue
        if c.isalpha() or c == '_' or c == '$':
            j = i
            while j < n and (text[j].isalnum() or text[j] in '_$.'):
                j += 1
            word = text[i:j]
            # Table structured reference: NAME immediately followed by '['
            if j < n and text[j] == '[':
                depth = 0
                k = j
                while k < n:
                    if text[k] == '[':
                        depth += 1
                    elif text[k] == ']':
                        depth -= 1
                        if depth == 0:
                            k += 1
                            break
                    k += 1
                tokens.append(('TABLEREF', (word, text[j:k])))
                i = k
                continue
            bare = _strip_dollars(word)
            # Range: WORD ':' WORD2, only if both sides look like cell refs
            if j < n and text[j] == ':':
                k = j + 1
                m = k
                while m < n and (text[m].isalnum() or text[m] in '_$'):
                    m += 1
                second_bare = _strip_dollars(text[k:m])
                if _CELL_RE.match(bare) and _CELL_RE.match(second_bare):
                    tokens.append(('RANGE', (bare, second_bare)))
                    i = m
                    continue
            if _CELL_RE.match(bare):
                tokens.append(('CELL', bare))
            else:
                tokens.append(('NAME', bare))
            i = j
            continue
        raise ValueError(f"Unexpected character {c!r} in formula: {text!r}")
    tokens.append(('EOF', None))
    return tokens


# ---------------------------------------------------------------------------
# Parser (recursive descent) -> tuple-based AST
# ---------------------------------------------------------------------------

_CELL_SPLIT_RE = re.compile(r'^([A-Z]{1,3})(\d+)$')


def _split_cell(ref):
    m = _CELL_SPLIT_RE.match(ref)
    return m.group(1), int(m.group(2))


def _parse_tableref(table_name, bracket_text):
    inner = bracket_text[1:-1]  # strip outer [ ]
    if inner.startswith('[#This Row]'):
        m = re.search(r'\[([^\]]*)\]\s*$', inner)
        return ('tableref_row', table_name, m.group(1))
    col_name = inner.strip('[]')
    return ('tableref_col', table_name, col_name)


class _Parser:
    def __init__(self, tokens):
        self.tokens = tokens
        self.pos = 0

    def peek(self):
        return self.tokens[self.pos]

    def advance(self):
        tok = self.tokens[self.pos]
        self.pos += 1
        return tok

    def expect_op(self, op):
        kind, val = self.advance()
        if kind != 'OP' or val != op:
            raise ValueError(f"Expected {op!r}, got {(kind, val)!r}")

    def parse(self):
        node = self.parse_comparison()
        if self.peek()[0] != 'EOF':
            raise ValueError(f"Unexpected trailing tokens: {self.tokens[self.pos:]}")
        return node

    def parse_comparison(self):
        left = self.parse_concat()
        while self.peek()[0] == 'OP' and self.peek()[1] in ('=', '<>', '>', '<', '>=', '<='):
            op = self.advance()[1]
            right = self.parse_concat()
            left = ('binop', op, left, right)
        return left

    def parse_concat(self):
        left = self.parse_add()
        while self.peek()[0] == 'OP' and self.peek()[1] == '&':
            self.advance()
            right = self.parse_add()
            left = ('binop', '&', left, right)
        return left

    def parse_add(self):
        left = self.parse_mul()
        while self.peek()[0] == 'OP' and self.peek()[1] in ('+', '-'):
            op = self.advance()[1]
            right = self.parse_mul()
            left = ('binop', op, left, right)
        return left

    def parse_mul(self):
        left = self.parse_unary()
        while self.peek()[0] == 'OP' and self.peek()[1] in ('*', '/'):
            op = self.advance()[1]
            right = self.parse_unary()
            left = ('binop', op, left, right)
        return left

    def parse_unary(self):
        if self.peek()[0] == 'OP' and self.peek()[1] == '-':
            self.advance()
            return ('unary', '-', self.parse_unary())
        return self.parse_pow()

    def parse_pow(self):
        left = self.parse_primary()
        if self.peek()[0] == 'OP' and self.peek()[1] == '^':
            self.advance()
            right = self.parse_unary()
            return ('binop', '^', left, right)
        return left

    def parse_primary(self):
        kind, val = self.peek()
        if kind == 'NUM':
            self.advance()
            return ('num', val)
        if kind == 'STR':
            self.advance()
            return ('str', val)
        if kind == 'CELL':
            self.advance()
            col, row = _split_cell(val)
            return ('cell', col, row)
        if kind == 'RANGE':
            self.advance()
            ref1, ref2 = val
            c1, r1 = _split_cell(ref1)
            c2, r2 = _split_cell(ref2)
            return ('range', c1, r1, c2, r2)
        if kind == 'TABLEREF':
            self.advance()
            table_name, bracket_text = val
            return _parse_tableref(table_name, bracket_text)
        if kind == 'NAME':
            self.advance()
            if self.peek() == ('OP', '('):
                self.advance()
                args = []
                if not (self.peek()[0] == 'OP' and self.peek()[1] == ')'):
                    args.append(self.parse_comparison())
                    while self.peek()[0] == 'OP' and self.peek()[1] == ',':
                        self.advance()
                        args.append(self.parse_comparison())
                self.expect_op(')')
                return ('call', val.upper(), args)
            # Bare name with no call -- not expected in our formula set.
            return ('str', val)
        if kind == 'OP' and val == '(':
            self.advance()
            node = self.parse_comparison()
            self.expect_op(')')
            return node
        raise ValueError(f"Unexpected token {(kind, val)!r}")


def parse(text):
    tokens = tokenize(text)
    return _Parser(tokens).parse()


# ---------------------------------------------------------------------------
# Value coercion -- mirrors Excel's own type-ordering and blank-cell rules
# ---------------------------------------------------------------------------

def _is_real_number(v):
    """True only for an actual number -- None/blank and non-numeric text are
    NOT real numbers. Used for COUNT/ISNUMBER/RANK.EQ, which count/inspect
    numeric cells specifically (unlike arithmetic, where blank means 0)."""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _coerce_num(v):
    """Coerce a value for use in arithmetic: blank/None -> 0 (Excel treats an
    empty cell as 0 in arithmetic), numeric text -> its float value, other
    text -> 0 (no #VALUE!-style errors in this bounded formula set)."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == '':
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _to_text(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _truthy(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return v.strip() != ''
    return bool(v)


def _is_numeric_like(v):
    """Excel's comparison type-ordering: blank cells count as numeric (0);
    an explicit blank/non-numeric STRING (e.g. the "" a formula's false-branch
    produces) does not -- Excel always treats text as greater than any number,
    so a blank-payback cell must never satisfy `<= 10`-style comparisons."""
    if v is None:
        return True
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    if isinstance(v, str):
        s = v.strip()
        if s == '':
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


def _compare(op, l, r):
    l_num, r_num = _is_numeric_like(l), _is_numeric_like(r)
    if l_num and r_num:
        lf, rf = _coerce_num(l), _coerce_num(r)
        if op == '=':
            return lf == rf
        if op == '<>':
            return lf != rf
        if op == '>':
            return lf > rf
        if op == '<':
            return lf < rf
        if op == '>=':
            return lf >= rf
        if op == '<=':
            return lf <= rf
    elif l_num != r_num:
        # Exactly one side is non-numeric text: Excel always ranks text
        # above any number, regardless of which side it's on.
        number_is_left = l_num
        if op in ('>', '>='):
            return not number_is_left
        if op in ('<', '<='):
            return number_is_left
        if op == '=':
            return False
        if op == '<>':
            return True
    # Both non-numeric: ordinary text comparison.
    ls, rs = _to_text(l), _to_text(r)
    if op == '=':
        return ls == rs
    if op == '<>':
        return ls != rs
    if op == '>':
        return ls > rs
    if op == '<':
        return ls < rs
    if op == '>=':
        return ls >= rs
    if op == '<=':
        return ls <= rs
    raise ValueError(f"Unsupported comparison operator {op!r}")


def _apply_binop(op, l, r):
    if isinstance(l, list) or isinstance(r, list):
        if isinstance(l, list) and isinstance(r, list):
            if len(l) != len(r):
                raise ValueError("Mismatched array lengths in binary operation")
            return [_apply_binop(op, a, b) for a, b in zip(l, r)]
        if isinstance(l, list):
            return [_apply_binop(op, a, r) for a in l]
        return [_apply_binop(op, l, b) for b in r]
    if op == '+':
        return _coerce_num(l) + _coerce_num(r)
    if op == '-':
        return _coerce_num(l) - _coerce_num(r)
    if op == '*':
        return _coerce_num(l) * _coerce_num(r)
    if op == '/':
        rv = _coerce_num(r)
        if rv == 0:
            raise ZeroDivisionError("division by zero")
        return _coerce_num(l) / rv
    if op == '^':
        return _coerce_num(l) ** _coerce_num(r)
    if op == '&':
        return _to_text(l) + _to_text(r)
    if op in ('=', '<>', '>', '<', '>=', '<='):
        return _compare(op, l, r)
    raise ValueError(f"Unsupported operator {op!r}")


def _flatten(args):
    for a in args:
        if isinstance(a, list):
            yield from a
        else:
            yield a


# ---------------------------------------------------------------------------
# Generic Excel function library -- standard, workbook-agnostic semantics
# ---------------------------------------------------------------------------

def _make_criteria_predicate(criteria):
    if isinstance(criteria, (int, float)) and not isinstance(criteria, bool):
        target = float(criteria)
        return lambda v: _is_real_number(v) and _coerce_num(v) == target
    s = _to_text(criteria)
    m = re.match(r'^(<>|>=|<=|>|<|=)?(.*)$', s)
    op, rest = m.group(1) or '=', m.group(2)
    try:
        threshold = float(rest)
        numeric = True
    except ValueError:
        threshold = rest
        numeric = False

    def predicate(v):
        if numeric:
            if not _is_real_number(v):
                return False
            vf = _coerce_num(v)
        else:
            vf = _to_text(v)
        if op == '=':
            return vf == threshold
        if op == '<>':
            return vf != threshold
        if op == '>':
            return vf > threshold
        if op == '<':
            return vf < threshold
        if op == '>=':
            return vf >= threshold
        if op == '<=':
            return vf <= threshold
        raise ValueError(f"Unsupported criteria operator {op!r}")

    return predicate


def fn_SUM(args):
    return sum(_coerce_num(v) for v in _flatten(args))


def fn_COUNT(args):
    return float(sum(1 for v in _flatten(args) if _is_real_number(v)))


def fn_COUNTIF(args):
    rng, criteria = args
    rng = rng if isinstance(rng, list) else [rng]
    pred = _make_criteria_predicate(criteria)
    return float(sum(1 for v in rng if pred(v)))


def fn_SUMIF(args):
    rng, criteria, sum_rng = args
    rng = rng if isinstance(rng, list) else [rng]
    sum_rng = sum_rng if isinstance(sum_rng, list) else [sum_rng]
    pred = _make_criteria_predicate(criteria)
    return sum(_coerce_num(sv) for v, sv in zip(rng, sum_rng) if pred(v))


def fn_SUMPRODUCT(args):
    arrays = [a if isinstance(a, list) else [a] for a in args]
    length = max(len(a) for a in arrays)
    total = 0.0
    for i in range(length):
        prod = 1.0
        for a in arrays:
            prod *= _coerce_num(a[i] if i < len(a) else a[0])
        total += prod
    return total


def fn_ISNUMBER(args):
    return _is_real_number(args[0])


def fn_RANK_EQ(args):
    value, ref = args[0], args[1]
    order = args[2] if len(args) > 2 else 0
    nums = [v for v in (ref if isinstance(ref, list) else [ref]) if _is_real_number(v)]
    val = _coerce_num(value)
    ascending = _coerce_num(order) != 0
    if ascending:
        return float(1 + sum(1 for v in nums if v < val))
    return float(1 + sum(1 for v in nums if v > val))


def fn_NPV(args):
    rate = _coerce_num(args[0])
    values = list(_flatten(args[1:]))
    return sum(_coerce_num(v) / (1 + rate) ** (i + 1) for i, v in enumerate(values))


def _irr_from_cashflows(cashflows, guess=0.1, max_iter=100, tol=1e-7):
    """Standard Newton-Raphson IRR root-find on an arbitrary cash-flow
    series (index 0 = period 0) -- a generic numerical method, not specific
    to any particular cash-flow model."""
    r = guess
    for _ in range(max_iter):
        npv = sum(cf / (1 + r) ** t for t, cf in enumerate(cashflows))
        d_npv = sum(-t * cf / (1 + r) ** (t + 1) for t, cf in enumerate(cashflows) if t > 0)
        if abs(d_npv) < 1e-12:
            break
        r_new = r - npv / d_npv
        if abs(r_new - r) < tol:
            return r_new
        if r_new <= -0.999999:
            r_new = (r - 0.999999) / 2  # keep (1+r) positive, avoid blow-up
        r = r_new
    # Bisection fallback if Newton-Raphson didn't converge cleanly.
    lo, hi = -0.999, 10.0

    def npv_at(rate):
        return sum(cf / (1 + rate) ** t for t, cf in enumerate(cashflows))

    f_lo, f_hi = npv_at(lo), npv_at(hi)
    if f_lo * f_hi > 0:
        raise ValueError("IRR did not converge")
    for _ in range(200):
        mid = (lo + hi) / 2
        f_mid = npv_at(mid)
        if abs(f_mid) < tol:
            return mid
        if f_lo * f_mid < 0:
            hi = mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2


def fn_IRR(args):
    values = args[0] if isinstance(args[0], list) else [args[0]]
    cashflows = [_coerce_num(v) for v in values]
    guess = _coerce_num(args[1]) if len(args) > 1 else 0.1
    return _irr_from_cashflows(cashflows, guess)


def fn_CONCATENATE(args):
    return ''.join(_to_text(v) for v in args)


_FUNCTIONS = {
    'SUM': fn_SUM,
    'COUNT': fn_COUNT,
    'COUNTIF': fn_COUNTIF,
    'SUMIF': fn_SUMIF,
    'SUMPRODUCT': fn_SUMPRODUCT,
    'ISNUMBER': fn_ISNUMBER,
    'RANK.EQ': fn_RANK_EQ,
    'NPV': fn_NPV,
    'IRR': fn_IRR,
    'CONCATENATE': fn_CONCATENATE,
}


# ---------------------------------------------------------------------------
# Evaluator
# ---------------------------------------------------------------------------

class FormulaError(Exception):
    """Raised for anything the evaluator cannot resolve -- caught by an
    enclosing IFERROR() the same way Excel would catch it, or propagated to
    the caller if there is none."""


class FormulaEngine:
    """
    Lazily evaluates formula cells on a single openpyxl worksheet, resolving
    dependencies on demand and memoizing every cell address it computes.
    Only cells actually reached while answering a get() call are ever
    touched -- unrelated formulas elsewhere on the sheet (e.g. a chart-data
    helper table) are never parsed, so they can never break evaluation of
    the cells this engine is actually asked about.
    """

    def __init__(self, ws, header_row, table_name):
        self.ws = ws
        self.header_row = header_row
        self.table_name = table_name
        self._cache = {}
        self._header_to_col = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                self._header_to_col[str(cell.value).strip()] = cell.column
        self._col_by_letters = {}

    def _col_index(self, letters):
        idx = self._col_by_letters.get(letters)
        if idx is None:
            idx = column_index_from_string(letters)
            self._col_by_letters[letters] = idx
        return idx

    def _col_for_header(self, name):
        try:
            return self._header_to_col[name]
        except KeyError:
            raise FormulaError(f"Unknown table column name {name!r}") from None

    def get(self, row, col):
        """Return the resolved value of ws cell (row, col): the literal
        value as-is if it isn't a formula, otherwise the evaluated result
        of its formula (memoized)."""
        key = (row, col)
        if key in self._cache:
            cached = self._cache[key]
            if cached is _CIRCULAR_SENTINEL:
                raise _CircularRef(
                    f"Circular reference detected at {self.ws.title}!"
                    f"{_col_letters(col)}{row}"
                )
            return cached
        raw = self.ws.cell(row=row, column=col).value
        if isinstance(raw, str) and raw.startswith('='):
            # Placeholder while evaluating, to raise a clear error on a
            # genuine circular reference rather than recursing forever.
            self._cache[key] = _CIRCULAR_SENTINEL
            try:
                node = parse(raw)
                value = self.eval_node(node, row)
            except _CircularRef:
                raise
            except Exception as exc:
                raise FormulaError(
                    f"Error evaluating {self.ws.title}!"
                    f"{_col_letters(col)}{row} ({raw!r}): {exc}"
                ) from exc
            self._cache[key] = value
            return value
        self._cache[key] = raw
        return raw

    def eval_node(self, node, current_row):
        kind = node[0]
        if kind == 'num' or kind == 'str':
            return node[1]
        if kind == 'cell':
            _, col_letters, row = node
            return self.get(row, self._col_index(col_letters))
        if kind == 'range':
            _, c1, r1, c2, r2 = node
            col1, col2 = self._col_index(c1), self._col_index(c2)
            rowlo, rowhi = min(r1, r2), max(r1, r2)
            collo, colhi = min(col1, col2), max(col1, col2)
            return [
                self.get(r, c)
                for r in range(rowlo, rowhi + 1)
                for c in range(collo, colhi + 1)
            ]
        if kind == 'tableref_row':
            _, table_name, col_name = node
            return self.get(current_row, self._col_for_header(col_name))
        if kind == 'tableref_col':
            raise FormulaError(
                "Whole-column table references are not needed by any "
                "formula this engine evaluates and are not implemented."
            )
        if kind == 'unary':
            _, op, expr = node
            v = self.eval_node(expr, current_row)
            return -_coerce_num(v) if op == '-' else v
        if kind == 'binop':
            _, op, lnode, rnode = node
            lv = self.eval_node(lnode, current_row)
            rv = self.eval_node(rnode, current_row)
            return _apply_binop(op, lv, rv)
        if kind == 'call':
            _, name, arg_nodes = node
            if name == 'IF':
                cond = self.eval_node(arg_nodes[0], current_row)
                if _truthy(cond):
                    return self.eval_node(arg_nodes[1], current_row)
                if len(arg_nodes) > 2:
                    return self.eval_node(arg_nodes[2], current_row)
                return False
            if name == 'IFERROR':
                try:
                    return self.eval_node(arg_nodes[0], current_row)
                except _CircularRef:
                    raise
                except Exception:
                    return self.eval_node(arg_nodes[1], current_row)
            argvals = [self.eval_node(a, current_row) for a in arg_nodes]
            fn = _FUNCTIONS.get(name)
            if fn is None:
                raise FormulaError(f"Unsupported function {name!r}")
            return fn(argvals)
        raise FormulaError(f"Unknown AST node {node!r}")


class _CircularRef(Exception):
    pass


_CIRCULAR_SENTINEL = object()


def _col_letters(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)
