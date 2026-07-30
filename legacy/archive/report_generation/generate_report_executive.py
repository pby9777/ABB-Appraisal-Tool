#!/usr/bin/env python3
"""
ABB Energy Appraisal – Executive Report Generator  (Phase 1: Scalar Tokens)
Template: ea_report_template_standard_top 10.docx
Regions:  Standard markets (India, LATAM, etc.)

Usage:
  python generate_report_executive.py  <saving_calc.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]

Requirements:
  pip install openpyxl
"""

import sys, os, re, math, io, zipfile
from datetime import datetime
import openpyxl

# ── CLI ───────────────────────────────────────────────────────────────────────
args = sys.argv[1:]
if len(args) < 3:
    print('Usage: python generate_report_executive.py  <excel.xlsx>  "CUSTOMER"  "Plant"  [Date]  [DataSource]')
    sys.exit(1)

XLSX_PATH   = args[0]
CUSTOMER    = args[1].upper()
PLANT       = args[2]
RPT_DATE    = args[3] if len(args) > 3 else datetime.now().strftime("%m.%d.%Y")
DATA_SOURCE = args[4] if len(args) > 4 else "Customer Input"

_script_dir   = os.path.dirname(os.path.abspath(__file__))
_template_dir = os.path.join(_script_dir, "report_templates")
TEMPLATE_NAME = "Spain_Global Switch_EA_Report_V1_top-10.docx"
TEMPLATE_PATH = os.path.join(_template_dir, TEMPLATE_NAME)

for path, label in [(XLSX_PATH, "Excel file"), (TEMPLATE_PATH, "Template")]:
    if not os.path.exists(path):
        print(f"ERROR: {label} not found: {path}")
        sys.exit(1)

def date_to_iso(d):
    """Convert any date string to YYYY-MM-DD."""
    for fmt in ('%m.%d.%Y', '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(d, fmt).strftime('%Y-%m-%d')
        except:
            pass
    return d

RPT_DATE_ISO = date_to_iso(RPT_DATE)

# ── Read Excel ────────────────────────────────────────────────────────────────
# The Saving Calculations workbook is the single source of truth for every
# financial/engineering calculation (NPV, payback, IRR, KPI totals, Top-10
# ranking). openpyxl never evaluates formulas, so a workbook that was filled
# and saved purely by openpyxl carries stale/blank cached results for those
# cells. recalculate_workbook() fixes that by evaluating the workbook's own
# formulas in-process (no external application involved), but it's only
# actually needed when the cache is stale — if the workbook was already
# opened/saved by a real spreadsheet app (or this script already
# recalculated it once), the cached values are trustworthy as-is and
# recalculating again would just be wasted time. See
# report_table_utils.recalculate_workbook()/cells_are_cached()/
# verify_recalculated().
from report_table_utils import recalculate_workbook, cells_are_cached, verify_recalculated

print(f"Reading Excel: {os.path.basename(XLSX_PATH)}")

def _is_sc_sheet(name):
    n = name.strip().lower().replace(' ', '_')
    prefixes = ('saving_calculations', 'savings_calculations',
                'saving_calculatios',  'savings_calculatios')
    return any(n.startswith(p) for p in prefixes)

def _find_sc_sheet(workbook):
    _sc_sheet = next((s for s in workbook.sheetnames if _is_sc_sheet(s)), None)
    if _sc_sheet is None:
        print(f"ERROR: No Saving_Calculations sheet found. Sheets: {workbook.sheetnames}")
        sys.exit(1)
    _sc_candidates = [s for s in workbook.sheetnames if _is_sc_sheet(s)]
    if len(_sc_candidates) > 1:
        import re as _re
        _vn = [s for s in _sc_candidates if _re.search(r'_v\d+$', s.strip().lower())]
        if _vn:
            _sc_sheet = _vn[0]
    if _sc_sheet not in ("Saving_Calculations", "Savings_Calculations"):
        print(f"  Note: using sheet '{_sc_sheet}'")
    return workbook[_sc_sheet]

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = _find_sc_sheet(wb)

_XL_ERRORS = frozenset({
    '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!', '#ERROR!',
    '-', '–', '—', 'N/A', 'n/a', 'NA', 'na', 'n.a.', 'N.A.',
})

def cv(row, col):
    """Read a cell value; return None for Excel error strings."""
    v = ws.cell(row=row, column=col).value
    return None if isinstance(v, str) and v.strip() in _XL_ERRORS else v

# ── Detect column positions from header row 36 ────────────────────────────────
HEADER_ROW = 36
col_map = {}
for c in range(1, 70):
    h = cv(HEADER_ROW, c)
    if h:
        col_map[str(h).strip().lower()] = c

def find_col(names, default):
    for n in names:
        if n.lower() in col_map:
            return col_map[n.lower()]
    return default

COL_NUM      = find_col(['#', '# no.', 'no.'], 2)
COL_E_CONS   = find_col(['annual energy. cons (kwh)', 'annual energy cons (kwh)'], 5)
COL_E_COST   = find_col(['annual energy cost'], 6)
COL_CO2_CONS = find_col(['annual co2 cons.'], 7)
COL_SAV_KWH  = find_col(['annual energy savings, kwh'], 8)
COL_SAV_COST = find_col(['annual energy cost savings'], 9)
COL_SAV_PCT  = find_col(['annual energy savings (%)'], 10)
COL_INVEST   = find_col(['investment'], 11)
_hdr13 = str(ws.cell(HEADER_ROW, 13).value or '').lower()
_hdr12 = str(ws.cell(HEADER_ROW, 12).value or '').lower()
_IS_V2_EXCEL = 'cee' in _hdr12 or 'cee' in _hdr13
if _IS_V2_EXCEL:
    COL_INVEST_NET = find_col(['investment - cee', 'investment-cee'], 13)
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 14)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 15)
    print(f"  Detected v2-style Excel (CEE columns) — using net investment col {COL_INVEST_NET}")
else:
    COL_INVEST_NET = COL_INVEST
    COL_PAYBACK    = find_col(['payback time, if npv positive', 'payback time'], 12)
    COL_CO2_SAV    = find_col(['annual co2 savings (kg)'], 13)
COL_NPV      = find_col(['npv'], 17)
COL_IE       = find_col(['ie eff class', 'ie'], 41)
COL_LOAD     = find_col(['driven load', 'application'], 42)
COL_FLOW     = find_col(['flow control', 'flow control method'], 43)
COL_CONN     = find_col(['dol vsd', 'connection'], 44)
COL_OUTPUT   = find_col(['output (kw)', 'rated power, kw'], 45)
COL_SHAFT    = find_col(['shaft height (frame)', 'shaft height'], 46)
COL_RUNHRS   = find_col(['running hours', 'running time (hours)'], 47)
COL_AVG      = find_col(['average flow % / average frequency (hz)',
                          'average flow', 'average frequency', 'avg. frequency'], 48)
COL_ESS      = find_col(['recommended ess motor'], 50)
COL_ESSC     = find_col(['ess connection'], 51)

# ── Recalculate only if the workbook's cached formula values are stale ───────
# Header labels (just read via col_map above) are literal text and identical
# whether or not the workbook has been recalculated, so column detection is
# safe to run before this check. Everything below it depends on real cached
# numbers, so it must run after.
_RECALC_CHECKS = [
    (13, 3,        "Total assets count"),
    (14, 3,        "NPV+ asset count"),
    (15, 3,        "Annual savings, NPV+ group"),
    (19, 3,        "NPV, NPV+ group"),
    (20, 3,        "Investment, NPV+ group"),
    (23, 3,        "IRR, NPV+ group"),
    (25, 4,        "BEV count"),
    (34, 5,        "NPV+ group totals row (row 34)"),
    (33, 5,        "Top-10 (lowest payback) totals row (row 33)"),
    (19, 4,        "Top-10 NPV"),
    (20, 4,        "Top-10 Investment"),
    (21, 4,        "Top-10 Payback"),
    (23, 4,        "Top-10 IRR"),
    (37, COL_NPV,  "First asset's NPV (per-asset formula)"),
]

if cells_are_cached(ws, _RECALC_CHECKS):
    print("  Workbook formulas already cached — skipping recalculation.")
else:
    print("  Workbook formulas not cached — recalculating...")
    RECALCED_PATH = recalculate_workbook(XLSX_PATH)
    wb = openpyxl.load_workbook(RECALCED_PATH, data_only=True)
    ws = _find_sc_sheet(wb)
    # Fail loudly if recalculation still didn't populate these cells — never
    # silently fall back to computing them ourselves.
    verify_recalculated(ws, _RECALC_CHECKS)

# ── KPI values ────────────────────────────────────────────────────────────────
CURRENCY       = str(cv(3, 3) or "INR").strip()
ELEC_PRICE     = cv(5, 4) or 8
CO2_INTENSITY  = cv(6, 4) or 0.54
TAX_RATE       = cv(8, 4) or 0.349
DISCOUNT_RATE  = float(cv(7, 4) or 0.065)
_NPV_YEARS     = 20
TOTAL_ASSETS   = int(cv(13, 3) or 0)
NPV_POS_CNT    = int(cv(14, 3) or 0)
ANNUAL_SAVINGS = cv(15, 3)
CONSUMP_BEFORE = cv(16, 3)
SAVINGS_KWH    = cv(17, 3)
NPV_VALUE      = cv(19, 3)
INVEST_COST    = cv(20, 3)
PAYBACK_TIME   = cv(21, 3)
CO2_SAVINGS    = cv(22, 3)
IRR_VALUE      = cv(23, 3)
BEV_COUNT      = cv(25, 4)
TOP10_PAYBACK  = cv(21, 4)
TOP10_NPV      = cv(19, 4)
TOP10_INVEST   = cv(20, 4)
TOP10_IRR      = cv(23, 4)

SENSITIVITY = []
for r in range(19, 24):
    delta   = cv(r, 20)
    payback = cv(r, 22)
    try:
        SENSITIVITY.append((float(delta), float(payback)))
    except (TypeError, ValueError):
        pass

# ── Load assets ───────────────────────────────────────────────────────────────
# TOTAL_ASSETS (Excel's own =COUNT(B37:B1100) cell) is metadata only, never a
# scan bound: openpyxl never evaluates formulas, so a workbook that was filled
# and saved by fill_saving_calculations.py without ever being reopened in a
# real spreadsheet app carries a stale/blank cached value for it (verified:
# after such a fill+save round-trip the cached result reads back as None). The
# actual end of data is detected the same way fill_saving_calculations.py
# itself bounds writes to this sheet (SCAN_LIMIT there) — by scanning the
# fixed structural row window and stopping at the first row whose '#' cell is
# genuinely blank/invalid, which the loop below already does. This list also
# feeds the All-Assets appendix tables further down, so under-scanning here
# would silently truncate that appendix too, not just the Complete report.
ASSET_START_ROW  = 37
ASSET_ROW_CEILING = 1100  # matches the sheet's own B37:B1100 formula range
assets = []
for r in range(ASSET_START_ROW, ASSET_START_ROW + ASSET_ROW_CEILING):
    num = cv(r, COL_NUM)
    if num is None: break
    try: num = int(float(num))
    except: break
    if not (1 <= num <= 9999): break

    ess_motor = str(cv(r, COL_ESS) or "")

    shaft_raw = cv(r, COL_SHAFT)
    try:
        shaft_val = str(int(float(shaft_raw))) if shaft_raw is not None else ""
    except:
        shaft_val = str(shaft_raw) if shaft_raw else ""

    avg_raw = cv(r, COL_AVG)
    if avg_raw is not None:
        try:
            v = float(avg_raw)
            avg_val = f"{round(v*100)}" if 0 < v <= 1 else str(avg_raw).replace(' ', '')
        except:
            avg_val = str(avg_raw).replace(' ', '')
    else:
        fm = re.search(r'([\d.]+)\s*Hz', ess_motor)
        avg_val = fm.group(1) if fm else ""

    conn_raw = cv(r, COL_CONN)
    conn_val = str(conn_raw) if conn_raw else "DOL"

    assets.append({
        "num":       num,
        "tag":       str(cv(r, 3) or ""),
        "load":      str(cv(r, COL_LOAD) or cv(r, 4) or ""),
        "e_cons":    cv(r, COL_E_CONS)   or 0,
        "e_cost":    cv(r, COL_E_COST)   or 0,
        "co2_cons":  cv(r, COL_CO2_CONS) or 0,
        "e_sav_kwh": cv(r, COL_SAV_KWH)  or 0,
        "e_sav_cost":cv(r, COL_SAV_COST) or 0,
        "e_sav_pct": cv(r, COL_SAV_PCT)  or 0,
        "invest":    cv(r, COL_INVEST)   or 0,
        "payback":   cv(r, COL_PAYBACK) if not isinstance(cv(r, COL_PAYBACK), str) or cv(r, COL_PAYBACK).strip() not in ('', '-') else None,
        "co2_sav":   cv(r, COL_CO2_SAV)  or 0,
        "npv":       cv(r, COL_NPV),
        "ie":        str(cv(r, COL_IE)   or ""),
        "flow_ctrl": str(cv(r, COL_FLOW) or "Information not available"),
        "connection":conn_val,
        "output_kw": str(cv(r, COL_OUTPUT) or ""),
        "shaft_h":   shaft_val,
        "run_hrs":   str(cv(r, COL_RUNHRS) or ""),
        "avg_val":   avg_val,
        "ess_motor": ess_motor,
        "ess_conn":  str(cv(r, COL_ESSC) or ""),
    })

if not assets:
    print("ERROR: No assets found. Check TOTAL_ASSETS in Excel row 13.")
    sys.exit(1)

NA = len(assets)
if NA != TOTAL_ASSETS:
    print(f"  WARNING: asset rows read ({NA}) != workbook's Total assets count "
          f"({TOTAL_ASSETS}) — check recalculation/scan range.")


def _payback_sort_key(a):
    """Ascending Payback Period; assets with no defined payback (NPV<=0) sort
    last. Ties are broken by original workbook row order — identical to the
    workbook's own 'Payback Rank (Top 10)' column (RANK.EQ + COUNTIF tie-
    break), so the two are always the same group for the same assets."""
    p = a["payback"]
    try:
        return (0, float(p)) if p is not None else (1, 0.0)
    except (TypeError, ValueError):
        return (1, 0.0)


# Single sort, reused for every section of the report. The Executive report's
# main body is simply the first 10 of this list; the appendix is the whole
# list -- both in the exact same order the workbook's own rank-based row 33
# aggregates (and therefore the Top-10 KPIs read above) were computed over.
assets_sorted = sorted(assets, key=_payback_sort_key)
top10 = assets_sorted[:10]
NT = len(top10)

print(f"  {NA} assets, {NPV_POS_CNT} NPV+, {NT} in Top-{NT} (lowest payback, from workbook)")

# ── Formatting helpers ────────────────────────────────────────────────────────
SYM = CURRENCY

def fmt(n, dp=0):
    if n is None or (isinstance(n, float) and math.isnan(n)): return "–"
    try:
        n = float(n)
        s = f"{abs(n):,.{dp}f}"
        return s if n >= 0 else f"-{s}"
    except: return str(n)

def fmtpct(n):
    if n is None: return "–"
    v = round(float(n) * 100)
    return f"{v}%" if v >= 0 else f"-{abs(v)}%"

def fmtyrs(n):
    if n is None: return ""
    try: return f"{float(n):.1f}"
    except: return ""

def fmtirr(n):
    if n is None: return "–"
    try: return f"{round(float(n)*100)}%"
    except: return str(n)

def xe(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

# ── Phase 2: shared table-rendering engine ───────────────────────────────────
from report_table_utils import (
    render_table_section, trim_stale_appendix,
    replace_literal_paragraphs, replace_literal_runs,
    replace_within_row, replace_within_textboxes,
)


# ── Top-10 (lowest payback) group totals ─────────────────────────────────────
# Read directly from the workbook's own row 33 aggregate — now rank-based
# (=SUMPRODUCT(($BF$37:$BF$1100<=10)*(...))), i.e. the true lowest-10-payback
# group, not recomputed from the asset list. TOP10_NPV/TOP10_INVEST/
# TOP10_PAYBACK/TOP10_IRR were already read from column D (rows 19-23) above;
# these are the same row-33 group's remaining per-column totals for the
# Top-10 table's own totals row.
t10_e_cons  = cv(33, 5)  or 0   # E33
t10_e_cost  = cv(33, 6)  or 0   # F33
t10_co2_c   = cv(33, 7)  or 0   # G33
t10_sav_kwh = cv(33, 8)  or 0   # H33
t10_sav_c   = cv(33, 9)  or 0   # I33
t10_pct     = cv(33, 10) or 0   # J33 = IFERROR(H33/E33,"-")
t10_invest  = cv(33, 11) or 0   # K33
t10_co2_sav = cv(33, 13) or 0   # M33
t10_npv     = cv(33, 17) or 0   # Q33

# ── NPV+ group totals (for total row of NPV+ table) ──────────────────────────
# Read directly from the workbook's own row 34 aggregate
# (=SUMIF($Q$37:$Q$1100,">0", ...)) — not recomputed from the asset list.
npv_e_cons  = cv(34, 5)  or 0   # E34
npv_e_cost  = cv(34, 6)  or 0   # F34
npv_co2_c   = cv(34, 7)  or 0   # G34
npv_sav_kwh = cv(34, 8)  or 0   # H34
npv_sav_c   = cv(34, 9)  or 0   # I34
npv_co2_sav = cv(34, 13) or 0   # M34
npv_sav_pct = cv(34, 10) or 0   # J34 = IFERROR(H34/E34,"-")

# ── Phase 1: Scalar Token Replacement Engine ──────────────────────────────────

REQUIRED_SCALAR_TOKENS = [
    '{{CUSTOMER}}',
    '{{PLANT}}',
    '{{DATA_SOURCE}}',
    '{{REPORT_DATE}}',
    '{{REPORT_DATE_ISO}}',
    '{{NPV_COUNT}}',
    '{{TOTAL_ASSETS}}',
    '{{ELEC_PRICE}}',
    '{{CURRENCY}}',
    '{{CO2_INTENSITY}}',
    '{{ANNUAL_SAVINGS}}',
    '{{NPV_INVEST}}',
    '{{NPV_VALUE}}',
    '{{PAYBACK}}',
    '{{IRR_DISPLAY}}',
    '{{CO2_SAVINGS}}',
    '{{BEV_COUNT}}',
    '{{TOP10_ANNUAL_SAVINGS}}',
    '{{TOP10_INVEST}}',
    '{{TOP10_NPV}}',
    '{{TOP10_PAYBACK}}',
    '{{TOTAL_TOP10_ENERGY_CONS}}',
    '{{TOTAL_TOP10_ENERGY_COST}}',
    '{{TOTAL_TOP10_CO2_CONS}}',
    '{{TOTAL_TOP10_SAVINGS_KWH}}',
    '{{TOTAL_TOP10_SAVING_PCT}}',
    '{{TOTAL_TOP10_CO2_AVOIDED}}',
    '{{TOTAL_NPV_ENERGY_CONS}}',
    '{{TOTAL_NPV_ENERGY_COST}}',
    '{{TOTAL_NPV_CO2_CONS}}',
    '{{TOTAL_NPV_SAVINGS_KWH}}',
    '{{TOTAL_NPV_SAVING_PCT}}',
    '{{TOTAL_NPV_CO2_AVOIDED}}',
]

# Structural Phase 2 markers — never replaced by this engine, never validated as missing
_PHASE2_TOKENS = {'{{ROW_TEMPLATE}}'}


def extract_scalar_values():
    """Return dict mapping each scalar token to its display string."""
    return {
        # Metadata
        '{{CUSTOMER}}':          CUSTOMER,
        '{{PLANT}}':             PLANT,
        '{{DATA_SOURCE}}':       DATA_SOURCE,
        '{{REPORT_DATE}}':       RPT_DATE,
        '{{REPORT_DATE_ISO}}':   RPT_DATE_ISO,
        # Cover KPIs
        '{{NPV_COUNT}}':         str(NPV_POS_CNT),
        '{{TOTAL_ASSETS}}':      str(TOTAL_ASSETS),
        '{{ELEC_PRICE}}':        f"{float(ELEC_PRICE):.2f}",
        '{{CURRENCY}}':          CURRENCY,
        '{{CO2_INTENSITY}}':     str(CO2_INTENSITY),
        # NPV+ summary KPIs (summary page)
        '{{ANNUAL_SAVINGS}}':    fmt(ANNUAL_SAVINGS),
        '{{NPV_INVEST}}':        fmt(INVEST_COST),
        '{{NPV_VALUE}}':         fmt(NPV_VALUE),
        '{{PAYBACK}}':           fmtyrs(PAYBACK_TIME),
        '{{IRR_DISPLAY}}':       fmtirr(TOP10_IRR),     # includes %, e.g. "42%"
        '{{CO2_SAVINGS}}':       str(round(float(CO2_SAVINGS or 0) / 1000)),
        '{{BEV_COUNT}}':         str(round(float(BEV_COUNT or 0))),
        # Top-10 summary KPIs
        '{{TOP10_ANNUAL_SAVINGS}}': fmt(t10_sav_c),
        '{{TOP10_INVEST}}':         fmt(t10_invest),
        '{{TOP10_NPV}}':            fmt(t10_npv),
        '{{TOP10_PAYBACK}}':        fmtyrs(TOP10_PAYBACK),
        # Top-10 total row
        '{{TOTAL_TOP10_ENERGY_CONS}}':  fmt(t10_e_cons),
        '{{TOTAL_TOP10_ENERGY_COST}}':  fmt(t10_e_cost),
        '{{TOTAL_TOP10_CO2_CONS}}':     fmt(t10_co2_c),
        '{{TOTAL_TOP10_SAVINGS_KWH}}':  fmt(t10_sav_kwh),
        '{{TOTAL_TOP10_SAVING_PCT}}':   fmtpct(t10_pct),
        '{{TOTAL_TOP10_CO2_AVOIDED}}':  fmt(t10_co2_sav),
        # NPV+ total row
        '{{TOTAL_NPV_ENERGY_CONS}}':    fmt(npv_e_cons),
        '{{TOTAL_NPV_ENERGY_COST}}':    fmt(npv_e_cost),
        '{{TOTAL_NPV_CO2_CONS}}':       fmt(npv_co2_c),
        '{{TOTAL_NPV_SAVINGS_KWH}}':    fmt(npv_sav_kwh),
        '{{TOTAL_NPV_SAVING_PCT}}':     fmtpct(npv_sav_pct),
        '{{TOTAL_NPV_CO2_AVOIDED}}':    fmt(npv_co2_sav),
    }


def replace_scalars(xml, token_map):
    """Replace all scalar tokens in xml with their display values (XML-escaped)."""
    for token, value in token_map.items():
        xml = xml.replace(token, xe(value))
    return xml


def validate_required_tokens(xml, footer_xmls, required_tokens):
    """Pre-flight: verify all required tokens exist in the template.
    Exits with error if any are missing (template was not properly tokenized)."""
    full = xml + ''.join(footer_xmls)
    missing = [t for t in required_tokens if t not in full]
    if missing:
        print("ERROR: Template is missing required tokens (run prepare_templates.py first):")
        for t in missing:
            print(f"  {t}")
        sys.exit(1)


def validate_no_unreplaced_tokens(xml, footer_xmls, skip_tokens=None):
    """Post-flight: verify no {{...}} tokens remain after replacement.
    skip_tokens: set of tokens intentionally left for later phases."""
    skip_tokens = skip_tokens or set()
    full = xml + ''.join(footer_xmls)
    remaining = re.findall(r'\{\{[A-Z0-9_]+\}\}', full)
    unreplaced = [t for t in remaining if t not in skip_tokens]
    if unreplaced:
        print("ERROR: Unreplaced tokens remain after scalar replacement:")
        for t in sorted(set(unreplaced)):
            print(f"  {t}")
        sys.exit(1)


# ── Main ──────────────────────────────────────────────────────────────────────
print(f"Loading template: {TEMPLATE_NAME}")
with open(TEMPLATE_PATH, 'rb') as f:
    template_bytes = f.read()

zin  = zipfile.ZipFile(io.BytesIO(template_bytes))
names = zin.namelist()

doc_xml = zin.read('word/document.xml').decode('utf-8')

footer_names = sorted(n for n in names if re.match(r'word/footer\d+\.xml', n))
footer_xmls  = [zin.read(n).decode('utf-8') for n in footer_names]

# NOTE: Spain_Global Switch_EA_Report_V1_top-10.docx is a real, previously
# filled-in sample report (not a {{TOKEN}}-tokenized template) - it has none
# of the REQUIRED_SCALAR_TOKENS, so the token-based Phase 1 above (extract_
# scalar_values/replace_scalars/validate_required_tokens/validate_no_
# unreplaced_tokens) does not apply to it and is intentionally not called
# here. Those functions are left unchanged for any tokenized template that
# may use them. Scalar substitution for this template is done further below,
# by literal value, after Phase 2 - see that block for why the ordering is
# reversed.

# ── Charts: update cached data points so Word renders real figures ──────────
# Chart XML caches its data as <c:numRef><c:f>...</c:f><c:numCache>...points...
# </c:numCache></c:numRef>. Word displays the cache, not a live link, so we
# overwrite the cached <c:v> values in place — this is robust to whatever
# sheet/cell the template's embedded workbook happens to reference. This
# block was previously missing entirely from this script (unlike
# generate_report_standard.py), which is why none of the Executive report's
# summary-page charts ever reflected the actual customer data.
#
# Business rule: the Payback Sensitivity chart and the energy-consumption
# "Now vs Upgraded Fleet" chart must always show the same All-Assets figures
# as the Complete report, never a Top-10-only subset. CONSUMP_BEFORE/
# SAVINGS_KWH/SENSITIVITY above are already read from the same NPV+/
# all-assets workbook cells generate_report_standard.py uses (rows 16/17 and
# 19-23, column C) — reusing them here keeps the two reports' charts
# identical with no extra Top-10-specific logic required. The NPV-positive
# pie chart is fleet-wide metadata (same NPV_POS_CNT/TOTAL_ASSETS shown on
# both reports' cover pages via the "21/45 motors NPV positive" fraction),
# not a Top-10-scoped figure, so it uses the same values here too.

def _set_single_point_values(ctxt, values):
    """Overwrite the single <c:pt idx="0"> value of the first len(values)
    <c:numRef> blocks found, in document order."""
    refs = list(re.finditer(r'<c:numRef>.*?</c:numRef>', ctxt, re.DOTALL))
    out, prev = [], 0
    for ref, val in zip(refs, values):
        out.append(ctxt[prev:ref.start()])
        seg = re.sub(r'(<c:pt idx="0"><c:v>)[^<]*(</c:v>)',
                     lambda m: f'{m.group(1)}{val:.10f}{m.group(2)}',
                     ref.group(), count=1)
        out.append(seg)
        prev = ref.end()
    out.append(ctxt[prev:])
    return ''.join(out)


def _set_last_multi_point_values(ctxt, values):
    """Overwrite every <c:pt> value inside the LAST <c:numRef> block found —
    that block holds the plotted series regardless of how many numRefs
    (categories + series, or just series) the template's chart declares."""
    refs = list(re.finditer(r'<c:numRef>.*?</c:numRef>', ctxt, re.DOTALL))
    if not refs:
        return ctxt
    ref = refs[-1]
    seg = ref.group()
    pts = list(re.finditer(r'<c:pt idx="(\d+)"><c:v>[^<]*</c:v></c:pt>', seg))
    new_seg = seg
    for pt, val in list(zip(pts, values))[::-1]:
        idx = pt.group(1)
        new_seg = new_seg[:pt.start()] + f'<c:pt idx="{idx}"><c:v>{val}</c:v></c:pt>' + new_seg[pt.end():]
    return ctxt[:ref.start()] + new_seg + ctxt[ref.end():]


# Binary/zip parts rewritten below, applied at write time alongside
# doc_xml/footer_xmls (see the write loop at the bottom of this script).
binary_overrides = {}

# Chart 1: Energy Consumption bar (before / after, All-Assets — "Now vs
# Upgraded Fleet") — 2 single-point series
cons_before = float(CONSUMP_BEFORE or 0)
cons_after  = cons_before - float(SAVINGS_KWH or 0)
if 'word/charts/chart1.xml' in names:
    ctxt = zin.read('word/charts/chart1.xml').decode('utf-8')
    ctxt = _set_single_point_values(ctxt, [cons_before, cons_after])
    binary_overrides['word/charts/chart1.xml'] = ctxt.encode('utf-8')
    print(f"  Chart 1 (Energy Consumption): {fmt(cons_before)} kWh -> {fmt(cons_after)} kWh")

# Chart 2: NPV positive pie chart (fleet-wide, All-Assets) — update embedded
# Excel workbook + cache
if 'word/embeddings/Microsoft_Excel_Worksheet.xlsx' in names:
    emb_bytes = zin.read('word/embeddings/Microsoft_Excel_Worksheet.xlsx')
    ewb = openpyxl.load_workbook(io.BytesIO(emb_bytes))
    ews = ewb.active
    ews['B2'] = NPV_POS_CNT
    ews['B3'] = TOTAL_ASSETS - NPV_POS_CNT
    ews['A2'] = 'NPV Positive'
    ews['A3'] = 'Remaining Assets'
    buf2 = io.BytesIO()
    ewb.save(buf2)
    ewb.close()
    binary_overrides['word/embeddings/Microsoft_Excel_Worksheet.xlsx'] = buf2.getvalue()
    if 'word/charts/chart2.xml' in names:
        c2txt = zin.read('word/charts/chart2.xml').decode('utf-8')
        c2txt = _set_last_multi_point_values(c2txt, [NPV_POS_CNT, TOTAL_ASSETS - NPV_POS_CNT])
        c2txt = c2txt.replace('<c:v>2nd Qtr</c:v>', '<c:v>Remaining Assets</c:v>')
        binary_overrides['word/charts/chart2.xml'] = c2txt.encode('utf-8')
    print(f"  Chart 2 (NPV pie): {NPV_POS_CNT} NPV+ / {TOTAL_ASSETS - NPV_POS_CNT} remaining")

# Chart 3: Payback sensitivity (All-Assets) — overwrite the plotted payback series
if 'word/charts/chart3.xml' in names and SENSITIVITY:
    ctxt = zin.read('word/charts/chart3.xml').decode('utf-8')
    ctxt = _set_last_multi_point_values(ctxt, [s[1] for s in SENSITIVITY])
    binary_overrides['word/charts/chart3.xml'] = ctxt.encode('utf-8')
    print(f"  Chart 3 (Payback sensitivity): {[round(s[1], 2) for s in SENSITIVITY]}")

# ── Phase 2: Dynamic Table Rendering ─────────────────────────────────────────
# Field-map keys are normalize_header(template header text) — see report_table_utils.
# '#' / 'application' / etc. below match the templates' actual header row, not
# arbitrary invented labels.

_ENERGY_FIELD_MAP = {
    '#':                    lambda a: str(a['num']),
    'customer equipment id':lambda a: a['tag'],
    'application':          lambda a: a['load'],
    'energy. cons (kwh)':   lambda a: fmt(a['e_cons']),
    'energy cost':          lambda a: fmt(a['e_cost']),
    'co2 cons. (kg)':       lambda a: fmt(a['co2_cons']),
    'energy savings (kwh)': lambda a: fmt(a['e_sav_kwh']),
    'energy cost savings':  lambda a: fmt(a['e_sav_cost']),
    'energy saving (%)':    lambda a: fmtpct(a['e_sav_pct']),
    'investment':           lambda a: fmt(a['invest']),
    'payback time (years)': lambda a: fmtyrs(a['payback']),
    'avoided co2 (kg)':     lambda a: fmt(a['co2_sav']),
}

# The Application Details table has no "Customer Equipment ID" column — its
# 'application' column is the equipment tag; 'driven load' carries the load type.
_DETAILS_FIELD_MAP = {
    '#':                     lambda a: str(a['num']),
    'application':           lambda a: a['tag'],
    'ie':                    lambda a: a['ie'],
    'driven load':           lambda a: a['load'],
    'flow control method':   lambda a: a['flow_ctrl'],
    'connection':            lambda a: a['connection'],
    'output (kw)':           lambda a: str(a['output_kw']),
    'running time (hours)':  lambda a: str(a['run_hrs']),
    # Loading% only applies to DOL connections, frequency only to VSD — the
    # single avg_val field covers both depending on which was measured.
    'avg. loading (%)':      lambda a: a['avg_val'] if a['connection'] == 'DOL' else '-',
    'avg. frequency (hz)':   lambda a: a['avg_val'] if a['connection'] != 'DOL' else '-',
    'recommended ess motor': lambda a: a['ess_motor'],
    'ess connection':        lambda a: a['ess_conn'],
}

_ENERGY_REQUIRED = [
    '#', 'customer equipment id', 'application',
    'energy. cons (kwh)', 'energy cost',
    'energy savings (kwh)', 'investment',
]

_DETAILS_REQUIRED = [
    '#', 'application', 'ie', 'driven load',
]

# Every template has exactly one Energy Savings table and one Application
# Details table. The Executive report populates each with the top-10 ranked
# assets; the Complete Asset report (generate_report_standard.py) populates
# the same two sections with every asset. Same anchors, same renderer —
# only the asset list passed in differs.
print("Rendering dynamic tables (Phase 2)...")

doc_xml = render_table_section(
    doc_xml, 'Energy Savings (Top 10)',
    anchor='Energy savings with ABB premium efficiency solutions',
    asset_list=top10, field_map=_ENERGY_FIELD_MAP, required=_ENERGY_REQUIRED,
    n_totals=2,
)
doc_xml = render_table_section(
    doc_xml, 'Application Details (Top 10)',
    anchor='Application Details – Top 10',
    asset_list=top10, field_map=_DETAILS_FIELD_MAP, required=_DETAILS_REQUIRED,
    n_totals=0,
)

# Unlike the older tokenized top-10 template this engine was originally built
# for, Spain_Global Switch_EA_Report_V1_top-10.docx also carries a genuine
# All-Assets appendix after the Top-10 section (Application Details – All
# Assets / Details of Recommendation – All Assets - the heading names are
# swapped relative to what they actually contain: the "Application Details"
# heading precedes the Energy-Savings-shaped table, and "Details of
# Recommendation" precedes the motor-spec-shaped table; confirmed against
# each table's own header row, not assumed from the heading text). This is
# real, final content, not a stale leftover to be trimmed - so it is
# populated with every asset here instead of being dropped.
doc_xml = render_table_section(
    doc_xml, 'Energy Savings (All Assets, appendix)',
    anchor='Application Details – All Assets',
    asset_list=assets_sorted, field_map=_ENERGY_FIELD_MAP, required=_ENERGY_REQUIRED,
    n_totals=1,
)
doc_xml = render_table_section(
    doc_xml, 'Application Details (All Assets, appendix)',
    anchor='Details of Recommendation – All Assets',
    asset_list=assets_sorted, field_map=_DETAILS_FIELD_MAP, required=_DETAILS_REQUIRED,
    n_totals=0,
)

print("  Phase 2 complete.")

# ── Phase 1 (literal): scalar substitution for the un-tokenized V1 template ──
# Run this AFTER Phase 2 for the same reason as generate_report_standard.py:
# every per-asset row (Top-10 and the All-Assets appendix) has already been
# overwritten with this customer's real numbers by the time this runs, so a
# bare value here can only match the cover table / KPI tiles / totals rows /
# text boxes it's meant for - not some other asset's own cell.
print("Substituting scalar values (literal, no {{TOKEN}} markers in this template)...")

_cover_replacements = [
    ('Global Switch',      CUSTOMER),
    ('Spain',               PLANT),
    ('Customer Input',      DATA_SOURCE),
    ('23.07.2026',          RPT_DATE),
    ('21 / 45',             f"{NPV_POS_CNT} / {TOTAL_ASSETS}"),
    ('0.14 EUR/kWh',        f"{float(ELEC_PRICE):.2f} {CURRENCY}/kWh"),
    ('0.15 kg CO2/kWh',     f"{CO2_INTENSITY} kg CO2/kWh"),
]

_kpi_replacements = [
    ('EUR 26,529',          f"{CURRENCY} {fmt(t10_sav_c)}"),        # Top-10 annual savings
    ('EUR 31,229',          f"{CURRENCY} {fmt(ANNUAL_SAVINGS)}"),   # overall annual savings
    ('EUR 49,928',          f"{CURRENCY} {fmt(t10_invest)}"),       # Top-10 investment
    ('EUR 89,528',          f"{CURRENCY} {fmt(INVEST_COST)}"),      # overall investment
    ('EUR 178,280',         f"{CURRENCY} {fmt(t10_npv)}"),          # Top-10 NPV
    ('EUR 184,638',         f"{CURRENCY} {fmt(NPV_VALUE)}"),        # overall NPV
    ('33 tCO2',             f"{round(float(CO2_SAVINGS or 0) / 1000)} tCO2"),
    ('64 Vehicles',         f"{round(float(BEV_COUNT or 0))} Vehicles"),
]

doc_xml = replace_literal_paragraphs(doc_xml, _cover_replacements + _kpi_replacements)

# Energy Savings tables' own currency-denominated column headers are baked
# into the template as literal "(EUR)" text, never covered by the field-map
# based table renderer above (which only matches header text to a column
# index via normalize_header(), stripping the currency code rather than
# rewriting it). This template repeats the header row twice (Top-10 table +
# All-Assets appendix table), so both need it and 2 matches is expected —
# warn_ambiguous is off for that reason, not because the count is uncertain.
_currency_header_replacements = [
    ('Energy Cost (EUR)',         f"Energy Cost ({CURRENCY})"),
    ('Energy Cost Savings (EUR)', f"Energy Cost Savings ({CURRENCY})"),
    ('Investment (EUR)',          f"Investment ({CURRENCY})"),
]
doc_xml = replace_literal_paragraphs(doc_xml, _currency_header_replacements, warn_ambiguous=False)

# Payback / IRR are overlaid as a text box on top of a donut-shaped image
# (not a table cell, not a native chart). Scoped to text-box content only -
# a bare "42%" is otherwise too generic to match document-wide once Phase 2
# has written this customer's own per-asset savings-% data. This report's
# donut always shows the Top-10 figures (matching {{IRR_DISPLAY}}'s use of
# TOP10_IRR in the token-based version of this same script).
doc_xml = replace_within_textboxes(doc_xml, [
    ('1.9 yrs',             f"{fmtyrs(TOP10_PAYBACK)} yrs"),
    ('42%',                 fmtirr(TOP10_IRR)),
])

# NPV-positive fraction on the Summary tile ("21/45 motors NPV positive") -
# same mixed value+caption paragraph as generate_report_standard.py; scoped
# to that one paragraph so bare "21"/"45" runs elsewhere are not touched.
_m = re.search(r'<w:p[ >].*?NPV positive.*?</w:p>', doc_xml, re.DOTALL)
if _m:
    _scoped = replace_literal_runs(_m.group(), [
        ('21', str(NPV_POS_CNT)),
        ('45', str(TOTAL_ASSETS)),
    ], warn_ambiguous=False)
    doc_xml = doc_xml[:_m.start()] + _scoped + doc_xml[_m.end():]
else:
    print("  WARNING: 'NPV positive' paragraph not found, motors fraction left unchanged.")

# Top-10 Energy Savings table's own totals row - bare numeric cells, scoped
# to that row so they can't collide with an unrelated asset's own value.
doc_xml = replace_within_row(doc_xml, 'Total - Top 10', [
    ('5,85,356',            fmt(t10_e_cons)),
    ('81,950',              fmt(t10_e_cost)),
    ('87,803',              fmt(t10_co2_c)),
    ('1,89,495',            fmt(t10_sav_kwh)),
    ('26,529',              fmt(t10_sav_c)),
    ('32%',                 fmtpct(t10_pct)),
    ('49,928',              fmt(t10_invest)),
    ('1.9',                 fmtyrs(TOP10_PAYBACK)),
    ('28,424',              fmt(t10_co2_sav)),
])

# NPV-positive-assets totals row - appears twice (once under the Top-10
# table, once under the All-Assets appendix table); replace_within_row
# updates every matching row it finds.
doc_xml = replace_within_row(doc_xml, 'Total NPV Positive Assets (21)', [
    ('Total NPV Positive Assets (21)', f"Total NPV Positive Assets ({NPV_POS_CNT})"),
    ('9,61,687',            fmt(npv_e_cons)),
    ('1,34,636',            fmt(npv_e_cost)),
    ('1,44,253',            fmt(npv_co2_c)),
    ('2,23,067',            fmt(npv_sav_kwh)),
    ('31,229',              fmt(ANNUAL_SAVINGS)),
    ('23%',                 fmtpct(npv_sav_pct)),
    ('89,528',              fmt(INVEST_COST)),
    ('2.9',                 fmtyrs(PAYBACK_TIME)),
    ('33,460',              fmt(npv_co2_sav)),
])

# Footer report date: plain regex match, robust regardless of which sample
# date happens to be baked into the template's footer today.
footer_xmls = [re.sub(r'\d{4}-\d{2}-\d{2}', RPT_DATE_ISO, fx) for fx in footer_xmls]

print("  Scalar substitution complete.")

# ── Write output docx ─────────────────────────────────────────────────────────
safe_customer = re.sub(r'[^\w\- ]', '', CUSTOMER).strip().replace(' ', '_')
safe_plant    = re.sub(r'[^\w\- ]', '', PLANT).strip().replace(' ', '_')
out_name      = f"{safe_customer}_{safe_plant}_EA_Report_Executive.docx"
out_path      = os.path.join(os.path.dirname(os.path.abspath(XLSX_PATH)), out_name)

buf = io.BytesIO()
with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
    for name in names:
        if name == 'word/document.xml':
            zout.writestr(name, doc_xml.encode('utf-8'))
        elif name in footer_names:
            zout.writestr(name, footer_xmls[footer_names.index(name)].encode('utf-8'))
        elif name in binary_overrides:
            zout.writestr(name, binary_overrides[name])
        else:
            zout.writestr(name, zin.read(name))

zin.close()

with open(out_path, 'wb') as f:
    f.write(buf.getvalue())

print(f"\nDone: {out_name}")
print(f"Saved to: {out_path}")
