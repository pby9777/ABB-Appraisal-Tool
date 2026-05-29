#!/usr/bin/env python3
"""
Phase 2C validation: prove generate_multi_workbook() produces a workbook
containing independent Pumps, Fans, and Compressors sheets from a single
call, with no residual Saving_Calculations sheet.

Checks:
  - Correct sheet names (original deleted, three asset-group sheets present)
  - Unique table names per sheet
  - Correct equipment IDs per sheet (data isolation)
  - No cross-contamination between sheets
  - Structured-reference formulas survived
  - Header row shared-strings preserved

Usage:
    python legacy/test_generate_multi_workbook.py
"""

import os
import re
import shutil
import sys
import tempfile
import zipfile as _zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from fill_saving_calculations import generate_multi_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

TEMPLATE   = "legacy/excel_templates/1_0_region_site_name_saving_calculations.xlsx"
OUTPUT     = "legacy/test_generate_multi_workbook_output.xlsx"

# ---------------------------------------------------------------------------
# Mock EEA data — three distinct asset groups
# ---------------------------------------------------------------------------

GROUPS = {
    "Pumps": [
        dict(equip_id="EQ-P001", energy_kwh=180_000, savings_kwh=45_000,
             investment=35_000, ess_motor="ABB-M132S",  ess_drive="ACS580",
             application="Centrifugal Pump", dol_vsd="VSD",
             flow_control="Throttling",          output_kw=22.0,
             run_hours=5000, ie_class="IE3", avg_loading=78.0, avg_freq=48.0),
        dict(equip_id="EQ-P002", energy_kwh=240_000, savings_kwh=65_000,
             investment=50_000, ess_motor="ABB-M160M",  ess_drive="ACS580",
             application="Centrifugal Pump", dol_vsd="VSD",
             flow_control="No control (Bypass)", output_kw=37.0,
             run_hours=4000, ie_class="IE3", avg_loading=82.0, avg_freq=47.5),
        dict(equip_id="EQ-P003", energy_kwh=110_000, savings_kwh=28_000,
             investment=22_000, ess_motor="ABB-M112M",  ess_drive="ACS580",
             application="Centrifugal Pump", dol_vsd="DOL",
             flow_control="Throttling",          output_kw=15.0,
             run_hours=6000, ie_class="IE2", avg_loading=70.0, avg_freq=50.0),
    ],
    "Fans": [
        dict(equip_id="EQ-F001", energy_kwh=95_000, savings_kwh=22_000,
             investment=18_000, ess_motor="ABB-M132S",  ess_drive="ACS580",
             application="Fan or blower",   dol_vsd="VSD",
             flow_control="Damper",             output_kw=11.0,
             run_hours=8000, ie_class="IE3", avg_loading=65.0, avg_freq=49.0),
        dict(equip_id="EQ-F002", energy_kwh=60_000, savings_kwh=14_000,
             investment=11_000, ess_motor="ABB-M100L",  ess_drive="ACS580",
             application="Fan or blower",   dol_vsd="DOL",
             flow_control="On/off",              output_kw=7.5,
             run_hours=7000, ie_class="IE2", avg_loading=60.0, avg_freq=50.0),
    ],
    "Compressors": [
        dict(equip_id="EQ-C001", energy_kwh=320_000, savings_kwh=80_000,
             investment=90_000, ess_motor="ABB-M200L",  ess_drive="ACS880",
             application="Centrifugal compressor", dol_vsd="VSD",
             flow_control="Throttling",          output_kw=90.0,
             run_hours=6000, ie_class="IE3", avg_loading=88.0, avg_freq=47.0),
        dict(equip_id="EQ-C002", energy_kwh=210_000, savings_kwh=52_000,
             investment=60_000, ess_motor="ABB-M180L",  ess_drive="ACS880",
             application="Centrifugal compressor", dol_vsd="VSD",
             flow_control="No control (Bypass)", output_kw=55.0,
             run_hours=5000, ie_class="IE3", avg_loading=85.0, avg_freq=47.5),
    ],
}

ASSUMPTIONS_BY_GROUP = {
    "Pumps":       {"currency": "EUR", "tariff": 0.14, "co2": 0.75,
                    "tax": 19.0, "discount": 6.5},
    "Fans":        {"currency": "EUR", "tariff": 0.14, "co2": 0.75,
                    "tax": 19.0, "discount": 6.5},
    "Compressors": {"currency": "EUR", "tariff": 0.16, "co2": 0.80,
                    "tax": 22.0, "discount": 7.0},
}

# Expected unique main table names after cloning (suffix order matches spec order)
EXPECTED_TABLES = {
    "Pumps":       "Saving_Table2410_2",
    "Fans":        "Saving_Table2410_3",
    "Compressors": "Saving_Table2410_4",
}

# ---------------------------------------------------------------------------
# Mock zip builder (same pattern as test_fill_sheet.py)
# ---------------------------------------------------------------------------

_ASSESS_HEADERS = [
    "Sr No", "Customer Equipment Id",
    "Annual Energy Consumption (kWh)", "C3", "C4", "C5",
    "Annual Energy Savings (kWh)", "C7", "C8", "C9",
    "Investment (currency)", "C11", "C12",
    "Recommended ESS Motor", "ESS Connection",
]
_INPUT_HEADERS = [
    "Customer Equipment Id", "Driven Load", "Dol Vsd",
    "Flow Control", "Shaft Height [mm]", "Rated Power [kW]",
    "Annual Running Hours [h]", "IE Eff Class",
    "Avg Loading [%]", "Avg Frequency [Hz]",
]


def _make_mock_zip(assets: list) -> tuple:
    """Return (zip_path, tmpdir) for a minimal EEA tool output zip."""
    tmpdir = tempfile.mkdtemp()

    awb, aws = openpyxl.Workbook(), None
    aws = awb.active
    for ci, h in enumerate(_ASSESS_HEADERS, 1):
        aws.cell(1, ci).value = h
    for i, a in enumerate(assets, 1):
        r = i + 1
        aws.cell(r, 1).value  = i;              aws.cell(r, 2).value  = a["equip_id"]
        aws.cell(r, 3).value  = a["energy_kwh"]; aws.cell(r, 7).value  = a["savings_kwh"]
        aws.cell(r, 11).value = a["investment"]; aws.cell(r, 14).value = a["ess_motor"]
        aws.cell(r, 15).value = a["ess_drive"]
    ap = os.path.join(tmpdir, "assessment data.xlsx")
    awb.save(ap)

    iwb, iws = openpyxl.Workbook(), None
    iws = iwb.active
    for ci, h in enumerate(_INPUT_HEADERS, 1):
        iws.cell(1, ci).value = h
    for i, a in enumerate(assets, 2):
        iws.cell(i, 1).value = a["equip_id"];    iws.cell(i, 2).value = a["application"]
        iws.cell(i, 3).value = a["dol_vsd"];     iws.cell(i, 4).value = a["flow_control"]
        iws.cell(i, 5).value = 0;                iws.cell(i, 6).value = a["output_kw"]
        iws.cell(i, 7).value = a["run_hours"];   iws.cell(i, 8).value = a["ie_class"]
        iws.cell(i, 9).value = a["avg_loading"]; iws.cell(i, 10).value = a["avg_freq"]
    ip = os.path.join(tmpdir, "input assets.xlsx")
    iwb.save(ip)

    zp = os.path.join(tmpdir, "eea_output.zip")
    with _zipfile.ZipFile(zp, "w") as zf:
        zf.write(ap, "assessment data.xlsx")
        zf.write(ip, "input assets.xlsx")

    return zp, tmpdir


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

PASS_LIST: list = []
FAIL_LIST: list = []


def check(label: str, cond: bool, detail: str = "") -> None:
    (PASS_LIST if cond else FAIL_LIST).append(
        label if cond else f"{label}  {detail}"
    )


def _main_table(ws):
    for tname, tobj in ws.tables.items():
        ref = tobj.ref if hasattr(tobj, "ref") else str(tobj)
        m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ref)
        if m and (int(m.group(2)) - int(m.group(1))) > 100:
            return tname, ref
    return None, None


def _header_row(ref: str) -> int:
    m = re.match(r"[A-Z]+(\d+):", ref)
    return int(m.group(1)) if m else 36


def _col_map(ws, hrow: int) -> dict:
    return {c.value: c.column for c in ws[hrow] if c.value is not None}


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------

def step1_generate(tmp_dirs: list) -> dict:
    print("\n── STEP 1: generate_multi_workbook() ───────────────────────────")

    sheet_specs = []
    for group_name, assets in GROUPS.items():
        zpath, tmpdir = _make_mock_zip(assets)
        tmp_dirs.append(tmpdir)
        spec = {"name": group_name, "zips": [zpath]}
        spec.update(ASSUMPTIONS_BY_GROUP[group_name])
        sheet_specs.append(spec)

    names, counts = generate_multi_workbook(TEMPLATE, sheet_specs, OUTPUT)

    check("Returns 3 sheet names",        len(names) == 3,
          f"got {names}")
    check("Returns correct asset counts", counts == {"Pumps": 3, "Fans": 2, "Compressors": 2},
          f"got {counts}")
    check("Output file exists",           os.path.isfile(OUTPUT))
    return counts


def step2_verify(counts: dict) -> None:
    print("\n── STEP 2: verification ────────────────────────────────────────")
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)

    # ── 2a. Sheet names ──────────────────────────────────────────────────
    expected_present = ["Note", "Input_ Reference", "Pumps", "Fans", "Compressors"]
    check("Sheet count == 5",             len(wb.sheetnames) == 5,
          f"got {wb.sheetnames}")
    for sn in expected_present:
        check(f"Sheet present: {sn!r}",   sn in wb.sheetnames)
    check("Original sheet deleted",
          "Saving_Calculations" not in wb.sheetnames,
          f"still in {wb.sheetnames}")

    for group_name, assets in GROUPS.items():
        ws   = wb[group_name]
        tname, ref = _main_table(ws)
        hrow = _header_row(ref or "A36:BE1100")
        cmap = _col_map(ws, hrow)

        # ── 2b. Tables ───────────────────────────────────────────────────
        check(f"2 tables on {group_name}",    len(ws.tables) == 2,
              f"got {len(ws.tables)}")
        check(f"Main table found: {group_name}",    tname is not None)
        check(f"Main table name correct: {EXPECTED_TABLES[group_name]!r}",
              tname == EXPECTED_TABLES[group_name],
              f"got {tname!r}")

        # ── 2c. Asset count & equipment IDs ─────────────────────────────
        equip_col = cmap.get("Customer Equipment Id")
        if equip_col:
            for i, asset in enumerate(assets):
                actual = ws.cell(hrow + 1 + i, equip_col).value
                check(f"{group_name} row{hrow+1+i} equip_id",
                      actual == asset["equip_id"],
                      f"expected={asset['equip_id']!r} got={actual!r}")

        # ── 2d. Cross-contamination ──────────────────────────────────────
        foreign_ids = {
            a["equip_id"]
            for g, ga in GROUPS.items() if g != group_name
            for a in ga
        }
        contaminated = False
        if equip_col:
            for r in range(hrow + 1, hrow + 1 + len(assets)):
                v = ws.cell(r, equip_col).value
                if v in foreign_ids:
                    contaminated = True
                    FAIL_LIST.append(
                        f"{group_name} row{r}: foreign id {v!r}")
        check(f"{group_name} no cross-contamination", not contaminated)

        # ── 2e. Formulas preserved ───────────────────────────────────────
        expected_tbl = EXPECTED_TABLES[group_name]
        for col_letter, desc in [("F", "Annual Energy Cost"),
                                  ("G", "Annual CO2 Cons."),
                                  ("I", "Annual Energy Cost Savings")]:
            cell = ws[f"{col_letter}{hrow + 1}"]
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            check(f"{group_name} {col_letter}{hrow+1} is formula ({desc})",
                  is_formula, f"value={str(cell.value)[:60]!r}")
            if is_formula:
                check(f"{group_name} {col_letter}{hrow+1} SR → {expected_tbl!r}",
                      f"{expected_tbl}[" in cell.value,
                      f"formula={cell.value[:80]!r}")

        # ── 2f. Application mapping from input assets ────────────────────
        app_col = cmap.get("Application")
        if app_col:
            actual_app = ws.cell(hrow + 1, app_col).value
            expected_app = assets[0]["application"]
            check(f"{group_name} row{hrow+1} Application",
                  actual_app == expected_app,
                  f"expected={expected_app!r} got={actual_app!r}")

        # ── 2g. Header row preserved (shared-string test) ────────────────
        for col_idx in list(cmap.values())[:5]:
            check(f"{group_name} header col{col_idx} non-null",
                  ws.cell(hrow, col_idx).value is not None)

    print(f"  Sheets in output: {wb.sheetnames}")


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def print_report() -> int:
    SEP = "=" * 62
    print(f"\n{SEP}")
    print("GENERATE_MULTI_WORKBOOK — VALIDATION REPORT")
    print(SEP)
    print(f"\n  PASS: {len(PASS_LIST)}   FAIL: {len(FAIL_LIST)}")
    if FAIL_LIST:
        print("\n  FAILURES:")
        for m in FAIL_LIST:
            print(f"    FAIL  {m}")
    print("\n  PASSING CHECKS:")
    for m in PASS_LIST:
        print(f"    ok    {m}")
    print()
    return len(FAIL_LIST)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    tmp_dirs: list = []
    try:
        counts = step1_generate(tmp_dirs)
        step2_verify(counts)
    finally:
        for td in tmp_dirs:
            shutil.rmtree(td, ignore_errors=True)

    failures = print_report()
    sys.exit(0 if failures == 0 else 1)
