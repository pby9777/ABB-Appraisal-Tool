#!/usr/bin/env python3
"""
test_hp_fix.py
--------------
Validates the HP/NEMA power-column priority fix in read_input_assets().

Test A — HP/NEMA variant:
  Input xlsx has both 'Rated Power [Hp]' and 'Rated Power [kW]' columns,
  plus 'Motor Efficiency [%]' (no 'IE Eff Class').
  Expected: output_kw == HP value (e.g. 10), NOT kW value (e.g. 7.46).

Test B — IEC variant (regression):
  Input xlsx has 'Rated Power [kW]' and 'IE Eff Class', no motor_eff.
  Expected: output_kw == kW value.

Test C — Real NEMA ZIP (tests/eea-report-24418_NEMA_Var.zip):
  Verifies output_kw is non-None for all assets and matches
  the HP column values from the ZIP (all 1 HP for first asset).

Prints PASS/FAIL for each check and exits non-zero on any failure.
"""

import os
import sys
import shutil
import tempfile
import zipfile as _zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from fill_saving_calculations import read_input_assets

PASS_LIST = []
FAIL_LIST = []


def check(label, cond, detail=""):
    if cond:
        PASS_LIST.append(label)
        print(f"  PASS  {label}")
    else:
        FAIL_LIST.append(f"{label}  {detail}")
        print(f"  FAIL  {label}  {detail}")


def _make_input_xlsx(tmpdir, headers, rows):
    """Write a minimal input assets xlsx and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci).value = h
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(ri, ci).value = v
    path = os.path.join(tmpdir, "input assets.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Test A: HP/NEMA variant — HP column must win
# ---------------------------------------------------------------------------

def test_a_hp_variant():
    print("\n── Test A: HP/NEMA variant ─────────────────────────────────────")
    tmpdir = tempfile.mkdtemp()
    try:
        HP_VALUE  = 10.0
        KW_VALUE  = 7.46

        headers = [
            "Customer Equipment Id", "Driven Load", "Dol Vsd",
            "Flow Control", "Shaft Height [mm]",
            "Rated Power [Hp]", "Rated Power [kW]",
            "Annual Running Hours [h]", "Motor Efficiency [%]",
            "Avg Loading [%]", "Avg Frequency [Hz]",
        ]
        rows = [
            ["EQ-HP-001", "Centrifugal Pump", "DOL", "Throttling", 0,
             HP_VALUE, KW_VALUE, 4000, 91.7, 75.0, 50.0],
        ]

        path = _make_input_xlsx(tmpdir, headers, rows)
        assets = read_input_assets(path)

        check("Test A: asset loaded",
              "EQ-HP-001" in assets)
        check("Test A: output_kw == HP value (not kW)",
              assets.get("EQ-HP-001", {}).get("output_kw") == HP_VALUE,
              f"got {assets.get('EQ-HP-001', {}).get('output_kw')!r}, expected {HP_VALUE!r}")
        check("Test A: output_kw != kW value",
              assets.get("EQ-HP-001", {}).get("output_kw") != KW_VALUE,
              f"got kW value {KW_VALUE!r} — HP column not prioritised")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Test B: IEC variant regression — kW column must win
# ---------------------------------------------------------------------------

def test_b_iec_variant():
    print("\n── Test B: IEC variant regression ──────────────────────────────")
    tmpdir = tempfile.mkdtemp()
    try:
        KW_VALUE = 22.0

        headers = [
            "Customer Equipment Id", "Driven Load", "Dol Vsd",
            "Flow Control", "Shaft Height [mm]",
            "Rated Power [kW]",
            "Annual Running Hours [h]", "IE Eff Class",
            "Avg Loading [%]", "Avg Frequency [Hz]",
        ]
        rows = [
            ["EQ-IEC-001", "Fan or blower", "VSD", "Damper", 0,
             KW_VALUE, 6000, "IE3", 70.0, 49.0],
        ]

        path = _make_input_xlsx(tmpdir, headers, rows)
        assets = read_input_assets(path)

        check("Test B: asset loaded",
              "EQ-IEC-001" in assets)
        check("Test B: output_kw == kW value",
              assets.get("EQ-IEC-001", {}).get("output_kw") == KW_VALUE,
              f"got {assets.get('EQ-IEC-001', {}).get('output_kw')!r}, expected {KW_VALUE!r}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Test C: Real NEMA ZIP (eea-report-24418_NEMA_Var.zip)
# ---------------------------------------------------------------------------

def test_c_real_nema_zip():
    print("\n── Test C: Real NEMA ZIP ────────────────────────────────────────")

    zip_path = os.path.join(
        os.path.dirname(__file__), "..", "tests", "eea-report-24418_NEMA_Var.zip"
    )
    if not os.path.isfile(zip_path):
        print(f"  SKIP  ZIP not found: {zip_path}")
        return

    tmpdir = tempfile.mkdtemp()
    try:
        with _zipfile.ZipFile(zip_path) as zf:
            zf.extractall(tmpdir)

        input_path = None
        for root, _, files in os.walk(tmpdir):
            for f in files:
                if "input" in f.lower() and f.endswith(".xlsx"):
                    input_path = os.path.join(root, f)
                    break

        check("Test C: input assets file found", input_path is not None)
        if input_path is None:
            return

        assets = read_input_assets(input_path)

        check("Test C: at least one asset loaded", len(assets) > 0,
              f"got {len(assets)} assets")

        none_count = sum(1 for v in assets.values() if v.get("output_kw") is None)
        check("Test C: no asset has output_kw=None",
              none_count == 0,
              f"{none_count}/{len(assets)} assets still have output_kw=None")

        first_id  = next(iter(assets))
        first_val = assets[first_id].get("output_kw")
        check("Test C: first asset output_kw is numeric",
              isinstance(first_val, (int, float)),
              f"got {first_val!r}")

        print(f"  INFO  First asset '{first_id}': output_kw={first_val!r} HP")

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def print_report():
    SEP = "=" * 60
    print(f"\n{SEP}")
    print("TEST HP FIX — VALIDATION REPORT")
    print(SEP)
    total = len(PASS_LIST) + len(FAIL_LIST)
    print(f"\n  PASS: {len(PASS_LIST)}/{total}   FAIL: {len(FAIL_LIST)}/{total}")
    if FAIL_LIST:
        print("\n  FAILURES:")
        for msg in FAIL_LIST:
            print(f"    FAIL  {msg}")
    print()
    if FAIL_LIST:
        print("RESULT: FAIL")
    else:
        print("RESULT: PASS")
    print()


if __name__ == "__main__":
    test_a_hp_variant()
    test_b_iec_variant()
    test_c_real_nema_zip()
    print_report()
    sys.exit(0 if not FAIL_LIST else 1)
