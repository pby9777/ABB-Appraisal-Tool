#!/usr/bin/env python3
"""
Phase 2A: Verify ZIP-level cloned workbooks can be populated by openpyxl.

Creates a 3-sheet workbook (Saving_Calculations, Saving_Calculations_2,
Saving_Calculations_3), writes sample motor-appraisal rows into each sheet,
then verifies that tables, structured-reference formulas, header shared-strings,
and written values all survive the openpyxl round-trip.

Usage:
    python legacy/test_multi_fill.py
"""

import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

# Resolve imports from the same legacy/ directory
sys.path.insert(0, os.path.dirname(__file__))
from excel_clone_poc import clone_saving_calculations

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

TEMPLATE    = "legacy/excel_templates/1_0_region_site_name_saving_calculations.xlsx"
STEP1_PATH  = "/tmp/test_multi_fill_step1.xlsx"   # original + _2
FILLED_PATH = "legacy/test_multi_fill_output.xlsx" # original + _2 + _3, then filled

# ---------------------------------------------------------------------------
# Target columns (exact names as they appear in the table header row 36)
# ---------------------------------------------------------------------------

WRITE_COLS = [
    "Customer Equipment Id",
    "Application",
    "Dol Vsd",
    "Flow Control Method",
    "Output (kW/HP)",
]

# Dropdown-valid values confirmed against the template DV rules
SAMPLE_ROWS = {
    "Saving_Calculations": [
        {
            "Customer Equipment Id": "EQ-A001",
            "Application":          "Centrifugal Pump",
            "Dol Vsd":              "VSD",
            "Flow Control Method":  "Throttling",
            "Output (kW/HP)":       22.0,
        },
        {
            "Customer Equipment Id": "EQ-A002",
            "Application":          "Fan or blower",
            "Dol Vsd":              "DOL",
            "Flow Control Method":  "On/off",
            "Output (kW/HP)":       15.0,
        },
        {
            "Customer Equipment Id": "EQ-A003",
            "Application":          "Centrifugal Pump",
            "Dol Vsd":              "VSD",
            "Flow Control Method":  "No control (Bypass)",
            "Output (kW/HP)":       37.0,
        },
    ],
    "Saving_Calculations_2": [
        {
            "Customer Equipment Id": "EQ-B001",
            "Application":          "Fan or blower",
            "Dol Vsd":              "VSD",
            "Flow Control Method":  "Damper",
            "Output (kW/HP)":       11.0,
        },
        {
            "Customer Equipment Id": "EQ-B002",
            "Application":          "Centrifugal Pump",
            "Dol Vsd":              "DOL",
            "Flow Control Method":  "Inlet vanes",
            "Output (kW/HP)":       7.5,
        },
        {
            "Customer Equipment Id": "EQ-B003",
            "Application":          "Screw chiller",
            "Dol Vsd":              "VSD",
            "Flow Control Method":  "Throttling",
            "Output (kW/HP)":       55.0,
        },
    ],
    "Saving_Calculations_3": [
        {
            "Customer Equipment Id": "EQ-C001",
            "Application":          "Mixer",
            "Dol Vsd":              "DOL",
            "Flow Control Method":  "Information not available",
            "Output (kW/HP)":       4.0,
        },
        {
            "Customer Equipment Id": "EQ-C002",
            "Application":          "Centrifugal compressor",
            "Dol Vsd":              "VSD",
            "Flow Control Method":  "Throttling",
            "Output (kW/HP)":       90.0,
        },
        {
            "Customer Equipment Id": "EQ-C003",
            "Application":          "Conveyor",
            "Dol Vsd":              "DOL",
            "Flow Control Method":  "On/off",
            "Output (kW/HP)":       18.5,
        },
    ],
}

# Expected main table names per sheet (verifies SR rename chain)
EXPECTED_MAIN_TABLES = {
    "Saving_Calculations":   "Saving_Table2410",
    "Saving_Calculations_2": "Saving_Table2410_2",
    "Saving_Calculations_3": "Saving_Table2410_3",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_col_map(ws, header_row: int) -> dict:
    """Return {column_name: column_index} for every non-empty cell in header_row."""
    return {
        cell.value: cell.column
        for cell in ws[header_row]
        if cell.value is not None
    }


def _find_main_table(ws):
    """
    Return (table_name, table_object) for the main data table on the sheet.
    Main table = the one with >100 data rows (the summary has only 20).
    """
    for tname, tobj in ws.tables.items():
        ref = tobj.ref if hasattr(tobj, "ref") else str(tobj)
        m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ref)
        if m and (int(m.group(2)) - int(m.group(1))) > 100:
            return tname, tobj
    return None, None


def _header_row_from_ref(ref: str) -> int:
    """Parse 'A36:BE1100' → 36."""
    m = re.match(r"[A-Z]+(\d+):", ref)
    return int(m.group(1)) if m else 36


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

PASS_LIST: list = []
FAIL_LIST: list = []


def check(label: str, condition: bool, detail: str = "") -> None:
    if condition:
        PASS_LIST.append(label)
    else:
        FAIL_LIST.append(f"{label}  {detail}")


# ---------------------------------------------------------------------------
# Step 1 — build 3-sheet workbook via ZIP-level cloning
# ---------------------------------------------------------------------------

def step1_clone() -> None:
    print("\n── STEP 1: ZIP-level clone ─────────────────────────────────────")

    d1 = clone_saving_calculations(
        TEMPLATE,
        STEP1_PATH,
        clone_display_name="Saving_Calculations_2",
        clone_suffix="_2",
    )
    print(f"  Clone _2: {d1['new_main_table_name']!r}  SR {d1['sr_in_source']}→{d1['sr_in_clone']}")
    check("Clone _2 SR count matches",  d1["sr_in_source"] == d1["sr_in_clone"])
    check("Clone _2 old name absent",   d1["sr_old_in_clone"] == 0)

    d2 = clone_saving_calculations(
        STEP1_PATH,
        FILLED_PATH,
        clone_display_name="Saving_Calculations_3",
        clone_suffix="_3",
    )
    print(f"  Clone _3: {d2['new_main_table_name']!r}  SR {d2['sr_in_source']}→{d2['sr_in_clone']}")
    check("Clone _3 SR count matches",  d2["sr_in_source"] == d2["sr_in_clone"])
    check("Clone _3 old name absent",   d2["sr_old_in_clone"] == 0)
    check("calcChain deleted (step1)",  d1["calc_chain_present"])
    # Step2 runs on step1's output, which already had calcChain deleted.
    # The meaningful check is that the final output has no calcChain at all.
    import zipfile as _zf
    with _zf.ZipFile(FILLED_PATH) as _z:
        check("calcChain absent from final output",
              "xl/calcChain.xml" not in _z.namelist())


# ---------------------------------------------------------------------------
# Step 2 — open with openpyxl and write sample rows
# ---------------------------------------------------------------------------

def step2_fill() -> dict:
    """Write sample data to all 3 Saving_Calculations sheets. Returns col_map."""
    print("\n── STEP 2: openpyxl fill ───────────────────────────────────────")

    wb = openpyxl.load_workbook(FILLED_PATH, data_only=False)

    # Capture header values BEFORE any writes (shared-string baseline)
    baselines = {}
    for sname in EXPECTED_MAIN_TABLES:
        ws = wb[sname]
        tname, tobj = _find_main_table(ws)
        ref = tobj.ref if tobj and hasattr(tobj, "ref") else "A36:BE1100"
        hrow = _header_row_from_ref(ref)
        col_map = _build_col_map(ws, hrow)
        # Capture first 5 header strings as shared-string baseline
        baselines[sname] = {
            col: ws.cell(hrow, col).value
            for col in list(col_map.values())[:5]
        }

    # Write sample rows
    col_maps_captured = {}
    for sname, rows in SAMPLE_ROWS.items():
        ws = wb[sname]
        tname, tobj = _find_main_table(ws)
        ref = tobj.ref if tobj and hasattr(tobj, "ref") else "A36:BE1100"
        hrow = _header_row_from_ref(ref)
        col_map = _build_col_map(ws, hrow)
        col_maps_captured[sname] = col_map

        for offset, row_data in enumerate(rows):
            data_row = hrow + 1 + offset          # 37, 38, 39
            for col_name, value in row_data.items():
                col_idx = col_map.get(col_name)
                if col_idx is None:
                    FAIL_LIST.append(f"Column not found: {col_name!r} on {sname}")
                    continue
                ws.cell(data_row, col_idx).value = value

        print(f"  Wrote {len(rows)} rows to {sname!r}  "
              f"(table={tname!r}, data starts row {hrow+1})")
        check(f"Table found on {sname}", tname is not None)
        check(
            f"Table name correct on {sname}",
            tname == EXPECTED_MAIN_TABLES[sname],
            f"expected={EXPECTED_MAIN_TABLES[sname]!r} got={tname!r}",
        )

    wb.save(FILLED_PATH)
    print(f"  Saved → {FILLED_PATH}")
    return col_maps_captured, baselines


# ---------------------------------------------------------------------------
# Step 3 — reopen and verify
# ---------------------------------------------------------------------------

def step3_verify(col_maps: dict, baselines: dict) -> None:
    print("\n── STEP 3: verification ────────────────────────────────────────")

    wb = openpyxl.load_workbook(FILLED_PATH, data_only=False)

    # ── 3a. Sheet count ──────────────────────────────────────────────────────
    expected_sheets = [
        "Note", "Input_ Reference",
        "Saving_Calculations", "Saving_Calculations_2", "Saving_Calculations_3",
    ]
    check(
        "Sheet count == 5",
        len(wb.sheetnames) == 5,
        f"got={wb.sheetnames}",
    )
    for sn in expected_sheets:
        check(f"Sheet exists: {sn!r}", sn in wb.sheetnames)

    for sname in EXPECTED_MAIN_TABLES:
        ws = wb[sname]
        col_map = col_maps[sname]
        tname, tobj = _find_main_table(ws)
        ref = tobj.ref if tobj and hasattr(tobj, "ref") else "A36:BE1100"
        hrow = _header_row_from_ref(ref)

        # ── 3b. Tables present ───────────────────────────────────────────────
        table_count = len(ws.tables)
        check(
            f"2 tables on {sname}",
            table_count == 2,
            f"got {table_count}",
        )
        check(
            f"Main table name correct: {EXPECTED_MAIN_TABLES[sname]!r}",
            tname == EXPECTED_MAIN_TABLES[sname],
        )

        # ── 3c. Written values preserved ─────────────────────────────────────
        rows = SAMPLE_ROWS[sname]
        for offset, row_data in enumerate(rows):
            data_row = hrow + 1 + offset
            for col_name, expected in row_data.items():
                col_idx = col_map.get(col_name)
                if col_idx is None:
                    continue
                actual = ws.cell(data_row, col_idx).value
                check(
                    f"{sname} row{data_row} {col_name!r}",
                    actual == expected,
                    f"expected={expected!r} got={actual!r}",
                )

        # ── 3d. Formulas still present in calculated columns ─────────────────
        # F = Annual Energy Cost (formula), G = Annual CO2 Cons. (formula)
        # I = Annual Energy Cost Savings,   J = Annual Energy Savings %
        formula_cols = [
            ("F", "Annual Energy Cost"),
            ("G", "Annual CO2 Cons."),
            ("I", "Annual Energy Cost Savings"),
            ("J", "Annual Energy Savings (%)"),
        ]
        expected_table = EXPECTED_MAIN_TABLES[sname]
        for col_letter, col_desc in formula_cols:
            cell_addr = f"{col_letter}{hrow + 1}"
            cell = ws[cell_addr]
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            check(
                f"{sname} {cell_addr} is formula ({col_desc})",
                is_formula,
                f"value={str(cell.value)[:60]!r}",
            )
            if is_formula:
                # Verify the SR uses the correct renamed table
                correct_sr = f"{expected_table}[" in cell.value
                check(
                    f"{sname} {cell_addr} SR references {expected_table!r}",
                    correct_sr,
                    f"formula={cell.value[:80]!r}",
                )

        # ── 3e. Shared-string header row preserved ────────────────────────────
        baseline = baselines[sname]
        for col_idx, expected_str in baseline.items():
            actual = ws.cell(hrow, col_idx).value
            check(
                f"{sname} header col{col_idx} preserved",
                actual == expected_str,
                f"expected={expected_str!r} got={actual!r}",
            )

    print(f"\n  Sheets in output: {wb.sheetnames}")


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def print_report() -> int:
    SEP = "=" * 62
    print(f"\n{SEP}")
    print("PHASE 2A VALIDATION REPORT")
    print(SEP)
    print(f"\n  PASS: {len(PASS_LIST)}   FAIL: {len(FAIL_LIST)}")

    if FAIL_LIST:
        print("\n  FAILURES:")
        for msg in FAIL_LIST:
            print(f"    FAIL  {msg}")

    print("\n  PASSING CHECKS:")
    for msg in PASS_LIST:
        print(f"    ok    {msg}")

    print()
    return len(FAIL_LIST)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    step1_clone()
    col_maps, baselines = step2_fill()
    step3_verify(col_maps, baselines)
    failures = print_report()
    sys.exit(0 if failures == 0 else 1)
